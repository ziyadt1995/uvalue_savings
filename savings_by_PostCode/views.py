from django.shortcuts import render, redirect
from django.http import HttpResponse
from django.contrib.auth.decorators import login_required
from django.views.decorators.csrf import csrf_exempt
from openpyxl.writer.excel import save_virtual_workbook
from rest_framework.views import APIView
from rest_framework.response import Response
import json
from django.http import JsonResponse
from rest_framework_simplejwt.authentication import JWTAuthentication
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from rest_framework.views import APIView
import base64
import requests
import string
import random # define the random module
#from backend.config import Config
# -*- coding: utf-8 -*-
from .forms import PCForm
import sys
import json



import requests
import numpy as np
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
import math




import requests
import numpy as np
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
import math

import urllib.request
from urllib.parse import urlencode
import re


token = "eml5YWR0MTk5NUBnbWFpbC5jb206MDkyODI2M2QxNTNhZDEzMDg0MDhhYjNlYjFhMTE1MzI0MDYzMGM2Nw=="

headers = {
    'Accept': 'application/json',
    'Authorization': f'Basic {token}'
}
def removecharacters(x):
    x=str(x)
    refinedx=re.sub('\W+','', x)
    return refinedx



def home(request):
    form = PCForm()
    return render(request, 'homepage.html', {"form": form})



class savingslistpost(APIView):
    def post(self, request, format=None):
        postcode=str(request.POST.get('postcode'))
        #print(request.body)
        #data=json.loads(request.body)
        #postcode=str(data['postcode'])
        if(len(postcode)>=4 and len(postcode)<=7):

            #ps=data['postcode']
            #ps=str(ps)


            # Define base URL and query parameters separately
            base_url = 'https://epc.opendatacommunities.org/api/v1/non-domestic/search'
            query_params = {"postcode":postcode, 'size': 5000}



            # Encode query parameters
            encoded_params = urlencode(query_params)

            # Append parameters to the base URL
            full_url = f"{base_url}?{encoded_params}"


            base_url_domestic='https://epc.opendatacommunities.org/api/v1/domestic/search'


            # Encode query parameters
            encoded_params = urlencode(query_params)

            # Append parameters to the base URL
            full_url_domestic = f"{base_url_domestic}?{encoded_params}"



            uvalue_2010={"window_L":1.6, "window_leti": 1.0, "door_L":1.8, "door_leti":1.3, "roof_L":0.18, "roof_leti":0.12, "wall_L":0.28, "wall_leti":0.24, "floor_L": 0.22, "floor_leti":0.20}

            uvalue_2011={"window_L":1.8, "window_leti": 1.3, "door_L":1.6, "door_leti":1.0, "roof_L":0.18, "roof_leti":0.12, "wall_L":0.28, "wall_leti":0.24, "floor_L": 0.22, "floor_leti":0.20}

            uvalue_2013={"window_L":2.0, "window_leti": 1.3, "door_L":2.0, "door_leti":1.0, "roof_L":0.20, "roof_leti":0.12, "wall_L":0.3, "wall_leti":0.18, "floor_L": 0.25, "floor_leti":0.20}

            uvalue_2018={"window_L":1.6, "window_leti": 1.3, "door_L":1.8, "door_leti":1.0, "roof_L":0.18, "roof_leti":0.12, "wall_L":0.28, "wall_leti":0.18, "floor_L": 0.22, "floor_leti":0.15}

            uvalue_2019={"window_L":1.6, "window_leti": 1.3, "door_L":1.8, "door_leti":1.0, "roof_L":0.18, "roof_leti":0.12, "wall_L":0.28, "wall_leti":0.18, "floor_L": 0.18, "floor_leti":0.15}

            uvalue_2021={"window_L":1.6, "window_leti": 1.3, "door_L":1.6, "door_leti":1.0, "roof_L":0.16, "roof_leti":0.12, "wall_L":0.30, "wall_leti":0.18, "floor_L": 0.18, "floor_leti":0.15}

            wb=Workbook()

            #ws.title="uvalue_"+postcode["postcode"]
            #print(ws.title)

            wb.create_sheet("FinalTables")
            finaltable=wb["FinalTables"]

            wb.create_sheet("UserInputs")
            userinputs=wb["UserInputs"]

            char=get_column_letter(1)
            userinputs[char+str(1)]="Property Address"
            char=get_column_letter(2)
            userinputs[char+str(1)]="Elec Monthly Bill"
            char=get_column_letter(3)
            userinputs[char+str(1)]="Gas Monthly Bill"
            char=get_column_letter(4)
            userinputs[char+str(1)]="Years on lease"


            wb.create_sheet("Portfolio")
            portfolio=wb["Portfolio"]


            char=get_column_letter(2)
            portfolio[char+str(1)]="Property Type"
            char=get_column_letter(3)
            portfolio[char+str(1)]="Address 1"
            char=get_column_letter(4)
            portfolio[char+str(1)]="Full Address"
            char=get_column_letter(5)
            portfolio[char+str(1)]="Postcode"
            char=get_column_letter(6)
            portfolio[char+str(1)]="Inspection Date"
            char=get_column_letter(7)
            portfolio[char+str(1)]="Floor Area"
            char=get_column_letter(8)
            portfolio[char+str(1)]="UPRN"
            char=get_column_letter(9)
            portfolio[char+str(1)]="LMK KEY"
            char=get_column_letter(10)
            portfolio[char+str(1)]="Door U_value Difference"
            char=get_column_letter(11)
            portfolio[char+str(1)]="Window U_value Difference"
            char=get_column_letter(12)
            portfolio[char+str(1)]="Roof U_value Difference"
            char=get_column_letter(13)
            portfolio[char+str(1)]="Wall U_value Difference"
            char=get_column_letter(14)
            portfolio[char+str(1)]="Floor U_value Difference"
            char=get_column_letter(15)
            portfolio[char+str(1)]="Door Area"
            char=get_column_letter(16)
            portfolio[char+str(1)]="Window Area"
            char=get_column_letter(17)
            portfolio[char+str(1)]="Roof Area"
            char=get_column_letter(18)
            portfolio[char+str(1)]="Wall Area"
            char=get_column_letter(19)
            portfolio[char+str(1)]="Floor Area"
            char=get_column_letter(20)
            portfolio[char+str(1)]="Door Improvement (Area)"
            char=get_column_letter(21)
            portfolio[char+str(1)]="Window Improvement (Area)"
            char=get_column_letter(22)
            portfolio[char+str(1)]="Roof Improvement (Area)"
            char=get_column_letter(23)
            portfolio[char+str(1)]="Wall Improvement (Area)"
            char=get_column_letter(24)
            portfolio[char+str(1)]="Floor Improvement (Area)"
            char=get_column_letter(25)
            portfolio[char+str(1)]="TotalImprovement (Area)"

            char=get_column_letter(26)
            portfolio[char+str(1)]="Door Savings"
            char=get_column_letter(27)
            portfolio[char+str(1)]="Window Savings"
            char=get_column_letter(28)
            portfolio[char+str(1)]="Roof Savings"
            char=get_column_letter(29)
            portfolio[char+str(1)]="Wall Savings"
            char=get_column_letter(30)
            portfolio[char+str(1)]="Floor Savings"
            char=get_column_letter(31)
            portfolio[char+str(1)]="Total Savings (%)"
            char=get_column_letter(32)
            portfolio[char+str(1)]="Final Savings (GBP)"


            char=get_column_letter(33)
            portfolio[char+str(1)]="Annual Saving Passed to Landlord"
            char=get_column_letter(34)
            portfolio[char+str(1)]="Monthly tenant Fee"
            char=get_column_letter(35)
            portfolio[char+str(1)]="Referral Fee"
            char=get_column_letter(36)
            portfolio[char+str(1)]="Cut of Tenant Fee"
            char=get_column_letter(37)
            portfolio[char+str(1)]="UpGreen Annual Income"
            char=get_column_letter(38)
            portfolio[char+str(1)]="Annual Payback to Landlord"
            char=get_column_letter(39)
            portfolio[char+str(1)]="Annual Savings Passed to Tenants"
            char=get_column_letter(40)
            portfolio[char+str(1)]="Value from fees passed back to the landlord for the remainder of the lease"
            char=get_column_letter(41)
            portfolio[char+str(1)]="Total cost of Retrofit per square meter for constituents"

            char=get_column_letter(51)
            portfolio[char+str(1)]="Building Emission Savings (kg CO2/m^2)"
            char=get_column_letter(52)
            portfolio[char+str(1)]="Door Energy Savings (GBP)"
            char=get_column_letter(53)
            portfolio[char+str(1)]="window Energy Savings (GBP)"
            char=get_column_letter(54)
            portfolio[char+str(1)]="Roof Energy Savings (GBP)"
            char=get_column_letter(55)
            portfolio[char+str(1)]="Wall Energy Savings (GBP)"
            char=get_column_letter(56)
            portfolio[char+str(1)]="Floor Energy Savings (GBP)"
            char=get_column_letter(57)
            portfolio[char+str(1)]="Full Energy Savings (kWh)"
            char=get_column_letter(58)
            portfolio[char+str(1)]="High ROI Energy Saving (kWh)"
            char=get_column_letter(59)
            portfolio[char+str(1)]="High ROI Building Emission Saving (kg co2/m2)"
            char=get_column_letter(60)
            portfolio[char+str(1)]="High ROI Energy Saving (GBP)"
            char=get_column_letter(61)
            portfolio[char+str(1)]="U_value L-Doc Based on Area"





            #COST OF RETROFITTING PER METER SQUARED
            cost_of_retrofit_constituent_per_m2_windowlow=30
            cost_of_retrofit_constituent_per_m2_doorlow=150
            cost_of_retrofit_constituent_per_m2_walllow=50
            cost_of_retrofit_constituent_per_m2_rooflow=50
            cost_of_retrofit_constituent_per_m2_floorlow=20


            cost_of_retrofit_constituent_per_m2_window=45
            cost_of_retrofit_constituent_per_m2_door=250
            cost_of_retrofit_constituent_per_m2_wall=100
            cost_of_retrofit_constituent_per_m2_roof=100
            cost_of_retrofit_constituent_per_m2_floor=40


            cost_of_retrofit_constituent_per_m2_windowhigh=60
            cost_of_retrofit_constituent_per_m2_doorhigh=350
            cost_of_retrofit_constituent_per_m2_wallhigh=150
            cost_of_retrofit_constituent_per_m2_roofhigh=300
            cost_of_retrofit_constituent_per_m2_floorhigh=80






            #THIS IS THE CODE THAT AUTOMATES THE FINAL SAVINGS AND CUTS BASED ON PUJAS LOOM VIDEO

            origination_fees=0
            installer_commission=0.05
            cut_of_savings=0.015
            cost_of_retrofitting=3000
            interest_rate=0
            loan_term=0
            years_left_on_current_lease=3
            savings_to_tenant=0.10
            discount_rate=0.05
            cashback=0
            temperature_difference=6
            totalfloorarea='=(K2/I2)'





            char=get_column_letter(3)
            portfolio[char+str(40)]="Origination Fees"
            char=get_column_letter(3)
            portfolio[char+str(41)]="Installer Commission (%)"
            char=get_column_letter(3)
            portfolio[char+str(42)]="Cut of Savings (%)"
            char=get_column_letter(3)
            portfolio[char+str(43)]="Cost of Retrofitting (Calculated)"
            char=get_column_letter(3)
            portfolio[char+str(44)]="Interest Rate (%)"
            char=get_column_letter(3)
            portfolio[char+str(45)]="Loan Term"
            char=get_column_letter(3)
            portfolio[char+str(46)]="Years left on Current Lease"
            char=get_column_letter(3)
            portfolio[char+str(47)]="Savings to Tenants (%)"
            char=get_column_letter(3)
            portfolio[char+str(48)]="Discount Rate (%)"
            char=get_column_letter(3)
            portfolio[char+str(49)]="Cashback"
            char=get_column_letter(3)
            portfolio[char+str(50)]="Temperature Difference"
            char=get_column_letter(3)
            portfolio[char+str(51)]="Window low (Cost of Retrofitting)"
            char=get_column_letter(3)
            portfolio[char+str(52)]="Door low (Cost of Retrofitting)"
            char=get_column_letter(3)
            portfolio[char+str(53)]="Wall low (Cost of Retrofitting)"
            char=get_column_letter(3)
            portfolio[char+str(54)]="Roof low (Cost of Retrofitting)"
            char=get_column_letter(3)
            portfolio[char+str(55)]="Floor low (Cost of Retrofitting)"
            char=get_column_letter(3)
            portfolio[char+str(56)]="Window Avg. (Cost of Retrofitting)"
            char=get_column_letter(3)
            portfolio[char+str(57)]="Door Avg. (Cost of Retrofitting)"
            char=get_column_letter(3)
            portfolio[char+str(58)]="Wall Avg. (Cost of Retrofitting)"
            char=get_column_letter(3)
            portfolio[char+str(59)]="Roof Avg. (Cost of Retrofitting)"
            char=get_column_letter(3)
            portfolio[char+str(60)]="Floor Avg. (Cost of Retrofitting)"
            char=get_column_letter(3)
            portfolio[char+str(61)]="Window High (Cost of Retrofitting)"
            char=get_column_letter(3)
            portfolio[char+str(62)]="Door High (Cost of Retrofitting)"
            char=get_column_letter(3)
            portfolio[char+str(63)]="Wall High (Cost of Retrofitting)"
            char=get_column_letter(3)
            portfolio[char+str(64)]="Roof High (Cost of Retrofitting)"
            char=get_column_letter(3)
            portfolio[char+str(65)]="Floor High (Cost of Retrofitting)"
            char=get_column_letter(3)
            portfolio[char+str(66)]="Average Monthly Bill"
            char=get_column_letter(3)
            portfolio[char+str(67)]="Unit elec/gas cost per kWh"
            char=get_column_letter(3)
            portfolio[char+str(68)]="U_value L-Doc Based on Area (Calculated)"

            char=get_column_letter(4)
            portfolio[char+str(40)]=origination_fees
            char=get_column_letter(4)
            portfolio[char+str(41)]=installer_commission
            char=get_column_letter(4)
            portfolio[char+str(42)]=cut_of_savings
            char=get_column_letter(4)
            portfolio[char+str(43)]=cost_of_retrofitting
            char=get_column_letter(4)
            portfolio[char+str(44)]=interest_rate
            char=get_column_letter(4)
            portfolio[char+str(45)]=loan_term
            char=get_column_letter(4)
            portfolio[char+str(46)]=years_left_on_current_lease
            char=get_column_letter(4)
            portfolio[char+str(47)]=savings_to_tenant
            char=get_column_letter(4)
            portfolio[char+str(48)]=discount_rate
            char=get_column_letter(4)
            portfolio[char+str(49)]=cashback
            char=get_column_letter(4)
            portfolio[char+str(50)]=temperature_difference
            char=get_column_letter(4)
            portfolio[char+str(51)]=30
            char=get_column_letter(4)
            portfolio[char+str(52)]=150
            char=get_column_letter(4)
            portfolio[char+str(53)]=50
            char=get_column_letter(4)
            portfolio[char+str(54)]=50
            char=get_column_letter(4)
            portfolio[char+str(55)]=20
            char=get_column_letter(4)
            portfolio[char+str(56)]=45
            char=get_column_letter(4)
            portfolio[char+str(57)]=250
            char=get_column_letter(4)
            portfolio[char+str(58)]=100
            char=get_column_letter(4)
            portfolio[char+str(59)]=100
            char=get_column_letter(4)
            portfolio[char+str(60)]=40
            char=get_column_letter(4)
            portfolio[char+str(61)]=60
            char=get_column_letter(4)
            portfolio[char+str(62)]=350
            char=get_column_letter(4)
            portfolio[char+str(63)]=150
            char=get_column_letter(4)
            portfolio[char+str(64)]=300
            char=get_column_letter(4)
            portfolio[char+str(65)]=80
            char=get_column_letter(4)
            portfolio[char+str(66)]=2632
            char=get_column_letter(4)
            portfolio[char+str(67)]=0.28




            heightt=2
            properties=1
            EPCDetails=requests.get(full_url_domestic, headers=headers)
            epcstr=str(EPCDetails.content)
            print(len(epcstr))


            uprns=[]
            if(len(epcstr)>3):

                for i in range(len(EPCDetails.json()['rows'])):

                    print(EPCDetails.json()['rows'][i])

                    uprn=EPCDetails.json()['rows'][i]['uprn']
                    lmk=EPCDetails.json()['rows'][i]['lmk-key']

                    lmkkeyflag=False
                    if(uprn==''):
                        uprn=lmk
                        uprn=uprn[len(str(uprn))-10:len(str(uprn))]


                    if(uprn not in uprns):
                            uprns.append(uprn)
                    else:
                            uprn=lmk
                            uprn=uprn[len(str(uprn))-10:len(str(uprn))]


                    wb.create_sheet(str(uprn))
                    properties=properties+1
                    ws=wb[str(uprn)]
                    ws.title=str(uprn)
                    floor=False

                    inspectionDate=EPCDetails.json()['rows'][i]['inspection-date']
                    add_to_portfolio=True

                    add=EPCDetails.json()['rows'][i]['address']
                    add=re.sub('\W+','', add)
                    add=add.lower()
                    print(add)
                    propertyrow=properties+0
                    for j in range(2,properties,1):
                        add2=str(portfolio[get_column_letter(4)+str(j)].value)
                        add2=re.sub('\W+','', add2)
                        add2=add2.lower()

                        if(add2 in add or add in add2):
                                    print("REPEATED:",add)
                                    firstdate=portfolio[get_column_letter(6)+str(j)].value
                                    seconddate=inspectionDate

                                    firstdate=int(firstdate[0:4])
                                    seconddate=int(seconddate[0:4])

                                    if(firstdate>seconddate):
                                        add_to_portfolio=False
                                        properties=properties-1
                                    else:
                                        add_to_portfolio=True
                                        properties=properties-1
                                        propertyrow=j

                                    break


                    for attribute,value in (EPCDetails.json()['rows'][i].items()):
                        #print("Attribute: ",attribute)
                        if(attribute=="floor-area"):
                            floorarea=float(value)
                            floor=True
                            if(add_to_portfolio==True):
                                char=get_column_letter(7)
                                portfolio[char+str(propertyrow)]=value
                        if(attribute=="address1"):
                            if(add_to_portfolio==True):
                                char=get_column_letter(3)
                                portfolio[char+str(propertyrow)]=value
                        if(attribute=="address"):
                            if(add_to_portfolio==True):
                                char=get_column_letter(4)
                                portfolio[char+str(propertyrow)]=value
                                char=get_column_letter(1)
                                userinputs[char+str(propertyrow)]=value
                        if(attribute=="postcode"):
                            if(add_to_portfolio==True):
                                char=get_column_letter(5)
                                portfolio[char+str(propertyrow)]=value
                        if(attribute=="inspection-date"):
                            if(add_to_portfolio==True):
                                char=get_column_letter(6)
                                portfolio[char+str(propertyrow)]=value
                        if(attribute=="uprn"):
                            if(add_to_portfolio==True):
                                char=get_column_letter(8)
                                portfolio[char+str(propertyrow)]=value
                        if(attribute=="lmk-key"):
                            if(add_to_portfolio==True):
                                char=get_column_letter(9)
                                portfolio[char+str(propertyrow)]=value
                        if(attribute=="building-emissions"):
                            buildingEmissions=(value)

                            if(len(str(buildingEmissions))==0):
                                buildingEmissions=0
                                value="0"
                                print(len(str(buildingEmissions)))


                        #print("Value:",value)
                        #print("Attribute: ",attribute)
                        #print("Value:",value)
                        ws.append([str(attribute), str(value)])



                    #PROPERTY TYPE
                    if(add_to_portfolio==True):
                        char=get_column_letter(2)
                        portfolio[char+str(propertyrow)]="Domestic"

                    if(floor==False):
                        print("THIS UPRNS floor area is N/A",uprn)
                        continue






                    if(inspectionDate[0:4]=="2020" or inspectionDate[0:4]=="2021" or inspectionDate[0:4]=="2022" or inspectionDate[0:4]=="2023"):


                        char=get_column_letter(4)
                        ws[char+str(1)]="Building Attribute"
                        char=get_column_letter(5)
                        ws[char+str(1)]="U_Value L-Doc (W/m^2K)"
                        char=get_column_letter(6)
                        ws[char+str(1)]="U_Value Leti (W/m^2K)"
                        char=get_column_letter(7)
                        ws[char+str(1)]="U_Value Difference (W/m^2K)"
                        char=get_column_letter(8)
                        ws[char+str(1)]="Area"
                        char=get_column_letter(9)
                        ws[char+str(1)]="U_value L-Doc Based on Area"
                        char=get_column_letter(10)
                        ws[char+str(1)]="U_value Leti Based on Area"
                        char=get_column_letter(11)
                        ws[char+str(1)]="U-value Improvement Based on Area"
                        char=get_column_letter(12)
                        ws[char+str(1)]="Savings"

                        #code for cost of retrofit per contituent
                        char=get_column_letter(13)
                        ws[char+str(1)]="low Cost of Retrofit Constituent (m^2)"
                        char=get_column_letter(14)
                        ws[char+str(1)]="Cost of Retrofit Constituent (m^2) / Improvement based on area"
                        char=get_column_letter(15)
                        ws[char+str(1)]="Avg. Cost of Retrofit Constituent (m^2)"
                        char=get_column_letter(16)
                        ws[char+str(1)]="Cost of Retrofit Constituent (m^2) / Improvement based on area"
                        char=get_column_letter(17)
                        ws[char+str(1)]="High Cost of Retrofit Constituent (m^2)"
                        char=get_column_letter(18)
                        ws[char+str(1)]="Cost of Retrofit Constituent (m^2) / Improvement based on area"
                        char=get_column_letter(19)
                        ws[char+str(1)]="Cost of Retrofit Constituent (m^2) * area"
                        char=get_column_letter(20)
                        ws[char+str(1)]="Energy Savings (GBP)"
                        char=get_column_letter(20)
                        ws[char+str(1)]="Energy Savings (kWh)"




                        char=get_column_letter(4)
                        ws[char+str(2)]="Window"
                        char=get_column_letter(5)
                        ws[char+str(2)]=uvalue_2021["window_L"]
                        char=get_column_letter(6)
                        ws[char+str(2)]=uvalue_2021["window_leti"]
                        char=get_column_letter(7)
                        ws[char+str(2)]=uvalue_2021["window_L"]-uvalue_2021["window_leti"]
                        char=get_column_letter(8)
                        ws[char+str(2)]=(math.sqrt(floorarea)*heightt)*(0.20)
                        char=get_column_letter(9)
                        ws[char+str(2)]=(((math.sqrt(floorarea)*heightt)*(0.20)) * uvalue_2021["window_L"])
                        char=get_column_letter(10)
                        ws[char+str(2)]=((math.sqrt(floorarea)*heightt)*(0.20)) * uvalue_2021["window_leti"]
                        char=get_column_letter(11)
                        ws[char+str(2)]=(((math.sqrt(floorarea)*heightt)*(0.20)) * uvalue_2021["window_L"]) - (((math.sqrt(floorarea)*heightt)*(0.20)) * uvalue_2021["window_leti"])
                        char=get_column_letter(12)
                        ws[char+str(2)]=((((math.sqrt(floorarea)*heightt)*(0.20)) * uvalue_2021["window_L"]) - (((math.sqrt(floorarea)*heightt)*(0.20)) * uvalue_2021["window_leti"])) / ((((math.sqrt(floorarea)*heightt)*(0.20)) * uvalue_2021["window_L"]))
                        #BuildingData.json()['data'][0]['height']) THIS IS THE BUILDINGS HEIGHT FROM WHICH WE CANNOT CALCULATE THE WINDOW SIZE



                        char=get_column_letter(4)
                        ws[char+str(3)]="Door"
                        char=get_column_letter(5)
                        ws[char+str(3)]=uvalue_2021["door_L"]
                        char=get_column_letter(6)
                        ws[char+str(3)]=uvalue_2021["door_leti"]
                        char=get_column_letter(7)
                        ws[char+str(3)]=uvalue_2021["door_L"]-uvalue_2021["door_leti"]
                        char=get_column_letter(8)
                        ws[char+str(3)]=(((math.sqrt(floorarea)*heightt)*(0.20))*(0.20))
                        char=get_column_letter(9)
                        ws[char+str(3)]=(((((math.sqrt(floorarea)*heightt)*(0.20))*(0.20))) * uvalue_2021["door_L"])
                        char=get_column_letter(10)
                        ws[char+str(3)]=((((math.sqrt(floorarea)*heightt)*(0.20))*(0.20))) * uvalue_2021["door_leti"]
                        char=get_column_letter(11)
                        ws[char+str(3)]=((((((math.sqrt(floorarea)*heightt)*(0.20))*(0.20))) * uvalue_2021["door_L"]) - (((((math.sqrt(floorarea)*heightt)*(0.20))*(0.20))) * uvalue_2021["door_leti"]))
                        char=get_column_letter(12)
                        ws[char+str(3)]=((((((math.sqrt(floorarea)*heightt)*(0.20))*(0.20))) * uvalue_2021["door_L"]) - (((((math.sqrt(floorarea)*heightt)*(0.20))*(0.20))) * uvalue_2021["door_leti"])) / ((((((math.sqrt(floorarea)*heightt)*(0.20))*(0.20))) * uvalue_2021["door_L"]))

                        char=get_column_letter(4)
                        ws[char+str(4)]="Roof"
                        char=get_column_letter(5)
                        ws[char+str(4)]=uvalue_2021["roof_L"]
                        char=get_column_letter(6)
                        ws[char+str(4)]=uvalue_2021["roof_leti"]
                        char=get_column_letter(7)
                        ws[char+str(4)]=uvalue_2021["roof_L"]-uvalue_2021["roof_leti"]
                        char=get_column_letter(8)
                        ws[char+str(4)]=(floorarea)
                        char=get_column_letter(9)
                        ws[char+str(4)]=((floorarea) * uvalue_2021["roof_L"])
                        char=get_column_letter(10)
                        ws[char+str(4)]=(floorarea) * uvalue_2021["roof_leti"]
                        char=get_column_letter(11)
                        ws[char+str(4)]=(((floorarea) * uvalue_2021["roof_L"]) - ((floorarea) * uvalue_2021["roof_leti"]))
                        char=get_column_letter(12)
                        ws[char+str(4)]=(((floorarea) * uvalue_2021["roof_L"]) - ((floorarea) * uvalue_2021["roof_leti"])) / ((floorarea) * uvalue_2021["roof_L"])


                        char=get_column_letter(4)
                        ws[char+str(5)]="Wall"
                        char=get_column_letter(5)
                        ws[char+str(5)]=uvalue_2021["wall_L"]
                        char=get_column_letter(6)
                        ws[char+str(5)]=uvalue_2021["wall_leti"]
                        char=get_column_letter(7)
                        ws[char+str(5)]=uvalue_2021["wall_L"]-uvalue_2021["wall_leti"]
                        char=get_column_letter(8)
                        ws[char+str(5)]=(math.sqrt(floorarea)*heightt)*(0.80)
                        char=get_column_letter(9)
                        ws[char+str(5)]=(((math.sqrt(floorarea)*heightt)*(0.80)) * uvalue_2021["wall_L"])
                        char=get_column_letter(10)
                        ws[char+str(5)]=((math.sqrt(floorarea)*heightt)*(0.80)) * uvalue_2021["wall_leti"]
                        char=get_column_letter(11)
                        ws[char+str(5)]=((((math.sqrt(floorarea)*heightt)*(0.80)) * uvalue_2021["wall_L"]) - (((math.sqrt(floorarea)*heightt)*(0.80)) * uvalue_2021["wall_leti"]))
                        char=get_column_letter(12)
                        ws[char+str(5)]=((((math.sqrt(floorarea)*heightt)*(0.80)) * uvalue_2021["wall_L"]) - (((math.sqrt(floorarea)*heightt)*(0.80)) * uvalue_2021["wall_leti"])) / (((math.sqrt(floorarea)*heightt)*(0.80)) * uvalue_2021["wall_leti"]  )

                        char=get_column_letter(4)
                        ws[char+str(6)]="Floor"
                        char=get_column_letter(5)
                        ws[char+str(6)]=uvalue_2021["floor_L"]
                        char=get_column_letter(6)
                        ws[char+str(6)]=uvalue_2021["floor_leti"]
                        char=get_column_letter(7)
                        ws[char+str(6)]=uvalue_2021["floor_L"]-uvalue_2021["floor_leti"]
                        char=get_column_letter(8)
                        ws[char+str(6)]=(floorarea)
                        char=get_column_letter(9)
                        ws[char+str(6)]=((floorarea) * uvalue_2021["floor_L"])
                        char=get_column_letter(10)
                        ws[char+str(6)]=(floorarea) * uvalue_2021["floor_leti"]
                        char=get_column_letter(11)
                        ws[char+str(6)]=(((floorarea) * uvalue_2021["floor_L"]) - ((floorarea) * uvalue_2021["floor_leti"]))
                        char=get_column_letter(12)
                        ws[char+str(6)]= (((floorarea) * uvalue_2021["floor_L"]) - ((floorarea) * uvalue_2021["floor_leti"])) / ((floorarea) * uvalue_2021["floor_L"])

                        char=get_column_letter(8)
                        ws[char+str(7)]="Total U-Value"

                        ws['L2']='=(K2/I2)'
                        ws['L3']='=(K3/I3)'
                        ws['L4']='=(K4/I4)'
                        ws['L5']='=(K5/I5)'
                        ws['L6']='=(K6/I6)'
                        ws['L7']='=(K7/I7)'
                        ws['I7']='=SUM(I2:I6)'
                        ws['I8']=f"=UserInputs!B{propertyrow}+UserInputs!C{propertyrow}"
                        ws['J7']='=SUM(J2:J6)'
                        ws['K7']='=SUM(K2:K6)'
                        ws['L8']='=(L7)*I8'

                        #const of retrofit * area
                        ws['S2']='=(H2)*O2'
                        ws['S3']='=(H3)*O3'
                        ws['S4']='=(H4)*O4'
                        ws['S5']='=(H5)*O5'
                        ws['S6']='=(H6)*O6'

                        #TOTAL SAVINGS

                        char=get_column_letter(11)
                        ws[char+str(8)]="Total Savings"

                        #Energy Savings
                        ws['T2']='=((((K2)*(Portfolio!D50))/1000)*365*24*(Portfolio!D67))'
                        ws['T3']='=((((K3)*(Portfolio!D50))/1000)*365*24*(Portfolio!D67))'
                        ws['T4']='=((((K4)*(Portfolio!D50))/1000)*365*24*(Portfolio!D67))'
                        ws['T5']='=((((K5)*(Portfolio!D50))/1000)*365*24*(Portfolio!D67))'
                        ws['T6']='=((((K6)*(Portfolio!D50))/1000)*365*24*(Portfolio!D67))'
                        ws['T7']='=SUM(T2:T6)'


                        ws['U2']=f'=((((K2)*(Portfolio!D50))/1000)*365*24)/12'
                        ws['U3']=f'=((((K3)*(Portfolio!D50))/1000)*365*24)/12'
                        ws['U4']=f'=((((K4)*(Portfolio!D50))/1000)*365*24)/12'
                        ws['U5']=f'=((((K5)*(Portfolio!D50))/1000)*365*24)/12'
                        ws['U6']=f'=((((K6)*(Portfolio!D50))/1000)*365*24)/12'
                        ws['U7']='=SUM(U2:U6)'


                        if(add_to_portfolio==True):
                            #TO ADD TO PORTFOLIO WHERE ALL THE PROPERTY INFORMATION IS SPECIFIED

                            #difference u value
                            char=get_column_letter(11)
                            portfolio[char+str(propertyrow)]=uvalue_2021["window_L"]-uvalue_2021["window_leti"]
                            char=get_column_letter(10)
                            portfolio[char+str(propertyrow)]=uvalue_2021["door_L"]-uvalue_2021["door_leti"]
                            char=get_column_letter(12)
                            portfolio[char+str(propertyrow)]=uvalue_2021["roof_L"]-uvalue_2021["roof_leti"]
                            char=get_column_letter(13)
                            portfolio[char+str(propertyrow)]=uvalue_2021["wall_L"]-uvalue_2021["wall_leti"]
                            char=get_column_letter(14)
                            portfolio[char+str(propertyrow)]=uvalue_2021["floor_L"]-uvalue_2021["floor_leti"]

                            #area
                            char=get_column_letter(16)
                            portfolio[char+str(propertyrow)]=(math.sqrt(floorarea)*heightt)*(0.20)
                            char=get_column_letter(15)
                            portfolio[char+str(propertyrow)]=(((math.sqrt(floorarea)*heightt)*(0.20))*(0.20))
                            char=get_column_letter(17)
                            portfolio[char+str(propertyrow)]=(floorarea)
                            char=get_column_letter(18)
                            portfolio[char+str(propertyrow)]=(math.sqrt(floorarea)*heightt)*(0.80)
                            char=get_column_letter(19)
                            portfolio[char+str(propertyrow)]=(floorarea)

                            #improvement
                            char=get_column_letter(21)
                            portfolio[char+str(propertyrow)]=(((math.sqrt(floorarea)*heightt)*(0.20)) * uvalue_2021["window_L"]) - (((math.sqrt(floorarea)*heightt)*(0.20)) * uvalue_2021["window_leti"])
                            char=get_column_letter(20)
                            portfolio[char+str(propertyrow)]=((((((math.sqrt(floorarea)*heightt)*(0.20))*(0.20))) * uvalue_2021["door_L"]) - (((((math.sqrt(floorarea)*heightt)*(0.20))*(0.20))) * uvalue_2021["door_leti"]))
                            char=get_column_letter(22)
                            portfolio[char+str(propertyrow)]=(((floorarea) * uvalue_2021["roof_L"]) - ((floorarea) * uvalue_2021["roof_leti"]))
                            char=get_column_letter(23)
                            portfolio[char+str(propertyrow)]=((((math.sqrt(floorarea)*heightt)*(0.80)) * uvalue_2021["wall_L"]) - (((math.sqrt(floorarea)*heightt)*(0.80)) * uvalue_2021["wall_leti"]))
                            char=get_column_letter(24)
                            portfolio[char+str(propertyrow)]=(((floorarea) * uvalue_2021["floor_L"]) - ((floorarea) * uvalue_2021["floor_leti"]))
                            char=get_column_letter(25)
                            portfolio[char+str(propertyrow)]=(((((math.sqrt(floorarea)*heightt)*(0.20)) * uvalue_2021["window_L"]) - (((math.sqrt(floorarea)*heightt)*(0.20)) * uvalue_2021["window_leti"])) + ((((((math.sqrt(floorarea)*heightt)*(0.20))*(0.20))) * uvalue_2021["door_L"]) - (((((math.sqrt(floorarea)*heightt)*(0.20))*(0.20))) * uvalue_2021["door_leti"])) + (((floorarea) * uvalue_2021["roof_L"]) - ((floorarea) * uvalue_2021["roof_leti"])) + ((((math.sqrt(floorarea)*heightt)*(0.80)) * uvalue_2021["wall_L"]) - (((math.sqrt(floorarea)*heightt)*(0.80)) * uvalue_2021["wall_leti"])) + (((floorarea) * uvalue_2021["floor_L"]) - ((floorarea) * uvalue_2021["floor_leti"])))




                            #savings
                            char=get_column_letter(27)
                            portfolio[char+str(propertyrow)]=((((math.sqrt(floorarea)*heightt)*(0.20)) * uvalue_2021["window_L"]) - (((math.sqrt(floorarea)*heightt)*(0.20)) * uvalue_2021["window_leti"])) / ((((math.sqrt(floorarea)*heightt)*(0.20)) * uvalue_2021["window_L"]))
                            char=get_column_letter(26)
                            portfolio[char+str(propertyrow)]=((((((math.sqrt(floorarea)*heightt)*(0.20))*(0.20))) * uvalue_2021["door_L"]) - (((((math.sqrt(floorarea)*heightt)*(0.20))*(0.20))) * uvalue_2021["door_leti"])) / ((((((math.sqrt(floorarea)*heightt)*(0.20))*(0.20))) * uvalue_2021["door_L"]))
                            char=get_column_letter(28)
                            portfolio[char+str(propertyrow)]=(((floorarea) * uvalue_2021["roof_L"]) - ((floorarea) * uvalue_2021["roof_leti"])) / ((floorarea) * uvalue_2021["roof_L"])
                            char=get_column_letter(29)
                            portfolio[char+str(propertyrow)]=((((math.sqrt(floorarea)*heightt)*(0.80)) * uvalue_2021["wall_L"]) - (((math.sqrt(floorarea)*heightt)*(0.80)) * uvalue_2021["wall_leti"])) / (((math.sqrt(floorarea)*heightt)*(0.80)) * uvalue_2021["wall_L"])
                            char=get_column_letter(30)
                            portfolio[char+str(propertyrow)]=(((floorarea) * uvalue_2021["floor_L"]) - ((floorarea) * uvalue_2021["floor_leti"])) / ((floorarea) * uvalue_2021["floor_L"])
                            char=get_column_letter(31)
                            portfolio[char+str(propertyrow)]=(((((math.sqrt(floorarea)*heightt)*(0.20)) * uvalue_2021["window_L"]) - (((math.sqrt(floorarea)*heightt)*(0.20)) * uvalue_2021["window_leti"])) + ((((((math.sqrt(floorarea)*heightt)*(0.20))*(0.20))) * uvalue_2021["door_L"]) - (((((math.sqrt(floorarea)*heightt)*(0.20))*(0.20))) * uvalue_2021["door_leti"])) + (((floorarea) * uvalue_2021["roof_L"]) - ((floorarea) * uvalue_2021["roof_leti"])) + ((((math.sqrt(floorarea)*heightt)*(0.80)) * uvalue_2021["wall_L"]) - (((math.sqrt(floorarea)*heightt)*(0.80)) * uvalue_2021["wall_leti"])) + (((floorarea) * uvalue_2021["floor_L"]) - ((floorarea) * uvalue_2021["floor_leti"]))) / ((((math.sqrt(floorarea)*heightt)*(0.20)) * uvalue_2021["window_L"]) + (((((math.sqrt(floorarea)*heightt)*(0.20))*(0.20))) * uvalue_2021["door_L"]) + ((floorarea) * uvalue_2021["roof_L"]) + (((math.sqrt(floorarea)*heightt)*(0.80)) * uvalue_2021["wall_L"]) + ((floorarea) * uvalue_2021["floor_L"]))

                            #FINAL SAVINGS
                            char=get_column_letter(32)
                            portfolio[char+str(propertyrow)]='='+str(uprn)+'!L8*12'



                            #cost of retrofit per meter square for all constituents
                            char=get_column_letter(33)
                            portfolio[get_column_letter(41)+str(propertyrow)]='=SUM('+str(uprn)+'!S2:S6)'

                            up=str(uprn)

                            char=get_column_letter(51)
                            portfolio[char+str(propertyrow)]=f"={up}!L7*{buildingEmissions}"
                            char=get_column_letter(52)
                            portfolio[char+str(propertyrow)]=f"={up}!T2"
                            char=get_column_letter(53)
                            portfolio[char+str(propertyrow)]=f"={up}!T3"
                            char=get_column_letter(54)
                            portfolio[char+str(propertyrow)]=f"={up}!T4"
                            char=get_column_letter(55)
                            portfolio[char+str(propertyrow)]=f"={up}!T5"
                            char=get_column_letter(56)
                            portfolio[char+str(propertyrow)]=f"={up}!T6"

                            char=get_column_letter(57)
                            portfolio[char+str(propertyrow)]=f"={up}!U7"
                            char=get_column_letter(58)
                            portfolio[char+str(propertyrow)]=f"=INDEX({up}!U2:U6, MATCH(MAX({up}!L2:L6), {up}!L2:L6, 0))"
                            char=get_column_letter(59)
                            portfolio[char+str(propertyrow)]=f"=(INDEX({up}!H2:H6, MATCH(MAX({up}!L2:L6), {up}!L2:L6, 0))/{up}!B36)*{up}!B1"
                            char=get_column_letter(60)
                            portfolio[char+str(propertyrow)]=f"=INDEX({up}!T2:T6, MATCH(MAX({up}!L2:L6), {up}!L2:L6, 0))"
                            char=get_column_letter(61)
                            portfolio[char+str(propertyrow)]=f"={up}!I7"






                        #cost of constituent outputs
                        #cost of constituent outputs
                        char=get_column_letter(13)
                        ws[char+str(2)]=f"=Portfolio!D51"
                        char=get_column_letter(13)
                        ws[char+str(3)]=f"=Portfolio!D52"
                        char=get_column_letter(13)
                        ws[char+str(4)]=f"=Portfolio!D53"
                        char=get_column_letter(13)
                        ws[char+str(5)]=f"=Portfolio!D54"
                        char=get_column_letter(13)
                        ws[char+str(6)]=f"=Portfolio!D55"

                        char=get_column_letter(15)
                        ws[char+str(2)]=f"=Portfolio!D56"
                        char=get_column_letter(15)
                        ws[char+str(3)]=f"=Portfolio!D57"
                        char=get_column_letter(15)
                        ws[char+str(4)]=f"=Portfolio!D58"
                        char=get_column_letter(15)
                        ws[char+str(5)]=f"=Portfolio!D59"
                        char=get_column_letter(15)
                        ws[char+str(6)]=f"=Portfolio!D60"

                        char=get_column_letter(17)
                        ws[char+str(2)]=f"=Portfolio!D61"
                        char=get_column_letter(17)
                        ws[char+str(3)]=f"=Portfolio!D62"
                        char=get_column_letter(17)
                        ws[char+str(4)]=f"=Portfolio!D63"
                        char=get_column_letter(17)
                        ws[char+str(5)]=f"=Portfolio!D64"
                        char=get_column_letter(17)
                        ws[char+str(6)]=f"=Portfolio!D65"


                        ws['N2']='=(M2/K2)'
                        ws['N3']='=(M3/K3)'
                        ws['N4']='=(M4/K4)'
                        ws['N5']='=(M5/K5)'
                        ws['N6']='=(M6/K6)'

                        ws['P2']='=(O2/K2)'
                        ws['P3']='=(O3/K3)'
                        ws['P4']='=(O4/K4)'
                        ws['P5']='=(O5/K5)'
                        ws['P6']='=(O6/K6)'

                        ws['R2']='=(Q2/K2)'
                        ws['R3']='=(Q3/K3)'
                        ws['R4']='=(Q4/K4)'
                        ws['R5']='=(Q5/K5)'
                        ws['R6']='=(Q6/K6)'


                    if(inspectionDate[0:4]=="2010" or inspectionDate[0:4]=="2008"):


                        char=get_column_letter(4)
                        ws[char+str(1)]="Building Attribute"
                        char=get_column_letter(5)
                        ws[char+str(1)]="U_Value L-Doc (W/m^2K)"
                        char=get_column_letter(6)
                        ws[char+str(1)]="U_Value Leti (W/m^2K)"
                        char=get_column_letter(7)
                        ws[char+str(1)]="U_Value Difference (W/m^2K)"
                        char=get_column_letter(8)
                        ws[char+str(1)]="Area"
                        char=get_column_letter(9)
                        ws[char+str(1)]="U_value L-Doc Based on Area"
                        char=get_column_letter(10)
                        ws[char+str(1)]="U_value Leti Based on Area"
                        char=get_column_letter(11)
                        ws[char+str(1)]="U-value Improvement Based on Area"
                        char=get_column_letter(12)
                        ws[char+str(1)]="Savings"

                        #code for cost of retrofit per contituent
                        char=get_column_letter(13)
                        ws[char+str(1)]="low Cost of Retrofit Constituent (m^2)"
                        char=get_column_letter(14)
                        ws[char+str(1)]="Cost of Retrofit Constituent (m^2) / Improvement based on area"
                        char=get_column_letter(15)
                        ws[char+str(1)]="Avg. Cost of Retrofit Constituent (m^2)"
                        char=get_column_letter(16)
                        ws[char+str(1)]="Cost of Retrofit Constituent (m^2) / Improvement based on area"
                        char=get_column_letter(17)
                        ws[char+str(1)]="High Cost of Retrofit Constituent (m^2)"
                        char=get_column_letter(18)
                        ws[char+str(1)]="Cost of Retrofit Constituent (m^2) / Improvement based on area"
                        char=get_column_letter(19)
                        ws[char+str(1)]="Cost of Retrofit Constituent (m^2) * area"
                        char=get_column_letter(20)
                        ws[char+str(1)]="Energy Savings (GBP)"
                        char=get_column_letter(21)
                        ws[char+str(1)]="Energy Savings (kWh)"

                        char=get_column_letter(4)
                        ws[char+str(2)]="Window"
                        char=get_column_letter(5)
                        ws[char+str(2)]=uvalue_2010["window_L"]
                        char=get_column_letter(6)
                        ws[char+str(2)]=uvalue_2010["window_leti"]
                        char=get_column_letter(7)
                        ws[char+str(2)]=uvalue_2010["window_L"]-uvalue_2010["window_leti"]
                        char=get_column_letter(8)
                        ws[char+str(2)]=(math.sqrt(floorarea)*heightt)*(0.20)
                        char=get_column_letter(9)
                        ws[char+str(2)]=(((math.sqrt(floorarea)*heightt)*(0.20)) * uvalue_2010["window_L"])
                        char=get_column_letter(10)
                        ws[char+str(2)]=((math.sqrt(floorarea)*heightt)*(0.20)) * uvalue_2010["window_leti"]
                        char=get_column_letter(11)
                        ws[char+str(2)]=(((math.sqrt(floorarea)*heightt)*(0.20)) * uvalue_2010["window_L"]) - (((math.sqrt(floorarea)*heightt)*(0.20)) * uvalue_2010["window_leti"])
                        char=get_column_letter(12)
                        ws[char+str(2)]=((((math.sqrt(floorarea)*heightt)*(0.20)) * uvalue_2010["window_L"]) - (((math.sqrt(floorarea)*heightt)*(0.20)) * uvalue_2010["window_leti"])) / ((((math.sqrt(floorarea)*heightt)*(0.20)) * uvalue_2010["window_L"]))
                        #BuildingData.json()['data'][0]['height']) THIS IS THE BUILDINGS HEIGHT FROM WHICH WE CANNOT CALCULATE THE WINDOW SIZE



                        char=get_column_letter(4)
                        ws[char+str(3)]="Door"
                        char=get_column_letter(5)
                        ws[char+str(3)]=uvalue_2010["door_L"]
                        char=get_column_letter(6)
                        ws[char+str(3)]=uvalue_2010["door_leti"]
                        char=get_column_letter(7)
                        ws[char+str(3)]=uvalue_2010["door_L"]-uvalue_2010["door_leti"]
                        char=get_column_letter(8)
                        ws[char+str(3)]=(((math.sqrt(floorarea)*heightt)*(0.20))*(0.20))
                        char=get_column_letter(9)
                        ws[char+str(3)]=(((((math.sqrt(floorarea)*heightt)*(0.20))*(0.20))) * uvalue_2010["door_L"])
                        char=get_column_letter(10)
                        ws[char+str(3)]=((((math.sqrt(floorarea)*heightt)*(0.20))*(0.20))) * uvalue_2010["door_leti"]
                        char=get_column_letter(11)
                        ws[char+str(3)]=((((((math.sqrt(floorarea)*heightt)*(0.20))*(0.20))) * uvalue_2010["door_L"]) - (((((math.sqrt(floorarea)*heightt)*(0.20))*(0.20))) * uvalue_2010["door_leti"]))
                        char=get_column_letter(12)
                        ws[char+str(3)]=((((((math.sqrt(floorarea)*heightt)*(0.20))*(0.20))) * uvalue_2010["door_L"]) - (((((math.sqrt(floorarea)*heightt)*(0.20))*(0.20))) * uvalue_2010["door_leti"])) / ((((((math.sqrt(floorarea)*heightt)*(0.20))*(0.20))) * uvalue_2010["door_L"]))

                        char=get_column_letter(4)
                        ws[char+str(4)]="Roof"
                        char=get_column_letter(5)
                        ws[char+str(4)]=uvalue_2010["roof_L"]
                        char=get_column_letter(6)
                        ws[char+str(4)]=uvalue_2010["roof_leti"]
                        char=get_column_letter(7)
                        ws[char+str(4)]=uvalue_2010["roof_L"]-uvalue_2010["roof_leti"]
                        char=get_column_letter(8)
                        ws[char+str(4)]=(floorarea)
                        char=get_column_letter(9)
                        ws[char+str(4)]=((floorarea) * uvalue_2010["roof_L"])
                        char=get_column_letter(10)
                        ws[char+str(4)]=(floorarea) * uvalue_2010["roof_leti"]
                        char=get_column_letter(11)
                        ws[char+str(4)]=(((floorarea) * uvalue_2010["roof_L"]) - ((floorarea) * uvalue_2010["roof_leti"]))
                        char=get_column_letter(12)
                        ws[char+str(4)]=(((floorarea) * uvalue_2010["roof_L"]) - ((floorarea) * uvalue_2010["roof_leti"])) / ((floorarea) * uvalue_2010["roof_L"])


                        char=get_column_letter(4)
                        ws[char+str(5)]="Wall"
                        char=get_column_letter(5)
                        ws[char+str(5)]=uvalue_2010["wall_L"]
                        char=get_column_letter(6)
                        ws[char+str(5)]=uvalue_2010["wall_leti"]
                        char=get_column_letter(7)
                        ws[char+str(5)]=uvalue_2010["wall_L"]-uvalue_2010["wall_leti"]
                        char=get_column_letter(8)
                        ws[char+str(5)]=(math.sqrt(floorarea)*heightt)*(0.80)
                        char=get_column_letter(9)
                        ws[char+str(5)]=(((math.sqrt(floorarea)*heightt)*(0.80)) * uvalue_2010["wall_L"])
                        char=get_column_letter(10)
                        ws[char+str(5)]=((math.sqrt(floorarea)*heightt)*(0.80)) * uvalue_2010["wall_leti"]
                        char=get_column_letter(11)
                        ws[char+str(5)]=((((math.sqrt(floorarea)*heightt)*(0.80)) * uvalue_2010["wall_L"]) - (((math.sqrt(floorarea)*heightt)*(0.80)) * uvalue_2010["wall_leti"]))
                        char=get_column_letter(12)
                        ws[char+str(5)]=((((math.sqrt(floorarea)*heightt)*(0.80)) * uvalue_2010["wall_L"]) - (((math.sqrt(floorarea)*heightt)*(0.80)) * uvalue_2010["wall_leti"])) / (((math.sqrt(floorarea)*heightt)*(0.80)) * uvalue_2010["wall_leti"]  )

                        char=get_column_letter(4)
                        ws[char+str(6)]="Floor"
                        char=get_column_letter(5)
                        ws[char+str(6)]=uvalue_2010["floor_L"]
                        char=get_column_letter(6)
                        ws[char+str(6)]=uvalue_2010["floor_leti"]
                        char=get_column_letter(7)
                        ws[char+str(6)]=uvalue_2010["floor_L"]-uvalue_2010["floor_leti"]
                        char=get_column_letter(8)
                        ws[char+str(6)]=(floorarea)
                        char=get_column_letter(9)
                        ws[char+str(6)]=((floorarea) * uvalue_2010["floor_L"])
                        char=get_column_letter(10)
                        ws[char+str(6)]=(floorarea) * uvalue_2010["floor_leti"]
                        char=get_column_letter(11)
                        ws[char+str(6)]=(((floorarea) * uvalue_2010["floor_L"]) - ((floorarea) * uvalue_2010["floor_leti"]))
                        char=get_column_letter(12)
                        ws[char+str(6)]= (((floorarea) * uvalue_2010["floor_L"]) - ((floorarea) * uvalue_2010["floor_leti"])) / ((floorarea) * uvalue_2010["floor_L"])

                        char=get_column_letter(8)
                        ws[char+str(7)]="Total U-Value"

                        ws['L2']='=(K2/I2)'
                        ws['L3']='=(K3/I3)'
                        ws['L4']='=(K4/I4)'
                        ws['L5']='=(K5/I5)'
                        ws['L6']='=(K6/I6)'
                        ws['L7']='=(K7/I7)'
                        ws['I7']='=SUM(I2:I6)'
                        ws['I8']=f"=UserInputs!B{propertyrow}+UserInputs!C{propertyrow}"
                        ws['J7']='=SUM(J2:J6)'
                        ws['K7']='=SUM(K2:K6)'
                        ws['L8']='=(L7)*I8'


                        #const of retrofit * area
                        ws['S2']='=(H2)*O2'
                        ws['S3']='=(H3)*O3'
                        ws['S4']='=(H4)*O4'
                        ws['S5']='=(H5)*O5'
                        ws['S6']='=(H6)*O6'

                        #TOTAL SAVINGS

                        char=get_column_letter(11)
                        ws[char+str(8)]="Total Savings"


                        #Energy Savings
                        ws['T2']='=((((K2)*(Portfolio!D50))/1000)*365*24*(Portfolio!D67))'
                        ws['T3']='=((((K3)*(Portfolio!D50))/1000)*365*24*(Portfolio!D67))'
                        ws['T4']='=((((K4)*(Portfolio!D50))/1000)*365*24*(Portfolio!D67))'
                        ws['T5']='=((((K5)*(Portfolio!D50))/1000)*365*24*(Portfolio!D67))'
                        ws['T6']='=((((K6)*(Portfolio!D50))/1000)*365*24*(Portfolio!D67))'
                        ws['T7']='=SUM(T2:T6)'


                        ws['U2']=f'=((((K2)*(Portfolio!D50))/1000)*365*24)/12'
                        ws['U3']=f'=((((K3)*(Portfolio!D50))/1000)*365*24)/12'
                        ws['U4']=f'=((((K4)*(Portfolio!D50))/1000)*365*24)/12'
                        ws['U5']=f'=((((K5)*(Portfolio!D50))/1000)*365*24)/12'
                        ws['U6']=f'=((((K6)*(Portfolio!D50))/1000)*365*24)/12'
                        ws['U7']='=SUM(U2:U6)'


                        if(add_to_portfolio==True):
                            #TO ADD TO PORTFOLIO

                            #difference u value
                            char=get_column_letter(11)
                            portfolio[char+str(propertyrow)]=uvalue_2010["window_L"]-uvalue_2010["window_leti"]
                            char=get_column_letter(10)
                            portfolio[char+str(propertyrow)]=uvalue_2010["door_L"]-uvalue_2010["door_leti"]
                            char=get_column_letter(12)
                            portfolio[char+str(propertyrow)]=uvalue_2010["roof_L"]-uvalue_2010["roof_leti"]
                            char=get_column_letter(13)
                            portfolio[char+str(propertyrow)]=uvalue_2010["wall_L"]-uvalue_2010["wall_leti"]
                            char=get_column_letter(14)
                            portfolio[char+str(propertyrow)]=uvalue_2010["floor_L"]-uvalue_2010["floor_leti"]

                            #area
                            char=get_column_letter(16)
                            portfolio[char+str(propertyrow)]=(math.sqrt(floorarea)*heightt)*(0.20)
                            char=get_column_letter(15)
                            portfolio[char+str(propertyrow)]=(((math.sqrt(floorarea)*heightt)*(0.20))*(0.20))
                            char=get_column_letter(17)
                            portfolio[char+str(propertyrow)]=(floorarea)
                            char=get_column_letter(18)
                            portfolio[char+str(propertyrow)]=(math.sqrt(floorarea)*heightt)*(0.80)
                            char=get_column_letter(19)
                            portfolio[char+str(propertyrow)]=(floorarea)

                            #improvement
                            char=get_column_letter(21)
                            portfolio[char+str(propertyrow)]=(((math.sqrt(floorarea)*heightt)*(0.20)) * uvalue_2010["window_L"]) - (((math.sqrt(floorarea)*heightt)*(0.20)) * uvalue_2010["window_leti"])
                            char=get_column_letter(20)
                            portfolio[char+str(propertyrow)]=((((((math.sqrt(floorarea)*heightt)*(0.20))*(0.20))) * uvalue_2010["door_L"]) - (((((math.sqrt(floorarea)*heightt)*(0.20))*(0.20))) * uvalue_2010["door_leti"]))
                            char=get_column_letter(22)
                            portfolio[char+str(propertyrow)]=(((floorarea) * uvalue_2010["roof_L"]) - ((floorarea) * uvalue_2010["roof_leti"]))
                            char=get_column_letter(23)
                            portfolio[char+str(propertyrow)]=((((math.sqrt(floorarea)*heightt)*(0.80)) * uvalue_2010["wall_L"]) - (((math.sqrt(floorarea)*heightt)*(0.80)) * uvalue_2010["wall_leti"]))
                            char=get_column_letter(24)
                            portfolio[char+str(propertyrow)]=(((floorarea) * uvalue_2010["floor_L"]) - ((floorarea) * uvalue_2010["floor_leti"]))
                            char=get_column_letter(25)
                            portfolio[char+str(propertyrow)]=(((((math.sqrt(floorarea)*heightt)*(0.20)) * uvalue_2010["window_L"]) - (((math.sqrt(floorarea)*heightt)*(0.20)) * uvalue_2010["window_leti"])) + ((((((math.sqrt(floorarea)*heightt)*(0.20))*(0.20))) * uvalue_2010["door_L"]) - (((((math.sqrt(floorarea)*heightt)*(0.20))*(0.20))) * uvalue_2010["door_leti"])) + (((floorarea) * uvalue_2010["roof_L"]) - ((floorarea) * uvalue_2010["roof_leti"])) + ((((math.sqrt(floorarea)*heightt)*(0.80)) * uvalue_2010["wall_L"]) - (((math.sqrt(floorarea)*heightt)*(0.80)) * uvalue_2010["wall_leti"])) + (((floorarea) * uvalue_2010["floor_L"]) - ((floorarea) * uvalue_2010["floor_leti"])))


                            #savings
                            char=get_column_letter(27)
                            portfolio[char+str(propertyrow)]=((((math.sqrt(floorarea)*heightt)*(0.20)) * uvalue_2010["window_L"]) - (((math.sqrt(floorarea)*heightt)*(0.20)) * uvalue_2010["window_leti"])) / ((((math.sqrt(floorarea)*heightt)*(0.20)) * uvalue_2010["window_L"]))
                            char=get_column_letter(26)
                            portfolio[char+str(propertyrow)]=((((((math.sqrt(floorarea)*heightt)*(0.20))*(0.20))) * uvalue_2010["door_L"]) - (((((math.sqrt(floorarea)*heightt)*(0.20))*(0.20))) * uvalue_2010["door_leti"])) / ((((((math.sqrt(floorarea)*heightt)*(0.20))*(0.20))) * uvalue_2010["door_L"]))
                            char=get_column_letter(28)
                            portfolio[char+str(propertyrow)]=(((floorarea) * uvalue_2010["roof_L"]) - ((floorarea) * uvalue_2010["roof_leti"])) / ((floorarea) * uvalue_2010["roof_L"])
                            char=get_column_letter(29)
                            portfolio[char+str(propertyrow)]=((((math.sqrt(floorarea)*heightt)*(0.80)) * uvalue_2010["wall_L"]) - (((math.sqrt(floorarea)*heightt)*(0.80)) * uvalue_2010["wall_leti"])) / (((math.sqrt(floorarea)*heightt)*(0.80)) * uvalue_2010["wall_L"]  )
                            char=get_column_letter(30)
                            portfolio[char+str(propertyrow)]=(((floorarea) * uvalue_2010["floor_L"]) - ((floorarea) * uvalue_2010["floor_leti"])) / ((floorarea) * uvalue_2010["floor_L"])
                            char=get_column_letter(31)
                            portfolio[char+str(propertyrow)]=(((((math.sqrt(floorarea)*heightt)*(0.20)) * uvalue_2010["window_L"]) - (((math.sqrt(floorarea)*heightt)*(0.20)) * uvalue_2010["window_leti"])) + ((((((math.sqrt(floorarea)*heightt)*(0.20))*(0.20))) * uvalue_2010["door_L"]) - (((((math.sqrt(floorarea)*heightt)*(0.20))*(0.20))) * uvalue_2010["door_leti"])) + (((floorarea) * uvalue_2010["roof_L"]) - ((floorarea) * uvalue_2010["roof_leti"])) + ((((math.sqrt(floorarea)*heightt)*(0.80)) * uvalue_2010["wall_L"]) - (((math.sqrt(floorarea)*heightt)*(0.80)) * uvalue_2010["wall_leti"])) + (((floorarea) * uvalue_2010["floor_L"]) - ((floorarea) * uvalue_2010["floor_leti"]))) / ((((math.sqrt(floorarea)*heightt)*(0.20)) * uvalue_2010["window_L"]) + (((((math.sqrt(floorarea)*heightt)*(0.20))*(0.20))) * uvalue_2010["door_L"]) + ((floorarea) * uvalue_2010["roof_L"]) + (((math.sqrt(floorarea)*heightt)*(0.80)) * uvalue_2010["wall_L"]) + ((floorarea) * uvalue_2010["floor_L"]))

                            #FINAL SAVINGS
                            char=get_column_letter(32)
                            portfolio[char+str(propertyrow)]='='+str(uprn)+'!L8*12'
                            #cost of retrofit per meter square for all constituents
                            char=get_column_letter(33)
                            portfolio[get_column_letter(41)+str(propertyrow)]='=SUM('+str(uprn)+'!S2:S6)'

                            up=str(uprn)

                            char=get_column_letter(51)
                            portfolio[char+str(propertyrow)]=f"={up}!L7*{buildingEmissions}"
                            char=get_column_letter(52)
                            portfolio[char+str(propertyrow)]=f"={up}!T2"
                            char=get_column_letter(53)
                            portfolio[char+str(propertyrow)]=f"={up}!T3"
                            char=get_column_letter(54)
                            portfolio[char+str(propertyrow)]=f"={up}!T4"
                            char=get_column_letter(55)
                            portfolio[char+str(propertyrow)]=f"={up}!T5"
                            char=get_column_letter(56)
                            portfolio[char+str(propertyrow)]=f"={up}!T6"

                            char=get_column_letter(57)
                            portfolio[char+str(propertyrow)]=f"={up}!U7"
                            char=get_column_letter(58)
                            portfolio[char+str(propertyrow)]=f"=INDEX({up}!U2:U6, MATCH(MAX({up}!L2:L6), {up}!L2:L6, 0))"
                            char=get_column_letter(59)
                            portfolio[char+str(propertyrow)]=f"=(INDEX({up}!H2:H6, MATCH(MAX({up}!L2:L6), {up}!L2:L6, 0))/{up}!B36)*{up}!B1"
                            char=get_column_letter(60)
                            portfolio[char+str(propertyrow)]=f"=INDEX({up}!T2:T6, MATCH(MAX({up}!L2:L6), {up}!L2:L6, 0))"
                            char=get_column_letter(61)
                            portfolio[char+str(propertyrow)]=f"={up}!I7"

                        #cost of constituent outputs
                        #cost of constituent outputs
                        char=get_column_letter(13)
                        ws[char+str(2)]=f"=Portfolio!D51"
                        char=get_column_letter(13)
                        ws[char+str(3)]=f"=Portfolio!D52"
                        char=get_column_letter(13)
                        ws[char+str(4)]=f"=Portfolio!D53"
                        char=get_column_letter(13)
                        ws[char+str(5)]=f"=Portfolio!D54"
                        char=get_column_letter(13)
                        ws[char+str(6)]=f"=Portfolio!D55"

                        char=get_column_letter(15)
                        ws[char+str(2)]=f"=Portfolio!D56"
                        char=get_column_letter(15)
                        ws[char+str(3)]=f"=Portfolio!D57"
                        char=get_column_letter(15)
                        ws[char+str(4)]=f"=Portfolio!D58"
                        char=get_column_letter(15)
                        ws[char+str(5)]=f"=Portfolio!D59"
                        char=get_column_letter(15)
                        ws[char+str(6)]=f"=Portfolio!D60"

                        char=get_column_letter(17)
                        ws[char+str(2)]=f"=Portfolio!D61"
                        char=get_column_letter(17)
                        ws[char+str(3)]=f"=Portfolio!D62"
                        char=get_column_letter(17)
                        ws[char+str(4)]=f"=Portfolio!D63"
                        char=get_column_letter(17)
                        ws[char+str(5)]=f"=Portfolio!D64"
                        char=get_column_letter(17)
                        ws[char+str(6)]=f"=Portfolio!D65"


                        ws['N2']='=(M2/K2)'
                        ws['N3']='=(M3/K3)'
                        ws['N4']='=(M4/K4)'
                        ws['N5']='=(M5/K5)'
                        ws['N6']='=(M6/K6)'

                        ws['P2']='=(O2/K2)'
                        ws['P3']='=(O3/K3)'
                        ws['P4']='=(O4/K4)'
                        ws['P5']='=(O5/K5)'
                        ws['P6']='=(O6/K6)'

                        ws['R2']='=(Q2/K2)'
                        ws['R3']='=(Q3/K3)'
                        ws['R4']='=(Q4/K4)'
                        ws['R5']='=(Q5/K5)'
                        ws['R6']='=(Q6/K6)'
                    if(inspectionDate[0:4]=="2011"):


                        char=get_column_letter(4)
                        ws[char+str(1)]="Building Attribute"
                        char=get_column_letter(5)
                        ws[char+str(1)]="U_Value L-Doc (W/m^2K)"
                        char=get_column_letter(6)
                        ws[char+str(1)]="U_Value Leti (W/m^2K)"
                        char=get_column_letter(7)
                        ws[char+str(1)]="U_Value Difference (W/m^2K)"
                        char=get_column_letter(8)
                        ws[char+str(1)]="Area"
                        char=get_column_letter(9)
                        ws[char+str(1)]="U_value L-Doc Based on Area"
                        char=get_column_letter(10)
                        ws[char+str(1)]="U_value Leti Based on Area"
                        char=get_column_letter(11)
                        ws[char+str(1)]="U-value Improvement Based on Area"
                        char=get_column_letter(12)
                        ws[char+str(1)]="Savings"
                        #code for cost of retrofit per contituent
                        char=get_column_letter(13)
                        ws[char+str(1)]="low Cost of Retrofit Constituent (m^2)"
                        char=get_column_letter(14)
                        ws[char+str(1)]="Cost of Retrofit Constituent (m^2) / Improvement based on area"
                        char=get_column_letter(15)
                        ws[char+str(1)]="Avg. Cost of Retrofit Constituent (m^2)"
                        char=get_column_letter(16)
                        ws[char+str(1)]="Cost of Retrofit Constituent (m^2) / Improvement based on area"
                        char=get_column_letter(17)
                        ws[char+str(1)]="High Cost of Retrofit Constituent (m^2)"
                        char=get_column_letter(18)
                        ws[char+str(1)]="Cost of Retrofit Constituent (m^2) / Improvement based on area"
                        char=get_column_letter(19)
                        ws[char+str(1)]="Cost of Retrofit Constituent (m^2) * area"
                        char=get_column_letter(20)
                        ws[char+str(1)]="Energy Savings (GBP)"
                        char=get_column_letter(21)
                        ws[char+str(1)]="Energy Savings (kWh)"

                        char=get_column_letter(4)
                        ws[char+str(2)]="Window"
                        char=get_column_letter(5)
                        ws[char+str(2)]=uvalue_2011["window_L"]
                        char=get_column_letter(6)
                        ws[char+str(2)]=uvalue_2011["window_leti"]
                        char=get_column_letter(7)
                        ws[char+str(2)]=uvalue_2011["window_L"]-uvalue_2011["window_leti"]
                        char=get_column_letter(8)
                        ws[char+str(2)]=(math.sqrt(floorarea)*heightt)*(0.20)
                        char=get_column_letter(9)
                        ws[char+str(2)]=(((math.sqrt(floorarea)*heightt)*(0.20)) * uvalue_2011["window_L"])
                        char=get_column_letter(10)
                        ws[char+str(2)]=((math.sqrt(floorarea)*heightt)*(0.20)) * uvalue_2011["window_leti"]
                        char=get_column_letter(11)
                        ws[char+str(2)]=(((math.sqrt(floorarea)*heightt)*(0.20)) * uvalue_2011["window_L"]) - (((math.sqrt(floorarea)*heightt)*(0.20)) * uvalue_2011["window_leti"])
                        char=get_column_letter(12)
                        ws[char+str(2)]=((((math.sqrt(floorarea)*heightt)*(0.20)) * uvalue_2011["window_L"]) - (((math.sqrt(floorarea)*heightt)*(0.20)) * uvalue_2011["window_leti"])) / ((((math.sqrt(floorarea)*heightt)*(0.20)) * uvalue_2011["window_L"]))
                        #BuildingData.json()['data'][0]['height']) THIS IS THE BUILDINGS HEIGHT FROM WHICH WE CANNOT CALCULATE THE WINDOW SIZE



                        char=get_column_letter(4)
                        ws[char+str(3)]="Door"
                        char=get_column_letter(5)
                        ws[char+str(3)]=uvalue_2011["door_L"]
                        char=get_column_letter(6)
                        ws[char+str(3)]=uvalue_2011["door_leti"]
                        char=get_column_letter(7)
                        ws[char+str(3)]=uvalue_2011["door_L"]-uvalue_2011["door_leti"]
                        char=get_column_letter(8)
                        ws[char+str(3)]=(((math.sqrt(floorarea)*heightt)*(0.20))*(0.20))
                        char=get_column_letter(9)
                        ws[char+str(3)]=(((((math.sqrt(floorarea)*heightt)*(0.20))*(0.20))) * uvalue_2011["door_L"])
                        char=get_column_letter(10)
                        ws[char+str(3)]=((((math.sqrt(floorarea)*heightt)*(0.20))*(0.20))) * uvalue_2011["door_leti"]
                        char=get_column_letter(11)
                        ws[char+str(3)]=((((((math.sqrt(floorarea)*heightt)*(0.20))*(0.20))) * uvalue_2011["door_L"]) - (((((math.sqrt(floorarea)*heightt)*(0.20))*(0.20))) * uvalue_2011["door_leti"]))
                        char=get_column_letter(12)
                        ws[char+str(3)]=((((((math.sqrt(floorarea)*heightt)*(0.20))*(0.20))) * uvalue_2011["door_L"]) - (((((math.sqrt(floorarea)*heightt)*(0.20))*(0.20))) * uvalue_2011["door_leti"])) / ((((((math.sqrt(floorarea)*heightt)*(0.20))*(0.20))) * uvalue_2011["door_L"]))

                        char=get_column_letter(4)
                        ws[char+str(4)]="Roof"
                        char=get_column_letter(5)
                        ws[char+str(4)]=uvalue_2011["roof_L"]
                        char=get_column_letter(6)
                        ws[char+str(4)]=uvalue_2011["roof_leti"]
                        char=get_column_letter(7)
                        ws[char+str(4)]=uvalue_2011["roof_L"]-uvalue_2011["roof_leti"]
                        char=get_column_letter(8)
                        ws[char+str(4)]=(floorarea)
                        char=get_column_letter(9)
                        ws[char+str(4)]=((floorarea) * uvalue_2011["roof_L"])
                        char=get_column_letter(10)
                        ws[char+str(4)]=(floorarea) * uvalue_2011["roof_leti"]
                        char=get_column_letter(11)
                        ws[char+str(4)]=(((floorarea) * uvalue_2011["roof_L"]) - ((floorarea) * uvalue_2011["roof_leti"]))
                        char=get_column_letter(12)
                        ws[char+str(4)]=(((floorarea) * uvalue_2011["roof_L"]) - ((floorarea) * uvalue_2011["roof_leti"])) / ((floorarea) * uvalue_2011["roof_L"])


                        char=get_column_letter(4)
                        ws[char+str(5)]="Wall"
                        char=get_column_letter(5)
                        ws[char+str(5)]=uvalue_2011["wall_L"]
                        char=get_column_letter(6)
                        ws[char+str(5)]=uvalue_2011["wall_leti"]
                        char=get_column_letter(7)
                        ws[char+str(5)]=uvalue_2011["wall_L"]-uvalue_2011["wall_leti"]
                        char=get_column_letter(8)
                        ws[char+str(5)]=(math.sqrt(floorarea)*heightt)*(0.80)
                        char=get_column_letter(9)
                        ws[char+str(5)]=(((math.sqrt(floorarea)*heightt)*(0.80)) * uvalue_2011["wall_L"])
                        char=get_column_letter(10)
                        ws[char+str(5)]=((math.sqrt(floorarea)*heightt)*(0.80)) * uvalue_2011["wall_leti"]
                        char=get_column_letter(11)
                        ws[char+str(5)]=((((math.sqrt(floorarea)*heightt)*(0.80)) * uvalue_2011["wall_L"]) - (((math.sqrt(floorarea)*heightt)*(0.80)) * uvalue_2011["wall_leti"]))
                        char=get_column_letter(12)
                        ws[char+str(5)]=((((math.sqrt(floorarea)*heightt)*(0.80)) * uvalue_2011["wall_L"]) - (((math.sqrt(floorarea)*heightt)*(0.80)) * uvalue_2011["wall_leti"])) / (((math.sqrt(floorarea)*heightt)*(0.80)) * uvalue_2011["wall_leti"]  )

                        char=get_column_letter(4)
                        ws[char+str(6)]="Floor"
                        char=get_column_letter(5)
                        ws[char+str(6)]=uvalue_2011["floor_L"]
                        char=get_column_letter(6)
                        ws[char+str(6)]=uvalue_2011["floor_leti"]
                        char=get_column_letter(7)
                        ws[char+str(6)]=uvalue_2011["floor_L"]-uvalue_2011["floor_leti"]
                        char=get_column_letter(8)
                        ws[char+str(6)]=(floorarea)
                        char=get_column_letter(9)
                        ws[char+str(6)]=((floorarea) * uvalue_2011["floor_L"])
                        char=get_column_letter(10)
                        ws[char+str(6)]=(floorarea) * uvalue_2011["floor_leti"]
                        char=get_column_letter(11)
                        ws[char+str(6)]=(((floorarea) * uvalue_2011["floor_L"]) - ((floorarea) * uvalue_2011["floor_leti"]))
                        char=get_column_letter(12)
                        ws[char+str(6)]= (((floorarea) * uvalue_2011["floor_L"]) - ((floorarea) * uvalue_2011["floor_leti"])) / ((floorarea) * uvalue_2011["floor_L"])

                        char=get_column_letter(8)
                        ws[char+str(7)]="Total U-Value"

                        ws['L2']='=(K2/I2)'
                        ws['L3']='=(K3/I3)'
                        ws['L4']='=(K4/I4)'
                        ws['L5']='=(K5/I5)'
                        ws['L6']='=(K6/I6)'
                        ws['L7']='=(K7/I7)'
                        ws['I7']='=SUM(I2:I6)'
                        ws['I8']=f"=UserInputs!B{propertyrow}+UserInputs!C{propertyrow}"
                        ws['J7']='=SUM(J2:J6)'
                        ws['K7']='=SUM(K2:K6)'
                        ws['L8']='=(L7)*I8'

                        #const of retrofit * area
                        ws['S2']='=(H2)*O2'
                        ws['S3']='=(H3)*O3'
                        ws['S4']='=(H4)*O4'
                        ws['S5']='=(H5)*O5'
                        ws['S6']='=(H6)*O6'

                        #TOTAL SAVINGS

                        char=get_column_letter(11)
                        ws[char+str(8)]="Total Savings"

                        #Energy Savings
                        ws['T2']='=((((K2)*(Portfolio!D50))/1000)*365*24*(Portfolio!D67))'
                        ws['T3']='=((((K3)*(Portfolio!D50))/1000)*365*24*(Portfolio!D67))'
                        ws['T4']='=((((K4)*(Portfolio!D50))/1000)*365*24*(Portfolio!D67))'
                        ws['T5']='=((((K5)*(Portfolio!D50))/1000)*365*24*(Portfolio!D67))'
                        ws['T6']='=((((K6)*(Portfolio!D50))/1000)*365*24*(Portfolio!D67))'
                        ws['T7']='=SUM(T2:T6)'


                        ws['U2']=f'=((((K2)*(Portfolio!D50))/1000)*365*24)/12'
                        ws['U3']=f'=((((K3)*(Portfolio!D50))/1000)*365*24)/12'
                        ws['U4']=f'=((((K4)*(Portfolio!D50))/1000)*365*24)/12'
                        ws['U5']=f'=((((K5)*(Portfolio!D50))/1000)*365*24)/12'
                        ws['U6']=f'=((((K6)*(Portfolio!D50))/1000)*365*24)/12'
                        ws['U7']='=SUM(U2:U6)'


                        if(add_to_portfolio==True):
                            #TO ADD TO PORTFOLIO

                            #difference u value
                            char=get_column_letter(11)
                            portfolio[char+str(propertyrow)]=uvalue_2011["window_L"]-uvalue_2011["window_leti"]
                            char=get_column_letter(10)
                            portfolio[char+str(propertyrow)]=uvalue_2011["door_L"]-uvalue_2011["door_leti"]
                            char=get_column_letter(12)
                            portfolio[char+str(propertyrow)]=uvalue_2011["roof_L"]-uvalue_2011["roof_leti"]
                            char=get_column_letter(13)
                            portfolio[char+str(propertyrow)]=uvalue_2011["wall_L"]-uvalue_2011["wall_leti"]
                            char=get_column_letter(14)
                            portfolio[char+str(propertyrow)]=uvalue_2011["floor_L"]-uvalue_2011["floor_leti"]

                            #area
                            char=get_column_letter(16)
                            portfolio[char+str(propertyrow)]=(math.sqrt(floorarea)*heightt)*(0.20)
                            char=get_column_letter(15)
                            portfolio[char+str(propertyrow)]=(((math.sqrt(floorarea)*heightt)*(0.20))*(0.20))
                            char=get_column_letter(17)
                            portfolio[char+str(propertyrow)]=(floorarea)
                            char=get_column_letter(18)
                            portfolio[char+str(propertyrow)]=(math.sqrt(floorarea)*heightt)*(0.80)
                            char=get_column_letter(19)
                            portfolio[char+str(propertyrow)]=(floorarea)


                            #improvement
                            char=get_column_letter(21)
                            portfolio[char+str(propertyrow)]=(((math.sqrt(floorarea)*heightt)*(0.20)) * uvalue_2011["window_L"]) - (((math.sqrt(floorarea)*heightt)*(0.20)) * uvalue_2011["window_leti"])
                            char=get_column_letter(20)
                            portfolio[char+str(propertyrow)]=((((((math.sqrt(floorarea)*heightt)*(0.20))*(0.20))) * uvalue_2011["door_L"]) - (((((math.sqrt(floorarea)*heightt)*(0.20))*(0.20))) * uvalue_2011["door_leti"]))
                            char=get_column_letter(22)
                            portfolio[char+str(propertyrow)]=(((floorarea) * uvalue_2011["roof_L"]) - ((floorarea) * uvalue_2011["roof_leti"]))
                            char=get_column_letter(23)
                            portfolio[char+str(propertyrow)]=((((math.sqrt(floorarea)*heightt)*(0.80)) * uvalue_2011["wall_L"]) - (((math.sqrt(floorarea)*heightt)*(0.80)) * uvalue_2011["wall_leti"]))
                            char=get_column_letter(24)
                            portfolio[char+str(propertyrow)]=(((floorarea) * uvalue_2011["floor_L"]) - ((floorarea) * uvalue_2011["floor_leti"]))
                            char=get_column_letter(25)
                            portfolio[char+str(propertyrow)]=(((((math.sqrt(floorarea)*heightt)*(0.20)) * uvalue_2011["window_L"]) - (((math.sqrt(floorarea)*heightt)*(0.20)) * uvalue_2011["window_leti"])) + ((((((math.sqrt(floorarea)*heightt)*(0.20))*(0.20))) * uvalue_2011["door_L"]) - (((((math.sqrt(floorarea)*heightt)*(0.20))*(0.20))) * uvalue_2011["door_leti"])) + (((floorarea) * uvalue_2011["roof_L"]) - ((floorarea) * uvalue_2011["roof_leti"])) + ((((math.sqrt(floorarea)*heightt)*(0.80)) * uvalue_2011["wall_L"]) - (((math.sqrt(floorarea)*heightt)*(0.80)) * uvalue_2011["wall_leti"])) + (((floorarea) * uvalue_2011["floor_L"]) - ((floorarea) * uvalue_2011["floor_leti"])))


                            #savings
                            char=get_column_letter(27)
                            portfolio[char+str(propertyrow)]=((((math.sqrt(floorarea)*heightt)*(0.20)) * uvalue_2011["window_L"]) - (((math.sqrt(floorarea)*heightt)*(0.20)) * uvalue_2011["window_leti"])) / ((((math.sqrt(floorarea)*heightt)*(0.20)) * uvalue_2011["window_L"]))
                            char=get_column_letter(26)
                            portfolio[char+str(propertyrow)]=((((((math.sqrt(floorarea)*heightt)*(0.20))*(0.20))) * uvalue_2011["door_L"]) - (((((math.sqrt(floorarea)*heightt)*(0.20))*(0.20))) * uvalue_2011["door_leti"])) / ((((((math.sqrt(floorarea)*heightt)*(0.20))*(0.20))) * uvalue_2011["door_L"]))
                            char=get_column_letter(28)
                            portfolio[char+str(propertyrow)]=(((floorarea) * uvalue_2011["roof_L"]) - ((floorarea) * uvalue_2011["roof_leti"])) / ((floorarea) * uvalue_2011["roof_L"])
                            char=get_column_letter(29)
                            portfolio[char+str(propertyrow)]=((((math.sqrt(floorarea)*heightt)*(0.80)) * uvalue_2011["wall_L"]) - (((math.sqrt(floorarea)*heightt)*(0.80)) * uvalue_2011["wall_leti"])) / (((math.sqrt(floorarea)*heightt)*(0.80)) * uvalue_2011["wall_L"]  )
                            char=get_column_letter(30)
                            portfolio[char+str(propertyrow)]=(((floorarea) * uvalue_2011["floor_L"]) - ((floorarea) * uvalue_2011["floor_leti"])) / ((floorarea) * uvalue_2011["floor_L"])
                            char=get_column_letter(31)
                            portfolio[char+str(propertyrow)]=(((((math.sqrt(floorarea)*heightt)*(0.20)) * uvalue_2011["window_L"]) - (((math.sqrt(floorarea)*heightt)*(0.20)) * uvalue_2011["window_leti"])) + ((((((math.sqrt(floorarea)*heightt)*(0.20))*(0.20))) * uvalue_2011["door_L"]) - (((((math.sqrt(floorarea)*heightt)*(0.20))*(0.20))) * uvalue_2011["door_leti"])) + (((floorarea) * uvalue_2011["roof_L"]) - ((floorarea) * uvalue_2011["roof_leti"])) + ((((math.sqrt(floorarea)*heightt)*(0.80)) * uvalue_2011["wall_L"]) - (((math.sqrt(floorarea)*heightt)*(0.80)) * uvalue_2011["wall_leti"])) + (((floorarea) * uvalue_2011["floor_L"]) - ((floorarea) * uvalue_2011["floor_leti"]))) / ((((math.sqrt(floorarea)*heightt)*(0.20)) * uvalue_2011["window_L"]) + (((((math.sqrt(floorarea)*heightt)*(0.20))*(0.20))) * uvalue_2011["door_L"]) + ((floorarea) * uvalue_2011["roof_L"]) + (((math.sqrt(floorarea)*heightt)*(0.80)) * uvalue_2011["wall_L"]) + ((floorarea) * uvalue_2011["floor_L"]))

                            #FINAL SAVINGS
                            char=get_column_letter(32)
                            portfolio[char+str(propertyrow)]='='+str(uprn)+'!L8*12'

                            #cost of retrofit per meter square for all constituents
                            char=get_column_letter(33)
                            portfolio[get_column_letter(41)+str(propertyrow)]='=SUM('+str(uprn)+'!S2:S6)'

                            up=str(uprn)

                            char=get_column_letter(51)
                            portfolio[char+str(propertyrow)]=f"={up}!L7*{buildingEmissions}"
                            char=get_column_letter(52)
                            portfolio[char+str(propertyrow)]=f"={up}!T2"
                            char=get_column_letter(53)
                            portfolio[char+str(propertyrow)]=f"={up}!T3"
                            char=get_column_letter(54)
                            portfolio[char+str(propertyrow)]=f"={up}!T4"
                            char=get_column_letter(55)
                            portfolio[char+str(propertyrow)]=f"={up}!T5"
                            char=get_column_letter(56)
                            portfolio[char+str(propertyrow)]=f"={up}!T6"
                            char=get_column_letter(57)
                            portfolio[char+str(propertyrow)]=f"={up}!U7"
                            char=get_column_letter(58)
                            portfolio[char+str(propertyrow)]=f"=INDEX({up}!U2:U6, MATCH(MAX({up}!L2:L6), {up}!L2:L6, 0))"
                            char=get_column_letter(59)
                            portfolio[char+str(propertyrow)]=f"=(INDEX({up}!H2:H6, MATCH(MAX({up}!L2:L6), {up}!L2:L6, 0))/{up}!B36)*{up}!B1"
                            char=get_column_letter(60)
                            portfolio[char+str(propertyrow)]=f"=INDEX({up}!T2:T6, MATCH(MAX({up}!L2:L6), {up}!L2:L6, 0))"
                            char=get_column_letter(61)
                            portfolio[char+str(propertyrow)]=f"={up}!I7"

                        #cost of constituent outputs
                        #cost of constituent outputs
                        char=get_column_letter(13)
                        ws[char+str(2)]=f"=Portfolio!D51"
                        char=get_column_letter(13)
                        ws[char+str(3)]=f"=Portfolio!D52"
                        char=get_column_letter(13)
                        ws[char+str(4)]=f"=Portfolio!D53"
                        char=get_column_letter(13)
                        ws[char+str(5)]=f"=Portfolio!D54"
                        char=get_column_letter(13)
                        ws[char+str(6)]=f"=Portfolio!D55"

                        char=get_column_letter(15)
                        ws[char+str(2)]=f"=Portfolio!D56"
                        char=get_column_letter(15)
                        ws[char+str(3)]=f"=Portfolio!D57"
                        char=get_column_letter(15)
                        ws[char+str(4)]=f"=Portfolio!D58"
                        char=get_column_letter(15)
                        ws[char+str(5)]=f"=Portfolio!D59"
                        char=get_column_letter(15)
                        ws[char+str(6)]=f"=Portfolio!D60"

                        char=get_column_letter(17)
                        ws[char+str(2)]=f"=Portfolio!D61"
                        char=get_column_letter(17)
                        ws[char+str(3)]=f"=Portfolio!D62"
                        char=get_column_letter(17)
                        ws[char+str(4)]=f"=Portfolio!D63"
                        char=get_column_letter(17)
                        ws[char+str(5)]=f"=Portfolio!D64"
                        char=get_column_letter(17)
                        ws[char+str(6)]=f"=Portfolio!D65"


                        ws['N2']='=(M2/K2)'
                        ws['N3']='=(M3/K3)'
                        ws['N4']='=(M4/K4)'
                        ws['N5']='=(M5/K5)'
                        ws['N6']='=(M6/K6)'

                        ws['P2']='=(O2/K2)'
                        ws['P3']='=(O3/K3)'
                        ws['P4']='=(O4/K4)'
                        ws['P5']='=(O5/K5)'
                        ws['P6']='=(O6/K6)'

                        ws['R2']='=(Q2/K2)'
                        ws['R3']='=(Q3/K3)'
                        ws['R4']='=(Q4/K4)'
                        ws['R5']='=(Q5/K5)'
                        ws['R6']='=(Q6/K6)'
                    if(inspectionDate[0:4]=="2013" or inspectionDate[0:4]=="2012" or inspectionDate[0:4]=="2014" or inspectionDate[0:4]=="2015" or inspectionDate[0:4]=="2016" or inspectionDate[0:4]=="2017"):


                        char=get_column_letter(4)
                        ws[char+str(1)]="Building Attribute"
                        char=get_column_letter(5)
                        ws[char+str(1)]="U_Value L-Doc (W/m^2K)"
                        char=get_column_letter(6)
                        ws[char+str(1)]="U_Value Leti (W/m^2K)"
                        char=get_column_letter(7)
                        ws[char+str(1)]="U_Value Difference (W/m^2K)"
                        char=get_column_letter(8)
                        ws[char+str(1)]="Area"
                        char=get_column_letter(9)
                        ws[char+str(1)]="U_value L-Doc Based on Area"
                        char=get_column_letter(10)
                        ws[char+str(1)]="U_value Leti Based on Area"
                        char=get_column_letter(11)
                        ws[char+str(1)]="U-value Improvement Based on Area"
                        char=get_column_letter(12)
                        ws[char+str(1)]="Savings"
                        #code for cost of retrofit per contituent
                        #code for cost of retrofit per contituent
                        char=get_column_letter(13)
                        ws[char+str(1)]="low Cost of Retrofit Constituent (m^2)"
                        char=get_column_letter(14)
                        ws[char+str(1)]="Cost of Retrofit Constituent (m^2) / Improvement based on area"
                        char=get_column_letter(15)
                        ws[char+str(1)]="Avg. Cost of Retrofit Constituent (m^2)"
                        char=get_column_letter(16)
                        ws[char+str(1)]="Cost of Retrofit Constituent (m^2) / Improvement based on area"
                        char=get_column_letter(17)
                        ws[char+str(1)]="High Cost of Retrofit Constituent (m^2)"
                        char=get_column_letter(18)
                        ws[char+str(1)]="Cost of Retrofit Constituent (m^2) / Improvement based on area"
                        char=get_column_letter(19)
                        ws[char+str(1)]="Cost of Retrofit Constituent (m^2) * area"
                        char=get_column_letter(20)
                        ws[char+str(1)]="Energy Savings (GBP)"
                        char=get_column_letter(21)
                        ws[char+str(1)]="Energy Savings (kWh)"



                        char=get_column_letter(4)
                        ws[char+str(2)]="Window"
                        char=get_column_letter(5)
                        ws[char+str(2)]=uvalue_2013["window_L"]
                        char=get_column_letter(6)
                        ws[char+str(2)]=uvalue_2013["window_leti"]
                        char=get_column_letter(7)
                        ws[char+str(2)]=uvalue_2013["window_L"]-uvalue_2013["window_leti"]
                        char=get_column_letter(8)
                        ws[char+str(2)]=(math.sqrt(floorarea)*heightt)*(0.20)
                        char=get_column_letter(9)
                        ws[char+str(2)]=(((math.sqrt(floorarea)*heightt)*(0.20)) * uvalue_2013["window_L"])
                        char=get_column_letter(10)
                        ws[char+str(2)]=((math.sqrt(floorarea)*heightt)*(0.20)) * uvalue_2013["window_leti"]
                        char=get_column_letter(11)
                        ws[char+str(2)]=(((math.sqrt(floorarea)*heightt)*(0.20)) * uvalue_2013["window_L"]) - (((math.sqrt(floorarea)*heightt)*(0.20)) * uvalue_2013["window_leti"])
                        char=get_column_letter(12)
                        ws[char+str(2)]=((((math.sqrt(floorarea)*heightt)*(0.20)) * uvalue_2013["window_L"]) - (((math.sqrt(floorarea)*heightt)*(0.20)) * uvalue_2013["window_leti"])) / ((((math.sqrt(floorarea)*heightt)*(0.20)) * uvalue_2013["window_L"]))
                        #BuildingData.json()['data'][0]['height']) THIS IS THE BUILDINGS HEIGHT FROM WHICH WE CANNOT CALCULATE THE WINDOW SIZE



                        char=get_column_letter(4)
                        ws[char+str(3)]="Door"
                        char=get_column_letter(5)
                        ws[char+str(3)]=uvalue_2013["door_L"]
                        char=get_column_letter(6)
                        ws[char+str(3)]=uvalue_2013["door_leti"]
                        char=get_column_letter(7)
                        ws[char+str(3)]=uvalue_2013["door_L"]-uvalue_2013["door_leti"]
                        char=get_column_letter(8)
                        ws[char+str(3)]=(((math.sqrt(floorarea)*heightt)*(0.20))*(0.20))
                        char=get_column_letter(9)
                        ws[char+str(3)]=(((((math.sqrt(floorarea)*heightt)*(0.20))*(0.20))) * uvalue_2013["door_L"])
                        char=get_column_letter(10)
                        ws[char+str(3)]=((((math.sqrt(floorarea)*heightt)*(0.20))*(0.20))) * uvalue_2013["door_leti"]
                        char=get_column_letter(11)
                        ws[char+str(3)]=((((((math.sqrt(floorarea)*heightt)*(0.20))*(0.20))) * uvalue_2013["door_L"]) - (((((math.sqrt(floorarea)*heightt)*(0.20))*(0.20))) * uvalue_2013["door_leti"]))
                        char=get_column_letter(12)
                        ws[char+str(3)]=((((((math.sqrt(floorarea)*heightt)*(0.20))*(0.20))) * uvalue_2013["door_L"]) - (((((math.sqrt(floorarea)*heightt)*(0.20))*(0.20))) * uvalue_2013["door_leti"])) / ((((((math.sqrt(floorarea)*heightt)*(0.20))*(0.20))) * uvalue_2013["door_L"]))

                        char=get_column_letter(4)
                        ws[char+str(4)]="Roof"
                        char=get_column_letter(5)
                        ws[char+str(4)]=uvalue_2013["roof_L"]
                        char=get_column_letter(6)
                        ws[char+str(4)]=uvalue_2013["roof_leti"]
                        char=get_column_letter(7)
                        ws[char+str(4)]=uvalue_2013["roof_L"]-uvalue_2013["roof_leti"]
                        char=get_column_letter(8)
                        ws[char+str(4)]=(floorarea)
                        char=get_column_letter(9)
                        ws[char+str(4)]=((floorarea) * uvalue_2013["roof_L"])
                        char=get_column_letter(10)
                        ws[char+str(4)]=(floorarea) * uvalue_2013["roof_leti"]
                        char=get_column_letter(11)
                        ws[char+str(4)]=(((floorarea) * uvalue_2013["roof_L"]) - ((floorarea) * uvalue_2013["roof_leti"]))
                        char=get_column_letter(12)
                        ws[char+str(4)]=(((floorarea) * uvalue_2013["roof_L"]) - ((floorarea) * uvalue_2013["roof_leti"])) / ((floorarea) * uvalue_2013["roof_L"])


                        char=get_column_letter(4)
                        ws[char+str(5)]="Wall"
                        char=get_column_letter(5)
                        ws[char+str(5)]=uvalue_2013["wall_L"]
                        char=get_column_letter(6)
                        ws[char+str(5)]=uvalue_2013["wall_leti"]
                        char=get_column_letter(7)
                        ws[char+str(5)]=uvalue_2013["wall_L"]-uvalue_2013["wall_leti"]
                        char=get_column_letter(8)
                        ws[char+str(5)]=(math.sqrt(floorarea)*heightt)*(0.80)
                        char=get_column_letter(9)
                        ws[char+str(5)]=(((math.sqrt(floorarea)*heightt)*(0.80)) * uvalue_2013["wall_L"])
                        char=get_column_letter(10)
                        ws[char+str(5)]=((math.sqrt(floorarea)*heightt)*(0.80)) * uvalue_2013["wall_leti"]
                        char=get_column_letter(11)
                        ws[char+str(5)]=((((math.sqrt(floorarea)*heightt)*(0.80)) * uvalue_2013["wall_L"]) - (((math.sqrt(floorarea)*heightt)*(0.80)) * uvalue_2013["wall_leti"]))
                        char=get_column_letter(12)
                        ws[char+str(5)]=((((math.sqrt(floorarea)*heightt)*(0.80)) * uvalue_2013["wall_L"]) - (((math.sqrt(floorarea)*heightt)*(0.80)) * uvalue_2013["wall_leti"])) / (((math.sqrt(floorarea)*heightt)*(0.80)) * uvalue_2013["wall_leti"]  )

                        char=get_column_letter(4)
                        ws[char+str(6)]="Floor"
                        char=get_column_letter(5)
                        ws[char+str(6)]=uvalue_2013["floor_L"]
                        char=get_column_letter(6)
                        ws[char+str(6)]=uvalue_2013["floor_leti"]
                        char=get_column_letter(7)
                        ws[char+str(6)]=uvalue_2013["floor_L"]-uvalue_2013["floor_leti"]
                        char=get_column_letter(8)
                        ws[char+str(6)]=(floorarea)
                        char=get_column_letter(9)
                        ws[char+str(6)]=((floorarea) * uvalue_2013["floor_L"])
                        char=get_column_letter(10)
                        ws[char+str(6)]=(floorarea) * uvalue_2013["floor_leti"]
                        char=get_column_letter(11)
                        ws[char+str(6)]=(((floorarea) * uvalue_2013["floor_L"]) - ((floorarea) * uvalue_2013["floor_leti"]))
                        char=get_column_letter(12)
                        ws[char+str(6)]= (((floorarea) * uvalue_2013["floor_L"]) - ((floorarea) * uvalue_2013["floor_leti"])) / ((floorarea) * uvalue_2013["floor_L"])

                        char=get_column_letter(8)
                        ws[char+str(7)]="Total U-Value"

                        ws['L2']='=(K2/I2)'
                        ws['L3']='=(K3/I3)'
                        ws['L4']='=(K4/I4)'
                        ws['L5']='=(K5/I5)'
                        ws['L6']='=(K6/I6)'
                        ws['L7']='=(K7/I7)'
                        ws['I7']='=SUM(I2:I6)'
                        ws['I8']=f"=UserInputs!B{propertyrow}+UserInputs!C{propertyrow}"
                        ws['J7']='=SUM(J2:J6)'
                        ws['K7']='=SUM(K2:K6)'
                        ws['L8']='=(L7)*I8'


                        #const of retrofit * area
                        ws['S2']='=(H2)*O2'
                        ws['S3']='=(H3)*O3'
                        ws['S4']='=(H4)*O4'
                        ws['S5']='=(H5)*O5'
                        ws['S6']='=(H6)*O6'

                        #TOTAL SAVINGS

                        char=get_column_letter(11)
                        ws[char+str(8)]="Total Savings"

                        #Energy Savings
                        ws['T2']='=((((K2)*(Portfolio!D50))/1000)*365*24*(Portfolio!D67))'
                        ws['T3']='=((((K3)*(Portfolio!D50))/1000)*365*24*(Portfolio!D67))'
                        ws['T4']='=((((K4)*(Portfolio!D50))/1000)*365*24*(Portfolio!D67))'
                        ws['T5']='=((((K5)*(Portfolio!D50))/1000)*365*24*(Portfolio!D67))'
                        ws['T6']='=((((K6)*(Portfolio!D50))/1000)*365*24*(Portfolio!D67))'
                        ws['T7']='=SUM(T2:T6)'


                        ws['U2']=f'=((((K2)*(Portfolio!D50))/1000)*365*24)/12'
                        ws['U3']=f'=((((K3)*(Portfolio!D50))/1000)*365*24)/12'
                        ws['U4']=f'=((((K4)*(Portfolio!D50))/1000)*365*24)/12'
                        ws['U5']=f'=((((K5)*(Portfolio!D50))/1000)*365*24)/12'
                        ws['U6']=f'=((((K6)*(Portfolio!D50))/1000)*365*24)/12'
                        ws['U7']='=SUM(U2:U6)'


                        if(add_to_portfolio==True):
                            #TO ADD TO PORTFOLIO

                            #difference u value
                            char=get_column_letter(11)
                            portfolio[char+str(propertyrow)]=uvalue_2013["window_L"]-uvalue_2013["window_leti"]
                            char=get_column_letter(10)
                            portfolio[char+str(propertyrow)]=uvalue_2013["door_L"]-uvalue_2013["door_leti"]
                            char=get_column_letter(12)
                            portfolio[char+str(propertyrow)]=uvalue_2013["roof_L"]-uvalue_2013["roof_leti"]
                            char=get_column_letter(13)
                            portfolio[char+str(propertyrow)]=uvalue_2013["wall_L"]-uvalue_2013["wall_leti"]
                            char=get_column_letter(14)
                            portfolio[char+str(propertyrow)]=uvalue_2013["floor_L"]-uvalue_2013["floor_leti"]

                            #area
                            char=get_column_letter(16)
                            portfolio[char+str(propertyrow)]=(math.sqrt(floorarea)*heightt)*(0.20)
                            char=get_column_letter(15)
                            portfolio[char+str(propertyrow)]=(((math.sqrt(floorarea)*heightt)*(0.20))*(0.20))
                            char=get_column_letter(17)
                            portfolio[char+str(propertyrow)]=(floorarea)
                            char=get_column_letter(18)
                            portfolio[char+str(propertyrow)]=(math.sqrt(floorarea)*heightt)*(0.80)
                            char=get_column_letter(19)
                            portfolio[char+str(propertyrow)]=(floorarea)

                            #improvement
                            char=get_column_letter(21)
                            portfolio[char+str(propertyrow)]=(((math.sqrt(floorarea)*heightt)*(0.20)) * uvalue_2013["window_L"]) - (((math.sqrt(floorarea)*heightt)*(0.20)) * uvalue_2013["window_leti"])
                            char=get_column_letter(20)
                            portfolio[char+str(propertyrow)]=((((((math.sqrt(floorarea)*heightt)*(0.20))*(0.20))) * uvalue_2013["door_L"]) - (((((math.sqrt(floorarea)*heightt)*(0.20))*(0.20))) * uvalue_2013["door_leti"]))
                            char=get_column_letter(22)
                            portfolio[char+str(propertyrow)]=(((floorarea) * uvalue_2013["roof_L"]) - ((floorarea) * uvalue_2013["roof_leti"]))
                            char=get_column_letter(23)
                            portfolio[char+str(propertyrow)]=((((math.sqrt(floorarea)*heightt)*(0.80)) * uvalue_2013["wall_L"]) - (((math.sqrt(floorarea)*heightt)*(0.80)) * uvalue_2013["wall_leti"]))
                            char=get_column_letter(24)
                            portfolio[char+str(propertyrow)]=(((floorarea) * uvalue_2013["floor_L"]) - ((floorarea) * uvalue_2013["floor_leti"]))
                            char=get_column_letter(25)
                            portfolio[char+str(propertyrow)]=(((((math.sqrt(floorarea)*heightt)*(0.20)) * uvalue_2013["window_L"]) - (((math.sqrt(floorarea)*heightt)*(0.20)) * uvalue_2013["window_leti"])) + ((((((math.sqrt(floorarea)*heightt)*(0.20))*(0.20))) * uvalue_2013["door_L"]) - (((((math.sqrt(floorarea)*heightt)*(0.20))*(0.20))) * uvalue_2013["door_leti"])) + (((floorarea) * uvalue_2013["roof_L"]) - ((floorarea) * uvalue_2013["roof_leti"])) + ((((math.sqrt(floorarea)*heightt)*(0.80)) * uvalue_2013["wall_L"]) - (((math.sqrt(floorarea)*heightt)*(0.80)) * uvalue_2013["wall_leti"])) + (((floorarea) * uvalue_2013["floor_L"]) - ((floorarea) * uvalue_2013["floor_leti"])))


                            #savings
                            char=get_column_letter(27)
                            portfolio[char+str(propertyrow)]=((((math.sqrt(floorarea)*heightt)*(0.20)) * uvalue_2013["window_L"]) - (((math.sqrt(floorarea)*heightt)*(0.20)) * uvalue_2013["window_leti"])) / ((((math.sqrt(floorarea)*heightt)*(0.20)) * uvalue_2013["window_L"]))
                            char=get_column_letter(26)
                            portfolio[char+str(propertyrow)]=((((((math.sqrt(floorarea)*heightt)*(0.20))*(0.20))) * uvalue_2013["door_L"]) - (((((math.sqrt(floorarea)*heightt)*(0.20))*(0.20))) * uvalue_2013["door_leti"])) / ((((((math.sqrt(floorarea)*heightt)*(0.20))*(0.20))) * uvalue_2013["door_L"]))
                            char=get_column_letter(28)
                            portfolio[char+str(propertyrow)]=(((floorarea) * uvalue_2013["roof_L"]) - ((floorarea) * uvalue_2013["roof_leti"])) / ((floorarea) * uvalue_2013["roof_L"])
                            char=get_column_letter(29)
                            portfolio[char+str(propertyrow)]=((((math.sqrt(floorarea)*heightt)*(0.80)) * uvalue_2013["wall_L"]) - (((math.sqrt(floorarea)*heightt)*(0.80)) * uvalue_2013["wall_leti"])) / (((math.sqrt(floorarea)*heightt)*(0.80)) * uvalue_2013["wall_L"]  )
                            char=get_column_letter(30)
                            portfolio[char+str(propertyrow)]=(((floorarea) * uvalue_2013["floor_L"]) - ((floorarea) * uvalue_2013["floor_leti"])) / ((floorarea) * uvalue_2013["floor_L"])
                            char=get_column_letter(31)
                            portfolio[char+str(propertyrow)]=(((((math.sqrt(floorarea)*heightt)*(0.20)) * uvalue_2013["window_L"]) - (((math.sqrt(floorarea)*heightt)*(0.20)) * uvalue_2013["window_leti"])) + ((((((math.sqrt(floorarea)*heightt)*(0.20))*(0.20))) * uvalue_2013["door_L"]) - (((((math.sqrt(floorarea)*heightt)*(0.20))*(0.20))) * uvalue_2013["door_leti"])) + (((floorarea) * uvalue_2013["roof_L"]) - ((floorarea) * uvalue_2013["roof_leti"])) + ((((math.sqrt(floorarea)*heightt)*(0.80)) * uvalue_2013["wall_L"]) - (((math.sqrt(floorarea)*heightt)*(0.80)) * uvalue_2013["wall_leti"])) + (((floorarea) * uvalue_2013["floor_L"]) - ((floorarea) * uvalue_2013["floor_leti"]))) / ((((math.sqrt(floorarea)*heightt)*(0.20)) * uvalue_2013["window_L"]) + (((((math.sqrt(floorarea)*heightt)*(0.20))*(0.20))) * uvalue_2013["door_L"]) + ((floorarea) * uvalue_2013["roof_L"]) + (((math.sqrt(floorarea)*heightt)*(0.80)) * uvalue_2013["wall_L"]) + ((floorarea) * uvalue_2013["floor_L"]))


                            #FINAL SAVINGS
                            char=get_column_letter(32)
                            portfolio[char+str(propertyrow)]='='+str(uprn)+'!L8*12'

                            #cost of retrofit per meter square for all constituents
                            char=get_column_letter(33)
                            portfolio[get_column_letter(41)+str(propertyrow)]='=SUM('+str(uprn)+'!S2:S6)'

                            up=str(uprn)

                            char=get_column_letter(51)
                            portfolio[char+str(propertyrow)]=f"={up}!L7*{buildingEmissions}"
                            char=get_column_letter(52)
                            portfolio[char+str(propertyrow)]=f"={up}!T2"
                            char=get_column_letter(53)
                            portfolio[char+str(propertyrow)]=f"={up}!T3"
                            char=get_column_letter(54)
                            portfolio[char+str(propertyrow)]=f"={up}!T4"
                            char=get_column_letter(55)
                            portfolio[char+str(propertyrow)]=f"={up}!T5"
                            char=get_column_letter(56)
                            portfolio[char+str(propertyrow)]=f"={up}!T6"
                            char=get_column_letter(57)
                            portfolio[char+str(propertyrow)]=f"={up}!U7"
                            char=get_column_letter(58)
                            portfolio[char+str(propertyrow)]=f"=INDEX({up}!U2:U6, MATCH(MAX({up}!L2:L6), {up}!L2:L6, 0))"
                            char=get_column_letter(59)
                            portfolio[char+str(propertyrow)]=f"=(INDEX({up}!H2:H6, MATCH(MAX({up}!L2:L6), {up}!L2:L6, 0))/{up}!B36)*{up}!B1"
                            char=get_column_letter(60)
                            portfolio[char+str(propertyrow)]=f"=INDEX({up}!T2:T6, MATCH(MAX({up}!L2:L6), {up}!L2:L6, 0))"
                            char=get_column_letter(61)
                            portfolio[char+str(propertyrow)]=f"={up}!I7"

                        #Code fo Constituent retrofift calculation
                        #cost of constituent outputs
                        #cost of constituent outputs
                        char=get_column_letter(13)
                        ws[char+str(2)]=f"=Portfolio!D51"
                        char=get_column_letter(13)
                        ws[char+str(3)]=f"=Portfolio!D52"
                        char=get_column_letter(13)
                        ws[char+str(4)]=f"=Portfolio!D53"
                        char=get_column_letter(13)
                        ws[char+str(5)]=f"=Portfolio!D54"
                        char=get_column_letter(13)
                        ws[char+str(6)]=f"=Portfolio!D55"

                        char=get_column_letter(15)
                        ws[char+str(2)]=f"=Portfolio!D56"
                        char=get_column_letter(15)
                        ws[char+str(3)]=f"=Portfolio!D57"
                        char=get_column_letter(15)
                        ws[char+str(4)]=f"=Portfolio!D58"
                        char=get_column_letter(15)
                        ws[char+str(5)]=f"=Portfolio!D59"
                        char=get_column_letter(15)
                        ws[char+str(6)]=f"=Portfolio!D60"

                        char=get_column_letter(17)
                        ws[char+str(2)]=f"=Portfolio!D61"
                        char=get_column_letter(17)
                        ws[char+str(3)]=f"=Portfolio!D62"
                        char=get_column_letter(17)
                        ws[char+str(4)]=f"=Portfolio!D63"
                        char=get_column_letter(17)
                        ws[char+str(5)]=f"=Portfolio!D64"
                        char=get_column_letter(17)
                        ws[char+str(6)]=f"=Portfolio!D65"


                        ws['N2']='=(M2/K2)'
                        ws['N3']='=(M3/K3)'
                        ws['N4']='=(M4/K4)'
                        ws['N5']='=(M5/K5)'
                        ws['N6']='=(M6/K6)'

                        ws['P2']='=(O2/K2)'
                        ws['P3']='=(O3/K3)'
                        ws['P4']='=(O4/K4)'
                        ws['P5']='=(O5/K5)'
                        ws['P6']='=(O6/K6)'

                        ws['R2']='=(Q2/K2)'
                        ws['R3']='=(Q3/K3)'
                        ws['R4']='=(Q4/K4)'
                        ws['R5']='=(Q5/K5)'
                        ws['R6']='=(Q6/K6)'

                    if(inspectionDate[0:4]=="2018"):


                        char=get_column_letter(4)
                        ws[char+str(1)]="Building Attribute"
                        char=get_column_letter(5)
                        ws[char+str(1)]="U_Value L-Doc (W/m^2K)"
                        char=get_column_letter(6)
                        ws[char+str(1)]="U_Value Leti (W/m^2K)"
                        char=get_column_letter(7)
                        ws[char+str(1)]="U_Value Difference (W/m^2K)"
                        char=get_column_letter(8)
                        ws[char+str(1)]="Area"
                        char=get_column_letter(9)
                        ws[char+str(1)]="U_value L-Doc Based on Area"
                        char=get_column_letter(10)
                        ws[char+str(1)]="U_value Leti Based on Area"
                        char=get_column_letter(11)
                        ws[char+str(1)]="U-value Improvement Based on Area"
                        char=get_column_letter(12)
                        ws[char+str(1)]="Savings"

                        #code for cost of retrofit per contituent
                        char=get_column_letter(13)
                        ws[char+str(1)]="low Cost of Retrofit Constituent (m^2)"
                        char=get_column_letter(14)
                        ws[char+str(1)]="Cost of Retrofit Constituent (m^2) / Improvement based on area"
                        char=get_column_letter(15)
                        ws[char+str(1)]="Avg. Cost of Retrofit Constituent (m^2)"
                        char=get_column_letter(16)
                        ws[char+str(1)]="Cost of Retrofit Constituent (m^2) / Improvement based on area"
                        char=get_column_letter(17)
                        ws[char+str(1)]="High Cost of Retrofit Constituent (m^2)"
                        char=get_column_letter(18)
                        ws[char+str(1)]="Cost of Retrofit Constituent (m^2) / Improvement based on area"
                        char=get_column_letter(19)
                        ws[char+str(1)]="Cost of Retrofit Constituent (m^2) * area"
                        char=get_column_letter(20)
                        ws[char+str(1)]="Energy Savings (GBP)"
                        char=get_column_letter(21)
                        ws[char+str(1)]="Energy Savings (kWh)"

                        char=get_column_letter(4)
                        ws[char+str(2)]="Window"
                        char=get_column_letter(5)
                        ws[char+str(2)]=uvalue_2018["window_L"]
                        char=get_column_letter(6)
                        ws[char+str(2)]=uvalue_2018["window_leti"]
                        char=get_column_letter(7)
                        ws[char+str(2)]=uvalue_2018["window_L"]-uvalue_2018["window_leti"]
                        char=get_column_letter(8)
                        ws[char+str(2)]=(math.sqrt(floorarea)*heightt)*(0.20)
                        char=get_column_letter(9)
                        ws[char+str(2)]=(((math.sqrt(floorarea)*heightt)*(0.20)) * uvalue_2018["window_L"])
                        char=get_column_letter(10)
                        ws[char+str(2)]=((math.sqrt(floorarea)*heightt)*(0.20)) * uvalue_2018["window_leti"]
                        char=get_column_letter(11)
                        ws[char+str(2)]=(((math.sqrt(floorarea)*heightt)*(0.20)) * uvalue_2018["window_L"]) - (((math.sqrt(floorarea)*heightt)*(0.20)) * uvalue_2018["window_leti"])
                        char=get_column_letter(12)
                        ws[char+str(2)]=((((math.sqrt(floorarea)*heightt)*(0.20)) * uvalue_2018["window_L"]) - (((math.sqrt(floorarea)*heightt)*(0.20)) * uvalue_2018["window_leti"])) / ((((math.sqrt(floorarea)*heightt)*(0.20)) * uvalue_2018["window_L"]))
                        #BuildingData.json()['data'][0]['height']) THIS IS THE BUILDINGS HEIGHT FROM WHICH WE CANNOT CALCULATE THE WINDOW SIZE



                        char=get_column_letter(4)
                        ws[char+str(3)]="Door"
                        char=get_column_letter(5)
                        ws[char+str(3)]=uvalue_2018["door_L"]
                        char=get_column_letter(6)
                        ws[char+str(3)]=uvalue_2018["door_leti"]
                        char=get_column_letter(7)
                        ws[char+str(3)]=uvalue_2018["door_L"]-uvalue_2018["door_leti"]
                        char=get_column_letter(8)
                        ws[char+str(3)]=(((math.sqrt(floorarea)*heightt)*(0.20))*(0.20))
                        char=get_column_letter(9)
                        ws[char+str(3)]=(((((math.sqrt(floorarea)*heightt)*(0.20))*(0.20))) * uvalue_2018["door_L"])
                        char=get_column_letter(10)
                        ws[char+str(3)]=((((math.sqrt(floorarea)*heightt)*(0.20))*(0.20))) * uvalue_2018["door_leti"]
                        char=get_column_letter(11)
                        ws[char+str(3)]=((((((math.sqrt(floorarea)*heightt)*(0.20))*(0.20))) * uvalue_2018["door_L"]) - (((((math.sqrt(floorarea)*heightt)*(0.20))*(0.20))) * uvalue_2018["door_leti"]))
                        char=get_column_letter(12)
                        ws[char+str(3)]=((((((math.sqrt(floorarea)*heightt)*(0.20))*(0.20))) * uvalue_2018["door_L"]) - (((((math.sqrt(floorarea)*heightt)*(0.20))*(0.20))) * uvalue_2018["door_leti"])) / ((((((math.sqrt(floorarea)*heightt)*(0.20))*(0.20))) * uvalue_2018["door_L"]))

                        char=get_column_letter(4)
                        ws[char+str(4)]="Roof"
                        char=get_column_letter(5)
                        ws[char+str(4)]=uvalue_2018["roof_L"]
                        char=get_column_letter(6)
                        ws[char+str(4)]=uvalue_2018["roof_leti"]
                        char=get_column_letter(7)
                        ws[char+str(4)]=uvalue_2018["roof_L"]-uvalue_2018["roof_leti"]
                        char=get_column_letter(8)
                        ws[char+str(4)]=(floorarea)
                        char=get_column_letter(9)
                        ws[char+str(4)]=((floorarea) * uvalue_2018["roof_L"])
                        char=get_column_letter(10)
                        ws[char+str(4)]=(floorarea) * uvalue_2018["roof_leti"]
                        char=get_column_letter(11)
                        ws[char+str(4)]=(((floorarea) * uvalue_2018["roof_L"]) - ((floorarea) * uvalue_2018["roof_leti"]))
                        char=get_column_letter(12)
                        ws[char+str(4)]=(((floorarea) * uvalue_2018["roof_L"]) - ((floorarea) * uvalue_2018["roof_leti"])) / ((floorarea) * uvalue_2018["roof_L"])


                        char=get_column_letter(4)
                        ws[char+str(5)]="Wall"
                        char=get_column_letter(5)
                        ws[char+str(5)]=uvalue_2018["wall_L"]
                        char=get_column_letter(6)
                        ws[char+str(5)]=uvalue_2018["wall_leti"]
                        char=get_column_letter(7)
                        ws[char+str(5)]=uvalue_2018["wall_L"]-uvalue_2018["wall_leti"]
                        char=get_column_letter(8)
                        ws[char+str(5)]=(math.sqrt(floorarea)*heightt)*(0.80)
                        char=get_column_letter(9)
                        ws[char+str(5)]=(((math.sqrt(floorarea)*heightt)*(0.80)) * uvalue_2018["wall_L"])
                        char=get_column_letter(10)
                        ws[char+str(5)]=((math.sqrt(floorarea)*heightt)*(0.80)) * uvalue_2018["wall_leti"]
                        char=get_column_letter(11)
                        ws[char+str(5)]=((((math.sqrt(floorarea)*heightt)*(0.80)) * uvalue_2018["wall_L"]) - (((math.sqrt(floorarea)*heightt)*(0.80)) * uvalue_2018["wall_leti"]))
                        char=get_column_letter(12)
                        ws[char+str(5)]=((((math.sqrt(floorarea)*heightt)*(0.80)) * uvalue_2018["wall_L"]) - (((math.sqrt(floorarea)*heightt)*(0.80)) * uvalue_2018["wall_leti"])) / (((math.sqrt(floorarea)*heightt)*(0.80)) * uvalue_2018["wall_leti"]  )

                        char=get_column_letter(4)
                        ws[char+str(6)]="Floor"
                        char=get_column_letter(5)
                        ws[char+str(6)]=uvalue_2018["floor_L"]
                        char=get_column_letter(6)
                        ws[char+str(6)]=uvalue_2018["floor_leti"]
                        char=get_column_letter(7)
                        ws[char+str(6)]=uvalue_2018["floor_L"]-uvalue_2018["floor_leti"]
                        char=get_column_letter(8)
                        ws[char+str(6)]=(floorarea)
                        char=get_column_letter(9)
                        ws[char+str(6)]=((floorarea) * uvalue_2018["floor_L"])
                        char=get_column_letter(10)
                        ws[char+str(6)]=(floorarea) * uvalue_2018["floor_leti"]
                        char=get_column_letter(11)
                        ws[char+str(6)]=(((floorarea) * uvalue_2018["floor_L"]) - ((floorarea) * uvalue_2018["floor_leti"]))
                        char=get_column_letter(12)
                        ws[char+str(6)]= (((floorarea) * uvalue_2018["floor_L"]) - ((floorarea) * uvalue_2018["floor_leti"])) / ((floorarea) * uvalue_2018["floor_L"])

                        char=get_column_letter(8)
                        ws[char+str(7)]="Total U-Value"

                        ws['L2']='=(K2/I2)'
                        ws['L3']='=(K3/I3)'
                        ws['L4']='=(K4/I4)'
                        ws['L5']='=(K5/I5)'
                        ws['L6']='=(K6/I6)'
                        ws['L7']='=(K7/I7)'
                        ws['I7']='=SUM(I2:I6)'
                        ws['I8']=f"=UserInputs!B{propertyrow}+UserInputs!C{propertyrow}"
                        ws['J7']='=SUM(J2:J6)'
                        ws['K7']='=SUM(K2:K6)'
                        ws['L8']='=(L7)*I8'

                        #const of retrofit * area
                        ws['S2']='=(H2)*O2'
                        ws['S3']='=(H3)*O3'
                        ws['S4']='=(H4)*O4'
                        ws['S5']='=(H5)*O5'
                        ws['S6']='=(H6)*O6'


                        #TOTAL SAVINGS

                        char=get_column_letter(11)
                        ws[char+str(8)]="Total Savings"

                        #Energy Savings
                        ws['T2']='=((((K2)*(Portfolio!D50))/1000)*365*24*(Portfolio!D67))'
                        ws['T3']='=((((K3)*(Portfolio!D50))/1000)*365*24*(Portfolio!D67))'
                        ws['T4']='=((((K4)*(Portfolio!D50))/1000)*365*24*(Portfolio!D67))'
                        ws['T5']='=((((K5)*(Portfolio!D50))/1000)*365*24*(Portfolio!D67))'
                        ws['T6']='=((((K6)*(Portfolio!D50))/1000)*365*24*(Portfolio!D67))'
                        ws['T7']='=SUM(T2:T6)'


                        ws['U2']=f'=((((K2)*(Portfolio!D50))/1000)*365*24)/12'
                        ws['U3']=f'=((((K3)*(Portfolio!D50))/1000)*365*24)/12'
                        ws['U4']=f'=((((K4)*(Portfolio!D50))/1000)*365*24)/12'
                        ws['U5']=f'=((((K5)*(Portfolio!D50))/1000)*365*24)/12'
                        ws['U6']=f'=((((K6)*(Portfolio!D50))/1000)*365*24)/12'
                        ws['U7']='=SUM(U2:U6)'


                        if(add_to_portfolio==True):
                            #TO ADD TO PORTFOLIO

                            #difference u value
                            char=get_column_letter(11)
                            portfolio[char+str(propertyrow)]=uvalue_2018["window_L"]-uvalue_2018["window_leti"]
                            char=get_column_letter(10)
                            portfolio[char+str(propertyrow)]=uvalue_2018["door_L"]-uvalue_2018["door_leti"]
                            char=get_column_letter(12)
                            portfolio[char+str(propertyrow)]=uvalue_2018["roof_L"]-uvalue_2018["roof_leti"]
                            char=get_column_letter(13)
                            portfolio[char+str(propertyrow)]=uvalue_2018["wall_L"]-uvalue_2018["wall_leti"]
                            char=get_column_letter(14)
                            portfolio[char+str(propertyrow)]=uvalue_2018["floor_L"]-uvalue_2018["floor_leti"]

                            #area
                            char=get_column_letter(16)
                            portfolio[char+str(propertyrow)]=(math.sqrt(floorarea)*heightt)*(0.20)
                            char=get_column_letter(15)
                            portfolio[char+str(propertyrow)]=(((math.sqrt(floorarea)*heightt)*(0.20))*(0.20))
                            char=get_column_letter(17)
                            portfolio[char+str(propertyrow)]=(floorarea)
                            char=get_column_letter(18)
                            portfolio[char+str(propertyrow)]=(math.sqrt(floorarea)*heightt)*(0.80)
                            char=get_column_letter(19)
                            portfolio[char+str(propertyrow)]=(floorarea)

                            #improvement
                            char=get_column_letter(21)
                            portfolio[char+str(propertyrow)]=(((math.sqrt(floorarea)*heightt)*(0.20)) * uvalue_2018["window_L"]) - (((math.sqrt(floorarea)*heightt)*(0.20)) * uvalue_2018["window_leti"])
                            char=get_column_letter(20)
                            portfolio[char+str(propertyrow)]=((((((math.sqrt(floorarea)*heightt)*(0.20))*(0.20))) * uvalue_2018["door_L"]) - (((((math.sqrt(floorarea)*heightt)*(0.20))*(0.20))) * uvalue_2018["door_leti"]))
                            char=get_column_letter(22)
                            portfolio[char+str(propertyrow)]=(((floorarea) * uvalue_2018["roof_L"]) - ((floorarea) * uvalue_2018["roof_leti"]))
                            char=get_column_letter(23)
                            portfolio[char+str(propertyrow)]=((((math.sqrt(floorarea)*heightt)*(0.80)) * uvalue_2018["wall_L"]) - (((math.sqrt(floorarea)*heightt)*(0.80)) * uvalue_2018["wall_leti"]))
                            char=get_column_letter(24)
                            portfolio[char+str(propertyrow)]=(((floorarea) * uvalue_2018["floor_L"]) - ((floorarea) * uvalue_2018["floor_leti"]))
                            char=get_column_letter(25)
                            portfolio[char+str(propertyrow)]=(((((math.sqrt(floorarea)*heightt)*(0.20)) * uvalue_2018["window_L"]) - (((math.sqrt(floorarea)*heightt)*(0.20)) * uvalue_2018["window_leti"])) + ((((((math.sqrt(floorarea)*heightt)*(0.20))*(0.20))) * uvalue_2018["door_L"]) - (((((math.sqrt(floorarea)*heightt)*(0.20))*(0.20))) * uvalue_2018["door_leti"])) + (((floorarea) * uvalue_2018["roof_L"]) - ((floorarea) * uvalue_2018["roof_leti"])) + ((((math.sqrt(floorarea)*heightt)*(0.80)) * uvalue_2018["wall_L"]) - (((math.sqrt(floorarea)*heightt)*(0.80)) * uvalue_2018["wall_leti"])) + (((floorarea) * uvalue_2018["floor_L"]) - ((floorarea) * uvalue_2018["floor_leti"])))


                            #savings
                            char=get_column_letter(27)
                            portfolio[char+str(propertyrow)]=((((math.sqrt(floorarea)*heightt)*(0.20)) * uvalue_2018["window_L"]) - (((math.sqrt(floorarea)*heightt)*(0.20)) * uvalue_2018["window_leti"])) / ((((math.sqrt(floorarea)*heightt)*(0.20)) * uvalue_2018["window_L"]))
                            char=get_column_letter(26)
                            portfolio[char+str(propertyrow)]=((((((math.sqrt(floorarea)*heightt)*(0.20))*(0.20))) * uvalue_2018["door_L"]) - (((((math.sqrt(floorarea)*heightt)*(0.20))*(0.20))) * uvalue_2018["door_leti"])) / ((((((math.sqrt(floorarea)*heightt)*(0.20))*(0.20))) * uvalue_2018["door_L"]))
                            char=get_column_letter(28)
                            portfolio[char+str(propertyrow)]=(((floorarea) * uvalue_2018["roof_L"]) - ((floorarea) * uvalue_2018["roof_leti"])) / ((floorarea) * uvalue_2018["roof_L"])
                            char=get_column_letter(29)
                            portfolio[char+str(propertyrow)]=((((math.sqrt(floorarea)*heightt)*(0.80)) * uvalue_2018["wall_L"]) - (((math.sqrt(floorarea)*heightt)*(0.80)) * uvalue_2018["wall_leti"])) / (((math.sqrt(floorarea)*heightt)*(0.80)) * uvalue_2018["wall_L"]  )
                            char=get_column_letter(30)
                            portfolio[char+str(propertyrow)]=(((floorarea) * uvalue_2018["floor_L"]) - ((floorarea) * uvalue_2018["floor_leti"])) / ((floorarea) * uvalue_2018["floor_L"])
                            char=get_column_letter(31)
                            portfolio[char+str(propertyrow)]=(((((math.sqrt(floorarea)*heightt)*(0.20)) * uvalue_2018["window_L"]) - (((math.sqrt(floorarea)*heightt)*(0.20)) * uvalue_2018["window_leti"])) + ((((((math.sqrt(floorarea)*heightt)*(0.20))*(0.20))) * uvalue_2018["door_L"]) - (((((math.sqrt(floorarea)*heightt)*(0.20))*(0.20))) * uvalue_2018["door_leti"])) + (((floorarea) * uvalue_2018["roof_L"]) - ((floorarea) * uvalue_2018["roof_leti"])) + ((((math.sqrt(floorarea)*heightt)*(0.80)) * uvalue_2018["wall_L"]) - (((math.sqrt(floorarea)*heightt)*(0.80)) * uvalue_2018["wall_leti"])) + (((floorarea) * uvalue_2018["floor_L"]) - ((floorarea) * uvalue_2018["floor_leti"]))) / ((((math.sqrt(floorarea)*heightt)*(0.20)) * uvalue_2018["window_L"]) + (((((math.sqrt(floorarea)*heightt)*(0.20))*(0.20))) * uvalue_2018["door_L"]) + ((floorarea) * uvalue_2018["roof_L"]) + (((math.sqrt(floorarea)*heightt)*(0.80)) * uvalue_2018["wall_L"]) + ((floorarea) * uvalue_2018["floor_L"]))


                            #FINAL SAVINGS
                            char=get_column_letter(32)
                            portfolio[char+str(propertyrow)]='='+str(uprn)+'!L8*12'

                            #cost of retrofit per meter square for all constituents
                            char=get_column_letter(33)
                            portfolio[get_column_letter(41)+str(propertyrow)]='=SUM('+str(uprn)+'!S2:S6)'

                            up=str(uprn)

                            char=get_column_letter(51)
                            portfolio[char+str(propertyrow)]=f"={up}!L7*{buildingEmissions}"
                            char=get_column_letter(52)
                            portfolio[char+str(propertyrow)]=f"={up}!T2"
                            char=get_column_letter(53)
                            portfolio[char+str(propertyrow)]=f"={up}!T3"
                            char=get_column_letter(54)
                            portfolio[char+str(propertyrow)]=f"={up}!T4"
                            char=get_column_letter(55)
                            portfolio[char+str(propertyrow)]=f"={up}!T5"
                            char=get_column_letter(56)
                            portfolio[char+str(propertyrow)]=f"={up}!T6"
                            char=get_column_letter(57)
                            portfolio[char+str(propertyrow)]=f"={up}!U7"
                            char=get_column_letter(58)
                            portfolio[char+str(propertyrow)]=f"=INDEX({up}!U2:U6, MATCH(MAX({up}!L2:L6), {up}!L2:L6, 0))"
                            char=get_column_letter(59)
                            portfolio[char+str(propertyrow)]=f"=(INDEX({up}!H2:H6, MATCH(MAX({up}!L2:L6), {up}!L2:L6, 0))/{up}!B36)*{up}!B1"
                            char=get_column_letter(60)
                            portfolio[char+str(propertyrow)]=f"=INDEX({up}!T2:T6, MATCH(MAX({up}!L2:L6), {up}!L2:L6, 0))"
                            char=get_column_letter(61)
                            portfolio[char+str(propertyrow)]=f"={up}!I7"

                        #Code fo Constituent retrofift calculation
                        #cost of constituent outputs
                        #cost of constituent outputs
                        char=get_column_letter(13)
                        ws[char+str(2)]=f"=Portfolio!D51"
                        char=get_column_letter(13)
                        ws[char+str(3)]=f"=Portfolio!D52"
                        char=get_column_letter(13)
                        ws[char+str(4)]=f"=Portfolio!D53"
                        char=get_column_letter(13)
                        ws[char+str(5)]=f"=Portfolio!D54"
                        char=get_column_letter(13)
                        ws[char+str(6)]=f"=Portfolio!D55"

                        char=get_column_letter(15)
                        ws[char+str(2)]=f"=Portfolio!D56"
                        char=get_column_letter(15)
                        ws[char+str(3)]=f"=Portfolio!D57"
                        char=get_column_letter(15)
                        ws[char+str(4)]=f"=Portfolio!D58"
                        char=get_column_letter(15)
                        ws[char+str(5)]=f"=Portfolio!D59"
                        char=get_column_letter(15)
                        ws[char+str(6)]=f"=Portfolio!D60"

                        char=get_column_letter(17)
                        ws[char+str(2)]=f"=Portfolio!D61"
                        char=get_column_letter(17)
                        ws[char+str(3)]=f"=Portfolio!D62"
                        char=get_column_letter(17)
                        ws[char+str(4)]=f"=Portfolio!D63"
                        char=get_column_letter(17)
                        ws[char+str(5)]=f"=Portfolio!D64"
                        char=get_column_letter(17)
                        ws[char+str(6)]=f"=Portfolio!D65"


                        ws['N2']='=(M2/K2)'
                        ws['N3']='=(M3/K3)'
                        ws['N4']='=(M4/K4)'
                        ws['N5']='=(M5/K5)'
                        ws['N6']='=(M6/K6)'

                        ws['P2']='=(O2/K2)'
                        ws['P3']='=(O3/K3)'
                        ws['P4']='=(O4/K4)'
                        ws['P5']='=(O5/K5)'
                        ws['P6']='=(O6/K6)'

                        ws['R2']='=(Q2/K2)'
                        ws['R3']='=(Q3/K3)'
                        ws['R4']='=(Q4/K4)'
                        ws['R5']='=(Q5/K5)'
                        ws['R6']='=(Q6/K6)'

                    if(inspectionDate[0:4]=="2019"):


                        char=get_column_letter(4)
                        ws[char+str(1)]="Building Attribute"
                        char=get_column_letter(5)
                        ws[char+str(1)]="U_Value L-Doc (W/m^2K)"
                        char=get_column_letter(6)
                        ws[char+str(1)]="U_Value Leti (W/m^2K)"
                        char=get_column_letter(7)
                        ws[char+str(1)]="U_Value Difference (W/m^2K)"
                        char=get_column_letter(8)
                        ws[char+str(1)]="Area"
                        char=get_column_letter(9)
                        ws[char+str(1)]="U_value L-Doc Based on Area"
                        char=get_column_letter(10)
                        ws[char+str(1)]="U_value Leti Based on Area"
                        char=get_column_letter(11)
                        ws[char+str(1)]="U-value Improvement Based on Area"
                        char=get_column_letter(12)
                        ws[char+str(1)]="Savings"
                        #code for cost of retrofit per contituent
                        char=get_column_letter(13)
                        ws[char+str(1)]="low Cost of Retrofit Constituent (m^2)"
                        char=get_column_letter(14)
                        ws[char+str(1)]="Cost of Retrofit Constituent (m^2) / Improvement based on area"
                        char=get_column_letter(15)
                        ws[char+str(1)]="Avg. Cost of Retrofit Constituent (m^2)"
                        char=get_column_letter(16)
                        ws[char+str(1)]="Cost of Retrofit Constituent (m^2) / Improvement based on area"
                        char=get_column_letter(17)
                        ws[char+str(1)]="High Cost of Retrofit Constituent (m^2)"
                        char=get_column_letter(18)
                        ws[char+str(1)]="Cost of Retrofit Constituent (m^2) / Improvement based on area"
                        char=get_column_letter(19)
                        ws[char+str(1)]="Cost of Retrofit Constituent (m^2) * area"
                        char=get_column_letter(20)
                        ws[char+str(1)]="Energy Savings (GBP)"
                        char=get_column_letter(21)
                        ws[char+str(1)]="Energy Savings (kWh)"



                        char=get_column_letter(4)
                        ws[char+str(2)]="Window"
                        char=get_column_letter(5)
                        ws[char+str(2)]=uvalue_2019["window_L"]
                        char=get_column_letter(6)
                        ws[char+str(2)]=uvalue_2019["window_leti"]
                        char=get_column_letter(7)
                        ws[char+str(2)]=uvalue_2019["window_L"]-uvalue_2019["window_leti"]
                        char=get_column_letter(8)
                        ws[char+str(2)]=(math.sqrt(floorarea)*heightt)*(0.20)
                        char=get_column_letter(9)
                        ws[char+str(2)]=(((math.sqrt(floorarea)*heightt)*(0.20)) * uvalue_2019["window_L"])
                        char=get_column_letter(10)
                        ws[char+str(2)]=((math.sqrt(floorarea)*heightt)*(0.20)) * uvalue_2019["window_leti"]
                        char=get_column_letter(11)
                        ws[char+str(2)]=(((math.sqrt(floorarea)*heightt)*(0.20)) * uvalue_2019["window_L"]) - (((math.sqrt(floorarea)*heightt)*(0.20)) * uvalue_2019["window_leti"])
                        char=get_column_letter(12)
                        ws[char+str(2)]=((((math.sqrt(floorarea)*heightt)*(0.20)) * uvalue_2019["window_L"]) - (((math.sqrt(floorarea)*heightt)*(0.20)) * uvalue_2019["window_leti"])) / ((((math.sqrt(floorarea)*heightt)*(0.20)) * uvalue_2019["window_L"]))
                        #BuildingData.json()['data'][0]['height']) THIS IS THE BUILDINGS HEIGHT FROM WHICH WE CANNOT CALCULATE THE WINDOW SIZE



                        char=get_column_letter(4)
                        ws[char+str(3)]="Door"
                        char=get_column_letter(5)
                        ws[char+str(3)]=uvalue_2019["door_L"]
                        char=get_column_letter(6)
                        ws[char+str(3)]=uvalue_2019["door_leti"]
                        char=get_column_letter(7)
                        ws[char+str(3)]=uvalue_2019["door_L"]-uvalue_2019["door_leti"]
                        char=get_column_letter(8)
                        ws[char+str(3)]=(((math.sqrt(floorarea)*heightt)*(0.20))*(0.20))
                        char=get_column_letter(9)
                        ws[char+str(3)]=(((((math.sqrt(floorarea)*heightt)*(0.20))*(0.20))) * uvalue_2019["door_L"])
                        char=get_column_letter(10)
                        ws[char+str(3)]=((((math.sqrt(floorarea)*heightt)*(0.20))*(0.20))) * uvalue_2019["door_leti"]
                        char=get_column_letter(11)
                        ws[char+str(3)]=((((((math.sqrt(floorarea)*heightt)*(0.20))*(0.20))) * uvalue_2019["door_L"]) - (((((math.sqrt(floorarea)*heightt)*(0.20))*(0.20))) * uvalue_2019["door_leti"]))
                        char=get_column_letter(12)
                        ws[char+str(3)]=((((((math.sqrt(floorarea)*heightt)*(0.20))*(0.20))) * uvalue_2019["door_L"]) - (((((math.sqrt(floorarea)*heightt)*(0.20))*(0.20))) * uvalue_2019["door_leti"])) / ((((((math.sqrt(floorarea)*heightt)*(0.20))*(0.20))) * uvalue_2019["door_L"]))

                        char=get_column_letter(4)
                        ws[char+str(4)]="Roof"
                        char=get_column_letter(5)
                        ws[char+str(4)]=uvalue_2019["roof_L"]
                        char=get_column_letter(6)
                        ws[char+str(4)]=uvalue_2019["roof_leti"]
                        char=get_column_letter(7)
                        ws[char+str(4)]=uvalue_2019["roof_L"]-uvalue_2019["roof_leti"]
                        char=get_column_letter(8)
                        ws[char+str(4)]=(floorarea)
                        char=get_column_letter(9)
                        ws[char+str(4)]=((floorarea) * uvalue_2019["roof_L"])
                        char=get_column_letter(10)
                        ws[char+str(4)]=(floorarea) * uvalue_2019["roof_leti"]
                        char=get_column_letter(11)
                        ws[char+str(4)]=(((floorarea) * uvalue_2019["roof_L"]) - ((floorarea) * uvalue_2019["roof_leti"]))
                        char=get_column_letter(12)
                        ws[char+str(4)]=(((floorarea) * uvalue_2019["roof_L"]) - ((floorarea) * uvalue_2019["roof_leti"])) / ((floorarea) * uvalue_2019["roof_L"])


                        char=get_column_letter(4)
                        ws[char+str(5)]="Wall"
                        char=get_column_letter(5)
                        ws[char+str(5)]=uvalue_2019["wall_L"]
                        char=get_column_letter(6)
                        ws[char+str(5)]=uvalue_2019["wall_leti"]
                        char=get_column_letter(7)
                        ws[char+str(5)]=uvalue_2019["wall_L"]-uvalue_2019["wall_leti"]
                        char=get_column_letter(8)
                        ws[char+str(5)]=(math.sqrt(floorarea)*heightt)*(0.80)
                        char=get_column_letter(9)
                        ws[char+str(5)]=(((math.sqrt(floorarea)*heightt)*(0.80)) * uvalue_2019["wall_L"])
                        char=get_column_letter(10)
                        ws[char+str(5)]=((math.sqrt(floorarea)*heightt)*(0.80)) * uvalue_2019["wall_leti"]
                        char=get_column_letter(11)
                        ws[char+str(5)]=((((math.sqrt(floorarea)*heightt)*(0.80)) * uvalue_2019["wall_L"]) - (((math.sqrt(floorarea)*heightt)*(0.80)) * uvalue_2019["wall_leti"]))
                        char=get_column_letter(12)
                        ws[char+str(5)]=((((math.sqrt(floorarea)*heightt)*(0.80)) * uvalue_2019["wall_L"]) - (((math.sqrt(floorarea)*heightt)*(0.80)) * uvalue_2019["wall_leti"])) / (((math.sqrt(floorarea)*heightt)*(0.80)) * uvalue_2019["wall_leti"]  )

                        char=get_column_letter(4)
                        ws[char+str(6)]="Floor"
                        char=get_column_letter(5)
                        ws[char+str(6)]=uvalue_2019["floor_L"]
                        char=get_column_letter(6)
                        ws[char+str(6)]=uvalue_2019["floor_leti"]
                        char=get_column_letter(7)
                        ws[char+str(6)]=uvalue_2019["floor_L"]-uvalue_2019["floor_leti"]
                        char=get_column_letter(8)
                        ws[char+str(6)]=(floorarea)
                        char=get_column_letter(9)
                        ws[char+str(6)]=((floorarea) * uvalue_2019["floor_L"])
                        char=get_column_letter(10)
                        ws[char+str(6)]=(floorarea) * uvalue_2019["floor_leti"]
                        char=get_column_letter(11)
                        ws[char+str(6)]=(((floorarea) * uvalue_2019["floor_L"]) - ((floorarea) * uvalue_2019["floor_leti"]))
                        char=get_column_letter(12)
                        ws[char+str(6)]= (((floorarea) * uvalue_2019["floor_L"]) - ((floorarea) * uvalue_2019["floor_leti"])) / ((floorarea) * uvalue_2019["floor_L"])

                        char=get_column_letter(8)
                        ws[char+str(7)]="Total U-Value"

                        ws['L2']='=(K2/I2)'
                        ws['L3']='=(K3/I3)'
                        ws['L4']='=(K4/I4)'
                        ws['L5']='=(K5/I5)'
                        ws['L6']='=(K6/I6)'
                        ws['L7']='=(K7/I7)'
                        ws['I7']='=SUM(I2:I6)'
                        ws['I8']=f"=UserInputs!B{propertyrow}+UserInputs!C{propertyrow}"
                        ws['J7']='=SUM(J2:J6)'
                        ws['K7']='=SUM(K2:K6)'
                        ws['L8']='=(L7)*I8'


                        #const of retrofit * area
                        ws['S2']='=(H2)*O2'
                        ws['S3']='=(H3)*O3'
                        ws['S4']='=(H4)*O4'
                        ws['S5']='=(H5)*O5'
                        ws['S6']='=(H6)*O6'

                        #TOTAL SAVINGS

                        char=get_column_letter(11)
                        ws[char+str(8)]="Total Savings"

                        #Energy Savings
                        ws['T2']='=((((K2)*(Portfolio!D50))/1000)*365*24*(Portfolio!D67))'
                        ws['T3']='=((((K3)*(Portfolio!D50))/1000)*365*24*(Portfolio!D67))'
                        ws['T4']='=((((K4)*(Portfolio!D50))/1000)*365*24*(Portfolio!D67))'
                        ws['T5']='=((((K5)*(Portfolio!D50))/1000)*365*24*(Portfolio!D67))'
                        ws['T6']='=((((K6)*(Portfolio!D50))/1000)*365*24*(Portfolio!D67))'
                        ws['T7']='=SUM(T2:T6)'


                        ws['U2']=f'=((((K2)*(Portfolio!D50))/1000)*365*24)/12'
                        ws['U3']=f'=((((K3)*(Portfolio!D50))/1000)*365*24)/12'
                        ws['U4']=f'=((((K4)*(Portfolio!D50))/1000)*365*24)/12'
                        ws['U5']=f'=((((K5)*(Portfolio!D50))/1000)*365*24)/12'
                        ws['U6']=f'=((((K6)*(Portfolio!D50))/1000)*365*24)/12'
                        ws['U7']='=SUM(U2:U6)'



                        if(add_to_portfolio==True):
                            #TO ADD TO PORTFOLIO

                            #difference u value
                            char=get_column_letter(11)
                            portfolio[char+str(propertyrow)]=uvalue_2019["window_L"]-uvalue_2019["window_leti"]
                            char=get_column_letter(10)
                            portfolio[char+str(propertyrow)]=uvalue_2019["door_L"]-uvalue_2019["door_leti"]
                            char=get_column_letter(12)
                            portfolio[char+str(propertyrow)]=uvalue_2019["roof_L"]-uvalue_2019["roof_leti"]
                            char=get_column_letter(13)
                            portfolio[char+str(propertyrow)]=uvalue_2019["wall_L"]-uvalue_2019["wall_leti"]
                            char=get_column_letter(14)
                            portfolio[char+str(propertyrow)]=uvalue_2019["floor_L"]-uvalue_2019["floor_leti"]

                            #area
                            char=get_column_letter(16)
                            portfolio[char+str(propertyrow)]=(math.sqrt(floorarea)*heightt)*(0.20)
                            char=get_column_letter(15)
                            portfolio[char+str(propertyrow)]=(((math.sqrt(floorarea)*heightt)*(0.20))*(0.20))
                            char=get_column_letter(17)
                            portfolio[char+str(propertyrow)]=(floorarea)
                            char=get_column_letter(18)
                            portfolio[char+str(propertyrow)]=(math.sqrt(floorarea)*heightt)*(0.80)
                            char=get_column_letter(19)
                            portfolio[char+str(propertyrow)]=(floorarea)

                            #improvement
                            char=get_column_letter(21)
                            portfolio[char+str(propertyrow)]=(((math.sqrt(floorarea)*heightt)*(0.20)) * uvalue_2019["window_L"]) - (((math.sqrt(floorarea)*heightt)*(0.20)) * uvalue_2019["window_leti"])
                            char=get_column_letter(20)
                            portfolio[char+str(propertyrow)]=((((((math.sqrt(floorarea)*heightt)*(0.20))*(0.20))) * uvalue_2019["door_L"]) - (((((math.sqrt(floorarea)*heightt)*(0.20))*(0.20))) * uvalue_2019["door_leti"]))
                            char=get_column_letter(22)
                            portfolio[char+str(propertyrow)]=(((floorarea) * uvalue_2019["roof_L"]) - ((floorarea) * uvalue_2019["roof_leti"]))
                            char=get_column_letter(23)
                            portfolio[char+str(propertyrow)]=((((math.sqrt(floorarea)*heightt)*(0.80)) * uvalue_2019["wall_L"]) - (((math.sqrt(floorarea)*heightt)*(0.80)) * uvalue_2019["wall_leti"]))
                            char=get_column_letter(24)
                            portfolio[char+str(propertyrow)]=(((floorarea) * uvalue_2019["floor_L"]) - ((floorarea) * uvalue_2019["floor_leti"]))
                            char=get_column_letter(25)
                            portfolio[char+str(propertyrow)]=(((((math.sqrt(floorarea)*heightt)*(0.20)) * uvalue_2019["window_L"]) - (((math.sqrt(floorarea)*heightt)*(0.20)) * uvalue_2019["window_leti"])) + ((((((math.sqrt(floorarea)*heightt)*(0.20))*(0.20))) * uvalue_2019["door_L"]) - (((((math.sqrt(floorarea)*heightt)*(0.20))*(0.20))) * uvalue_2019["door_leti"])) + (((floorarea) * uvalue_2019["roof_L"]) - ((floorarea) * uvalue_2019["roof_leti"])) + ((((math.sqrt(floorarea)*heightt)*(0.80)) * uvalue_2019["wall_L"]) - (((math.sqrt(floorarea)*heightt)*(0.80)) * uvalue_2019["wall_leti"])) + (((floorarea) * uvalue_2019["floor_L"]) - ((floorarea) * uvalue_2019["floor_leti"])))


                            #savings
                            char=get_column_letter(27)
                            portfolio[char+str(propertyrow)]=((((math.sqrt(floorarea)*heightt)*(0.20)) * uvalue_2019["window_L"]) - (((math.sqrt(floorarea)*heightt)*(0.20)) * uvalue_2019["window_leti"])) / ((((math.sqrt(floorarea)*heightt)*(0.20)) * uvalue_2019["window_L"]))
                            char=get_column_letter(26)
                            portfolio[char+str(propertyrow)]=((((((math.sqrt(floorarea)*heightt)*(0.20))*(0.20))) * uvalue_2019["door_L"]) - (((((math.sqrt(floorarea)*heightt)*(0.20))*(0.20))) * uvalue_2019["door_leti"])) / ((((((math.sqrt(floorarea)*heightt)*(0.20))*(0.20))) * uvalue_2019["door_L"]))
                            char=get_column_letter(28)
                            portfolio[char+str(propertyrow)]=(((floorarea) * uvalue_2019["roof_L"]) - ((floorarea) * uvalue_2019["roof_leti"])) / ((floorarea) * uvalue_2019["roof_L"])
                            char=get_column_letter(29)
                            portfolio[char+str(propertyrow)]=((((math.sqrt(floorarea)*heightt)*(0.80)) * uvalue_2019["wall_L"]) - (((math.sqrt(floorarea)*heightt)*(0.80)) * uvalue_2019["wall_leti"])) / (((math.sqrt(floorarea)*heightt)*(0.80)) * uvalue_2019["wall_L"]  )
                            char=get_column_letter(30)
                            portfolio[char+str(propertyrow)]=(((floorarea) * uvalue_2019["floor_L"]) - ((floorarea) * uvalue_2019["floor_leti"])) / ((floorarea) * uvalue_2019["floor_L"])
                            char=get_column_letter(31)
                            portfolio[char+str(propertyrow)]=(((((math.sqrt(floorarea)*heightt)*(0.20)) * uvalue_2019["window_L"]) - (((math.sqrt(floorarea)*heightt)*(0.20)) * uvalue_2019["window_leti"])) + ((((((math.sqrt(floorarea)*heightt)*(0.20))*(0.20))) * uvalue_2019["door_L"]) - (((((math.sqrt(floorarea)*heightt)*(0.20))*(0.20))) * uvalue_2019["door_leti"])) + (((floorarea) * uvalue_2019["roof_L"]) - ((floorarea) * uvalue_2019["roof_leti"])) + ((((math.sqrt(floorarea)*heightt)*(0.80)) * uvalue_2019["wall_L"]) - (((math.sqrt(floorarea)*heightt)*(0.80)) * uvalue_2019["wall_leti"])) + (((floorarea) * uvalue_2019["floor_L"]) - ((floorarea) * uvalue_2019["floor_leti"]))) / ((((math.sqrt(floorarea)*heightt)*(0.20)) * uvalue_2019["window_L"]) + (((((math.sqrt(floorarea)*heightt)*(0.20))*(0.20))) * uvalue_2019["door_L"]) + ((floorarea) * uvalue_2019["roof_L"]) + (((math.sqrt(floorarea)*heightt)*(0.80)) * uvalue_2019["wall_L"]) + ((floorarea) * uvalue_2019["floor_L"]))

                            #FINAL SAVINGS
                            char=get_column_letter(32)
                            portfolio[char+str(propertyrow)]='='+str(uprn)+'!L8*12'

                            #cost of retrofit per meter square for all constituents
                            char=get_column_letter(33)
                            portfolio[get_column_letter(41)+str(propertyrow)]='=SUM('+str(uprn)+'!S2:S6)'

                            up=str(uprn)

                            char=get_column_letter(51)
                            portfolio[char+str(propertyrow)]=f"={up}!L7*{buildingEmissions}"
                            char=get_column_letter(52)
                            portfolio[char+str(propertyrow)]=f"={up}!T2"
                            char=get_column_letter(53)
                            portfolio[char+str(propertyrow)]=f"={up}!T3"
                            char=get_column_letter(54)
                            portfolio[char+str(propertyrow)]=f"={up}!T4"
                            char=get_column_letter(55)
                            portfolio[char+str(propertyrow)]=f"={up}!T5"
                            char=get_column_letter(56)
                            portfolio[char+str(propertyrow)]=f"={up}!T6"
                            char=get_column_letter(57)
                            portfolio[char+str(propertyrow)]=f"={up}!U7"
                            char=get_column_letter(58)
                            portfolio[char+str(propertyrow)]=f"=INDEX({up}!U2:U6, MATCH(MAX({up}!L2:L6), {up}!L2:L6, 0))"
                            char=get_column_letter(59)
                            portfolio[char+str(propertyrow)]=f"=(INDEX({up}!H2:H6, MATCH(MAX({up}!L2:L6), {up}!L2:L6, 0))/{up}!B36)*{up}!B1"
                            char=get_column_letter(60)
                            portfolio[char+str(propertyrow)]=f"=INDEX({up}!T2:T6, MATCH(MAX({up}!L2:L6), {up}!L2:L6, 0))"
                            char=get_column_letter(61)
                            portfolio[char+str(propertyrow)]=f"={up}!I7"

                        #cost of constituent outputs
                        #cost of constituent outputs
                        char=get_column_letter(13)
                        ws[char+str(2)]=f"=Portfolio!D51"
                        char=get_column_letter(13)
                        ws[char+str(3)]=f"=Portfolio!D52"
                        char=get_column_letter(13)
                        ws[char+str(4)]=f"=Portfolio!D53"
                        char=get_column_letter(13)
                        ws[char+str(5)]=f"=Portfolio!D54"
                        char=get_column_letter(13)
                        ws[char+str(6)]=f"=Portfolio!D55"

                        char=get_column_letter(15)
                        ws[char+str(2)]=f"=Portfolio!D56"
                        char=get_column_letter(15)
                        ws[char+str(3)]=f"=Portfolio!D57"
                        char=get_column_letter(15)
                        ws[char+str(4)]=f"=Portfolio!D58"
                        char=get_column_letter(15)
                        ws[char+str(5)]=f"=Portfolio!D59"
                        char=get_column_letter(15)
                        ws[char+str(6)]=f"=Portfolio!D60"

                        char=get_column_letter(17)
                        ws[char+str(2)]=f"=Portfolio!D61"
                        char=get_column_letter(17)
                        ws[char+str(3)]=f"=Portfolio!D62"
                        char=get_column_letter(17)
                        ws[char+str(4)]=f"=Portfolio!D63"
                        char=get_column_letter(17)
                        ws[char+str(5)]=f"=Portfolio!D64"
                        char=get_column_letter(17)
                        ws[char+str(6)]=f"=Portfolio!D65"


                        ws['N2']='=(M2/K2)'
                        ws['N3']='=(M3/K3)'
                        ws['N4']='=(M4/K4)'
                        ws['N5']='=(M5/K5)'
                        ws['N6']='=(M6/K6)'

                        ws['P2']='=(O2/K2)'
                        ws['P3']='=(O3/K3)'
                        ws['P4']='=(O4/K4)'
                        ws['P5']='=(O5/K5)'
                        ws['P6']='=(O6/K6)'

                        ws['R2']='=(Q2/K2)'
                        ws['R3']='=(Q3/K3)'
                        ws['R4']='=(Q4/K4)'
                        ws['R5']='=(Q5/K5)'
                        ws['R6']='=(Q6/K6)'
                    #print(BuildingData.json()['data'])

                    if(add_to_portfolio==True):

                        #COST OF RETROFITTING
                        char=get_column_letter(4)
                        portfolio[char+str(43)]='=SUM(AO2:AO'+str(properties+1)+')'



                        propertyy=str(propertyrow)
                        char=get_column_letter(33)
                        portfolio[char+str(propertyrow)]='=(0.84-D42)*AF'+str(propertyrow)
                        char=get_column_letter(34)
                        portfolio[char+str(propertyrow)]='=AG'+str(propertyrow)+'/12'

                        char=get_column_letter(35)
                        portfolio[char+str(propertyrow)]=f"=AO{propertyy}*D41"
                        char=get_column_letter(36)

                        portfolio[char+str(propertyrow)]=f"=D42*AF{propertyy}"

                        char=get_column_letter(37)
                        portfolio[char+str(propertyrow)]='=AI'+str(propertyrow)+'*AJ'+str(propertyrow)

                        char=get_column_letter(38)
                        portfolio[char+str(propertyrow)]=f"=AF{propertyy}-(AF{propertyy}*D47)-(((1-D47)*AF{propertyy})*D42)"


                        char=get_column_letter(39)
                        portfolio[char+str(propertyrow)]=f"=AF{propertyy}*D47"

                        char=get_column_letter(40)
                        portfolio[char+str(propertyrow)]=f"=AL{propertyy}*UserInputs!C{propertyy}"

                        #ROI
                        up=str(uprn)
                        char=get_column_letter(42)
                        portfolio[char+str(1)]="High ROI Savings"
                        char=get_column_letter(43)
                        portfolio[char+str(1)]="High ROI savings passed to landlord"
                        char=get_column_letter(44)
                        portfolio[char+str(1)]="High ROI monthly tenant fee"
                        char=get_column_letter(45)
                        portfolio[char+str(1)]="Referral fee on high ROI components"
                        char=get_column_letter(46)
                        portfolio[char+str(1)]="UpGreen Annual income on high ROI components"
                        char=get_column_letter(47)
                        portfolio[char+str(1)]="High ROI savings passed to tenants"
                        char=get_column_letter(48)
                        portfolio[char+str(1)]="High ROI savings passed back for remainder of lease"
                        char=get_column_letter(49)
                        portfolio[char+str(1)]="Cost of retrofit for prioritized components"
                        char=get_column_letter(50)
                        portfolio[char+str(1)]="Constituent to Focus on"
                        #AP
                        char=get_column_letter(42)
                        portfolio[char+str(propertyrow)]=f"=MAX(Z{propertyy}:AD{propertyy})*AF{propertyy}"
                        #AQ
                        char=get_column_letter(43)
                        portfolio[char+str(propertyrow)]=f"=AP{propertyy}*(1-D47)"
                        #AR
                        char=get_column_letter(44)
                        portfolio[char+str(propertyrow)]=f"=AQ{propertyy}/12"
                        #AS
                        char=get_column_letter(45)
                        portfolio[char+str(propertyrow)]=f"=AW{propertyy}*D41"
                        #AT
                        char=get_column_letter(46)
                        portfolio[char+str(propertyrow)]=f"=AS{propertyy}+D42"
                        #AU
                        char=get_column_letter(47)
                        portfolio[char+str(propertyrow)]=f"=AP{propertyy}*D47"
                        #AV
                        char=get_column_letter(48)
                        portfolio[char+str(propertyrow)]=f"=AQ{propertyy}*UserInputs!C{propertyy}"
                        #AW
                        char=get_column_letter(49)
                        portfolio[char+str(propertyrow)]=f"=INDEX({up}!S2:S6, MATCH(MAX({up}!L2:L6), {up}!L2:L6, 0))"
                        #AX
                        char=get_column_letter(50)
                        portfolio[char+str(propertyrow)]=f"=INDEX({up}!D2:D6, MATCH(MAX({up}!L2:L6), {up}!L2:L6, 0))"
                    print(EPCData.json()['data'][0]['inspectionDate'])



            EPCDetails=requests.get(full_url, headers=headers)
            print(len(EPCDetails.json()['rows']))



            epcstr=str(EPCDetails.content)
            print(len(epcstr))


            uprns=[]
            if(len(epcstr)>3):

                for i in range(len(EPCDetails.json()['rows'])):
                    print(i)
                    print(EPCDetails.json()['rows'][i])

                    uprn=EPCDetails.json()['rows'][i]['uprn']
                    lmk=EPCDetails.json()['rows'][i]['lmk-key']

                    lmkkeyflag=False
                    if(uprn==''):
                        uprn=lmk
                        uprn=uprn[len(str(uprn))-10:len(str(uprn))]


                    if(uprn not in uprns):
                            uprns.append(uprn)
                    else:
                            uprn=lmk
                            uprn=uprn[len(str(uprn))-10:len(str(uprn))]


                    wb.create_sheet(str(uprn))
                    properties=properties+1
                    ws=wb[str(uprn)]
                    ws.title=str(uprn)
                    floor=False

                    #REMOVING REPITITIONS


                    inspectionDate=EPCDetails.json()['rows'][i]['inspection-date']
                    add_to_portfolio=True

                    add=EPCDetails.json()['rows'][i]['address']
                    add=re.sub('\W+','', add)
                    add=add.lower()
                    print(add)
                    propertyrow=properties+0
                    for j in range(2,properties,1):
                        add2=str(portfolio[get_column_letter(4)+str(j)].value)
                        add2=re.sub('\W+','', add2)
                        add2=add2.lower()



                        if(add2 in add or add in add2):
                                    print("REPEATED:",add)
                                    firstdate=portfolio[get_column_letter(6)+str(j)].value
                                    seconddate=inspectionDate

                                    firstdate=int(firstdate[0:4])
                                    seconddate=int(seconddate[0:4])

                                    if(firstdate>seconddate):
                                        add_to_portfolio=False
                                        properties=properties-1
                                    else:
                                        add_to_portfolio=True
                                        properties=properties-1
                                        propertyrow=j

                                    break


                    for attribute,value in (EPCDetails.json()['rows'][i].items()):
                        #print("Attribute: ",attribute)
                        if(attribute=="floor-area"):
                            floorarea=float(value)
                            floor=True
                            if(add_to_portfolio==True):
                                char=get_column_letter(7)
                                portfolio[char+str(propertyrow)]=value
                        if(attribute=="address1"):
                            if(add_to_portfolio==True):
                                char=get_column_letter(3)
                                portfolio[char+str(propertyrow)]=value
                        if(attribute=="address"):
                            if(add_to_portfolio==True):
                                char=get_column_letter(4)
                                portfolio[char+str(propertyrow)]=value
                                char=get_column_letter(1)
                                userinputs[char+str(propertyrow)]=value
                        if(attribute=="postcode"):
                            if(add_to_portfolio==True):
                                char=get_column_letter(5)
                                portfolio[char+str(propertyrow)]=value
                        if(attribute=="inspection-date"):
                            if(add_to_portfolio==True):
                                char=get_column_letter(6)
                                portfolio[char+str(propertyrow)]=value
                        if(attribute=="uprn"):
                            if(add_to_portfolio==True):
                                char=get_column_letter(8)
                                portfolio[char+str(propertyrow)]=value
                        if(attribute=="lmk-key"):
                            if(add_to_portfolio==True):
                                char=get_column_letter(9)
                                portfolio[char+str(propertyrow)]=value
                        if(attribute=="building-emissions"):
                            buildingEmissions=(value)

                            if(len(str(buildingEmissions))==0):
                                buildingEmissions=0
                                value="0"
                                print(value)

                        #print("Value:",value)
                        #print("Attribute: ",attribute)
                        #print("Value:",value)
                        ws.append([str(attribute), str(value)])


                    #PROPERTY TYPE
                    if(add_to_portfolio==True):
                        char=get_column_letter(2)
                        portfolio[char+str(propertyrow)]="Non-Domestic"



                    if(floor==False):
                        print("THIS UPRNS/LMKKEY floor area is N/A",uprn)
                        continue



                    if(inspectionDate[0:4]=="2020" or inspectionDate[0:4]=="2021" or inspectionDate[0:4]=="2022" or inspectionDate[0:4]=="2023"):


                        char=get_column_letter(4)
                        ws[char+str(1)]="Building Attribute"
                        char=get_column_letter(5)
                        ws[char+str(1)]="U_Value L-Doc (W/m^2K)"
                        char=get_column_letter(6)
                        ws[char+str(1)]="U_Value Leti (W/m^2K)"
                        char=get_column_letter(7)
                        ws[char+str(1)]="U_Value Difference (W/m^2K)"
                        char=get_column_letter(8)
                        ws[char+str(1)]="Area"
                        char=get_column_letter(9)
                        ws[char+str(1)]="U_value L-Doc Based on Area"
                        char=get_column_letter(10)
                        ws[char+str(1)]="U_value Leti Based on Area"
                        char=get_column_letter(11)
                        ws[char+str(1)]="U-value Improvement Based on Area"
                        char=get_column_letter(12)
                        ws[char+str(1)]="Savings"
                        #code for cost of retrofit per contituent
                        char=get_column_letter(13)
                        ws[char+str(1)]="low Cost of Retrofit Constituent (m^2)"
                        char=get_column_letter(14)
                        ws[char+str(1)]="Cost of Retrofit Constituent (m^2) / Improvement based on area"
                        char=get_column_letter(15)
                        ws[char+str(1)]="Avg. Cost of Retrofit Constituent (m^2)"
                        char=get_column_letter(16)
                        ws[char+str(1)]="Cost of Retrofit Constituent (m^2) / Improvement based on area"
                        char=get_column_letter(17)
                        ws[char+str(1)]="High Cost of Retrofit Constituent (m^2)"
                        char=get_column_letter(18)
                        ws[char+str(1)]="Cost of Retrofit Constituent (m^2) / Improvement based on area"
                        char=get_column_letter(19)
                        ws[char+str(1)]="Cost of Retrofit Constituent (m^2) * area"
                        char=get_column_letter(20)
                        ws[char+str(1)]="Energy Savings (GBP)"
                        char=get_column_letter(21)
                        ws[char+str(1)]="Energy Savings (kWh)"








                        char=get_column_letter(4)
                        ws[char+str(2)]="Window"
                        char=get_column_letter(5)
                        ws[char+str(2)]=uvalue_2021["window_L"]
                        char=get_column_letter(6)
                        ws[char+str(2)]=uvalue_2021["window_leti"]
                        char=get_column_letter(7)
                        ws[char+str(2)]=uvalue_2021["window_L"]-uvalue_2021["window_leti"]
                        char=get_column_letter(8)
                        ws[char+str(2)]=(math.sqrt(float(floorarea))*heightt)*(0.20)
                        char=get_column_letter(9)
                        ws[char+str(2)]=(((math.sqrt(float(floorarea))*heightt)*(0.20)) * uvalue_2021["window_L"])
                        char=get_column_letter(10)
                        ws[char+str(2)]=((math.sqrt(float(floorarea))*heightt)*(0.20)) * uvalue_2021["window_leti"]
                        char=get_column_letter(11)
                        ws[char+str(2)]=(((math.sqrt(float(floorarea))*heightt)*(0.20)) * uvalue_2021["window_L"]) - (((math.sqrt(float(floorarea))*heightt)*(0.20)) * uvalue_2021["window_leti"])



                        char=get_column_letter(4)
                        ws[char+str(3)]="Door"
                        char=get_column_letter(5)
                        ws[char+str(3)]=uvalue_2021["door_L"]
                        char=get_column_letter(6)
                        ws[char+str(3)]=uvalue_2021["door_leti"]
                        char=get_column_letter(7)
                        ws[char+str(3)]=uvalue_2021["door_L"]-uvalue_2021["door_leti"]
                        char=get_column_letter(8)
                        ws[char+str(3)]=(math.sqrt(float(floorarea))*heightt)*(0.40)
                        char=get_column_letter(9)
                        ws[char+str(3)]=(((math.sqrt(float(floorarea))*heightt)*(0.40)) * uvalue_2021["door_L"])
                        char=get_column_letter(10)
                        ws[char+str(3)]=((math.sqrt(float(floorarea))*heightt)*(0.40)) * uvalue_2021["door_leti"]
                        char=get_column_letter(11)
                        ws[char+str(3)]=((((math.sqrt(float(floorarea))*heightt)*(0.40)) * uvalue_2021["door_L"]) - (((math.sqrt(float(floorarea))*heightt)*(0.40)) * uvalue_2021["door_leti"]))


                        char=get_column_letter(4)
                        ws[char+str(4)]="Roof"
                        char=get_column_letter(5)
                        ws[char+str(4)]=uvalue_2021["roof_L"]
                        char=get_column_letter(6)
                        ws[char+str(4)]=uvalue_2021["roof_leti"]
                        char=get_column_letter(7)
                        ws[char+str(4)]=uvalue_2021["roof_L"]-uvalue_2021["roof_leti"]
                        char=get_column_letter(8)
                        ws[char+str(4)]=(float(floorarea))
                        char=get_column_letter(9)
                        ws[char+str(4)]=((float(floorarea)) * uvalue_2021["roof_L"])
                        char=get_column_letter(10)
                        ws[char+str(4)]=(float(floorarea)) * uvalue_2021["roof_leti"]
                        char=get_column_letter(11)
                        ws[char+str(4)]=(((float(floorarea)) * uvalue_2021["roof_L"]) - ((float(floorarea)) * uvalue_2021["roof_leti"]))



                        char=get_column_letter(4)
                        ws[char+str(5)]="Wall"
                        char=get_column_letter(5)
                        ws[char+str(5)]=uvalue_2021["wall_L"]
                        char=get_column_letter(6)
                        ws[char+str(5)]=uvalue_2021["wall_leti"]
                        char=get_column_letter(7)
                        ws[char+str(5)]=uvalue_2021["wall_L"]-uvalue_2021["wall_leti"]
                        char=get_column_letter(8)
                        ws[char+str(5)]=(math.sqrt(float(floorarea))*heightt)*(0.80)
                        char=get_column_letter(9)
                        ws[char+str(5)]=(((math.sqrt(float(floorarea))*heightt)*(0.80)) * uvalue_2021["wall_L"])
                        char=get_column_letter(10)
                        ws[char+str(5)]=((math.sqrt(float(floorarea))*heightt)*(0.80)) * uvalue_2021["wall_leti"]
                        char=get_column_letter(11)
                        ws[char+str(5)]=((((math.sqrt(float(floorarea))*heightt)*(0.80)) * uvalue_2021["wall_L"]) - (((math.sqrt(float(floorarea))*heightt)*(0.80)) * uvalue_2021["wall_leti"]))


                        char=get_column_letter(4)
                        ws[char+str(6)]="Floor"
                        char=get_column_letter(5)
                        ws[char+str(6)]=uvalue_2021["floor_L"]
                        char=get_column_letter(6)
                        ws[char+str(6)]=uvalue_2021["floor_leti"]
                        char=get_column_letter(7)
                        ws[char+str(6)]=uvalue_2021["floor_L"]-uvalue_2021["floor_leti"]
                        char=get_column_letter(8)
                        ws[char+str(6)]=(float(floorarea))
                        char=get_column_letter(9)
                        ws[char+str(6)]=((float(floorarea)) * uvalue_2021["floor_L"])
                        char=get_column_letter(10)
                        ws[char+str(6)]=(float(floorarea)) * uvalue_2021["floor_leti"]
                        char=get_column_letter(11)
                        ws[char+str(6)]=(((float(floorarea)) * uvalue_2021["floor_L"]) - ((float(floorarea)) * uvalue_2021["floor_leti"]))


                        char=get_column_letter(8)
                        ws[char+str(7)]="Total U-Value"

                        ws['L2']='=(K2/I2)'
                        ws['L3']='=(K3/I3)'
                        ws['L4']='=(K4/I4)'
                        ws['L5']='=(K5/I5)'
                        ws['L6']='=(K6/I6)'
                        ws['L7']='=(K7/I7)'
                        ws['I7']='=SUM(I2:I6)'
                        ws['I8']=f"=UserInputs!B{propertyrow}+UserInputs!C{propertyrow}"
                        ws['J7']='=SUM(J2:J6)'
                        ws['K7']='=SUM(K2:K6)'
                        ws['L8']='=(L7)*I8'

                        #const of retrofit * area
                        ws['S2']='=(H2)*O2'
                        ws['S3']='=(H3)*O3'
                        ws['S4']='=(H4)*O4'
                        ws['S5']='=(H5)*O5'
                        ws['S6']='=(H6)*O6'

                        #TOTAL SAVINGS
                        char=get_column_letter(11)
                        ws[char+str(8)]="Total Savings"

                        #Energy Savings
                        ws['T2']='=((((K2)*(Portfolio!D50))/1000)*365*24*(Portfolio!D67))'
                        ws['T3']='=((((K3)*(Portfolio!D50))/1000)*365*24*(Portfolio!D67))'
                        ws['T4']='=((((K4)*(Portfolio!D50))/1000)*365*24*(Portfolio!D67))'
                        ws['T5']='=((((K5)*(Portfolio!D50))/1000)*365*24*(Portfolio!D67))'
                        ws['T6']='=((((K6)*(Portfolio!D50))/1000)*365*24*(Portfolio!D67))'
                        ws['T7']='=SUM(T2:T6)'


                        ws['U2']=f'=((((K2)*(Portfolio!D50))/1000)*365*24)/12'
                        ws['U3']=f'=((((K3)*(Portfolio!D50))/1000)*365*24)/12'
                        ws['U4']=f'=((((K4)*(Portfolio!D50))/1000)*365*24)/12'
                        ws['U5']=f'=((((K5)*(Portfolio!D50))/1000)*365*24)/12'
                        ws['U6']=f'=((((K6)*(Portfolio!D50))/1000)*365*24)/12'
                        ws['U7']='=SUM(U2:U6)'


                        if(add_to_portfolio==True):
                            #TO ADD TO PORTFOLIO

                            #difference u value
                            char=get_column_letter(11)
                            portfolio[char+str(propertyrow)]=uvalue_2021["window_L"]-uvalue_2021["window_leti"]
                            char=get_column_letter(10)
                            portfolio[char+str(propertyrow)]=uvalue_2021["door_L"]-uvalue_2021["door_leti"]
                            char=get_column_letter(12)
                            portfolio[char+str(propertyrow)]=uvalue_2021["roof_L"]-uvalue_2021["roof_leti"]
                            char=get_column_letter(13)
                            portfolio[char+str(propertyrow)]=uvalue_2021["wall_L"]-uvalue_2021["wall_leti"]
                            char=get_column_letter(14)
                            portfolio[char+str(propertyrow)]=uvalue_2021["floor_L"]-uvalue_2021["floor_leti"]

                            #area
                            char=get_column_letter(16)
                            portfolio[char+str(propertyrow)]=(math.sqrt(floorarea)*heightt)*(0.20)
                            char=get_column_letter(15)
                            portfolio[char+str(propertyrow)]=(((math.sqrt(floorarea)*heightt)*(0.20))*(0.20))
                            char=get_column_letter(17)
                            portfolio[char+str(propertyrow)]=(floorarea)
                            char=get_column_letter(18)
                            portfolio[char+str(propertyrow)]=(math.sqrt(floorarea)*heightt)*(0.80)
                            char=get_column_letter(19)
                            portfolio[char+str(propertyrow)]=(floorarea)

                            #improvement
                            char=get_column_letter(21)
                            portfolio[char+str(propertyrow)]=(((math.sqrt(floorarea)*heightt)*(0.20)) * uvalue_2021["window_L"]) - (((math.sqrt(floorarea)*heightt)*(0.20)) * uvalue_2021["window_leti"])
                            char=get_column_letter(20)
                            portfolio[char+str(propertyrow)]=((((((math.sqrt(floorarea)*heightt)*(0.20))*(0.20))) * uvalue_2021["door_L"]) - (((((math.sqrt(floorarea)*heightt)*(0.20))*(0.20))) * uvalue_2021["door_leti"]))
                            char=get_column_letter(22)
                            portfolio[char+str(propertyrow)]=(((floorarea) * uvalue_2021["roof_L"]) - ((floorarea) * uvalue_2021["roof_leti"]))
                            char=get_column_letter(23)
                            portfolio[char+str(propertyrow)]=((((math.sqrt(floorarea)*heightt)*(0.80)) * uvalue_2021["wall_L"]) - (((math.sqrt(floorarea)*heightt)*(0.80)) * uvalue_2021["wall_leti"]))
                            char=get_column_letter(24)
                            portfolio[char+str(propertyrow)]=(((floorarea) * uvalue_2021["floor_L"]) - ((floorarea) * uvalue_2021["floor_leti"]))
                            char=get_column_letter(25)
                            portfolio[char+str(propertyrow)]=(((((math.sqrt(floorarea)*heightt)*(0.20)) * uvalue_2021["window_L"]) - (((math.sqrt(floorarea)*heightt)*(0.20)) * uvalue_2021["window_leti"])) + ((((((math.sqrt(floorarea)*heightt)*(0.20))*(0.20))) * uvalue_2021["door_L"]) - (((((math.sqrt(floorarea)*heightt)*(0.20))*(0.20))) * uvalue_2021["door_leti"])) + (((floorarea) * uvalue_2021["roof_L"]) - ((floorarea) * uvalue_2021["roof_leti"])) + ((((math.sqrt(floorarea)*heightt)*(0.80)) * uvalue_2021["wall_L"]) - (((math.sqrt(floorarea)*heightt)*(0.80)) * uvalue_2021["wall_leti"])) + (((floorarea) * uvalue_2021["floor_L"]) - ((floorarea) * uvalue_2021["floor_leti"])))


                            #savings
                            char=get_column_letter(27)
                            portfolio[char+str(propertyrow)]=((((math.sqrt(floorarea)*heightt)*(0.20)) * uvalue_2021["window_L"]) - (((math.sqrt(floorarea)*heightt)*(0.20)) * uvalue_2021["window_leti"])) / ((((math.sqrt(floorarea)*heightt)*(0.20)) * uvalue_2021["window_L"]))
                            char=get_column_letter(26)
                            portfolio[char+str(propertyrow)]=((((((math.sqrt(floorarea)*heightt)*(0.20))*(0.20))) * uvalue_2021["door_L"]) - (((((math.sqrt(floorarea)*heightt)*(0.20))*(0.20))) * uvalue_2021["door_leti"])) / ((((((math.sqrt(floorarea)*heightt)*(0.20))*(0.20))) * uvalue_2021["door_L"]))
                            char=get_column_letter(28)
                            portfolio[char+str(propertyrow)]=(((floorarea) * uvalue_2021["roof_L"]) - ((floorarea) * uvalue_2021["roof_leti"])) / ((floorarea) * uvalue_2021["roof_L"])
                            char=get_column_letter(29)
                            portfolio[char+str(propertyrow)]=((((math.sqrt(floorarea)*heightt)*(0.80)) * uvalue_2021["wall_L"]) - (((math.sqrt(floorarea)*heightt)*(0.80)) * uvalue_2021["wall_leti"])) / (((math.sqrt(floorarea)*heightt)*(0.80)) * uvalue_2021["wall_L"]  )
                            char=get_column_letter(30)
                            portfolio[char+str(propertyrow)]=(((floorarea) * uvalue_2021["floor_L"]) - ((floorarea) * uvalue_2021["floor_leti"])) / ((floorarea) * uvalue_2021["floor_L"])
                            char=get_column_letter(31)
                            portfolio[char+str(propertyrow)]=(((((math.sqrt(floorarea)*heightt)*(0.20)) * uvalue_2021["window_L"]) - (((math.sqrt(floorarea)*heightt)*(0.20)) * uvalue_2021["window_leti"])) + ((((((math.sqrt(floorarea)*heightt)*(0.20))*(0.20))) * uvalue_2021["door_L"]) - (((((math.sqrt(floorarea)*heightt)*(0.20))*(0.20))) * uvalue_2021["door_leti"])) + (((floorarea) * uvalue_2021["roof_L"]) - ((floorarea) * uvalue_2021["roof_leti"])) + ((((math.sqrt(floorarea)*heightt)*(0.80)) * uvalue_2021["wall_L"]) - (((math.sqrt(floorarea)*heightt)*(0.80)) * uvalue_2021["wall_leti"])) + (((floorarea) * uvalue_2021["floor_L"]) - ((floorarea) * uvalue_2021["floor_leti"]))) / ((((math.sqrt(floorarea)*heightt)*(0.20)) * uvalue_2021["window_L"]) + (((((math.sqrt(floorarea)*heightt)*(0.20))*(0.20))) * uvalue_2021["door_L"]) + ((floorarea) * uvalue_2021["roof_L"]) + (((math.sqrt(floorarea)*heightt)*(0.80)) * uvalue_2021["wall_L"]) + ((floorarea) * uvalue_2021["floor_L"]))
                            #FINAL SAVINGS
                            char=get_column_letter(32)
                            portfolio[char+str(propertyrow)]='='+str(uprn)+'!L8*12'

                            #cost of retrofit per meter square for all constituents
                            char=get_column_letter(33)
                            portfolio[get_column_letter(41)+str(propertyrow)]='=SUM('+str(uprn)+'!S2:S6)'

                            #Energy and Emissions Savings
                            up=str(uprn)

                            char=get_column_letter(51)
                            portfolio[char+str(propertyrow)]=f"={up}!L7*{buildingEmissions}"
                            char=get_column_letter(52)
                            portfolio[char+str(propertyrow)]=f"={up}!T2"
                            char=get_column_letter(53)
                            portfolio[char+str(propertyrow)]=f"={up}!T3"
                            char=get_column_letter(54)
                            portfolio[char+str(propertyrow)]=f"={up}!T4"
                            char=get_column_letter(55)
                            portfolio[char+str(propertyrow)]=f"={up}!T5"
                            char=get_column_letter(56)
                            portfolio[char+str(propertyrow)]=f"={up}!T6"
                            char=get_column_letter(57)
                            portfolio[char+str(propertyrow)]=f"={up}!U7"
                            char=get_column_letter(58)
                            portfolio[char+str(propertyrow)]=f"=INDEX({up}!U2:U6, MATCH(MAX({up}!L2:L6), {up}!L2:L6, 0))"
                            char=get_column_letter(59)
                            portfolio[char+str(propertyrow)]=f"=(INDEX({up}!H2:H6, MATCH(MAX({up}!L2:L6), {up}!L2:L6, 0))/{up}!B36)*{up}!B1"
                            char=get_column_letter(60)
                            portfolio[char+str(propertyrow)]=f"=INDEX({up}!T2:T6, MATCH(MAX({up}!L2:L6), {up}!L2:L6, 0))"
                            char=get_column_letter(61)
                            portfolio[char+str(propertyrow)]=f"={up}!I7"







                        #Code fo Constituent retrofift calculation
                        #cost of constituent outputs
                        char=get_column_letter(13)
                        ws[char+str(2)]=f"=Portfolio!D51"
                        char=get_column_letter(13)
                        ws[char+str(3)]=f"=Portfolio!D52"
                        char=get_column_letter(13)
                        ws[char+str(4)]=f"=Portfolio!D53"
                        char=get_column_letter(13)
                        ws[char+str(5)]=f"=Portfolio!D54"
                        char=get_column_letter(13)
                        ws[char+str(6)]=f"=Portfolio!D55"

                        char=get_column_letter(15)
                        ws[char+str(2)]=f"=Portfolio!D56"
                        char=get_column_letter(15)
                        ws[char+str(3)]=f"=Portfolio!D57"
                        char=get_column_letter(15)
                        ws[char+str(4)]=f"=Portfolio!D58"
                        char=get_column_letter(15)
                        ws[char+str(5)]=f"=Portfolio!D59"
                        char=get_column_letter(15)
                        ws[char+str(6)]=f"=Portfolio!D60"

                        char=get_column_letter(17)
                        ws[char+str(2)]=f"=Portfolio!D61"
                        char=get_column_letter(17)
                        ws[char+str(3)]=f"=Portfolio!D62"
                        char=get_column_letter(17)
                        ws[char+str(4)]=f"=Portfolio!D63"
                        char=get_column_letter(17)
                        ws[char+str(5)]=f"=Portfolio!D64"
                        char=get_column_letter(17)
                        ws[char+str(6)]=f"=Portfolio!D65"


                        ws['N2']='=(M2/K2)'
                        ws['N3']='=(M3/K3)'
                        ws['N4']='=(M4/K4)'
                        ws['N5']='=(M5/K5)'
                        ws['N6']='=(M6/K6)'

                        ws['P2']='=(O2/K2)'
                        ws['P3']='=(O3/K3)'
                        ws['P4']='=(O4/K4)'
                        ws['P5']='=(O5/K5)'
                        ws['P6']='=(O6/K6)'

                        ws['R2']='=(Q2/K2)'
                        ws['R3']='=(Q3/K3)'
                        ws['R4']='=(Q4/K4)'
                        ws['R5']='=(Q5/K5)'
                        ws['R6']='=(Q6/K6)'
                    if(inspectionDate[0:4]=="2010"  or inspectionDate[0:4]=="2008"):


                        char=get_column_letter(4)
                        ws[char+str(1)]="Building Attribute"
                        char=get_column_letter(5)
                        ws[char+str(1)]="U_Value L-Doc (W/m^2K)"
                        char=get_column_letter(6)
                        ws[char+str(1)]="U_Value Leti (W/m^2K)"
                        char=get_column_letter(7)
                        ws[char+str(1)]="U_Value Difference (W/m^2K)"
                        char=get_column_letter(8)
                        ws[char+str(1)]="Area"
                        char=get_column_letter(9)
                        ws[char+str(1)]="U_value L-Doc Based on Area"
                        char=get_column_letter(10)
                        ws[char+str(1)]="U_value Leti Based on Area"
                        char=get_column_letter(11)
                        ws[char+str(1)]="U-value Improvement Based on Area"
                        char=get_column_letter(12)
                        ws[char+str(1)]="Savings"
                        #code for cost of retrofit per contituent
                        char=get_column_letter(13)
                        ws[char+str(1)]="low Cost of Retrofit Constituent (m^2)"
                        char=get_column_letter(14)
                        ws[char+str(1)]="Cost of Retrofit Constituent (m^2) / Improvement based on area"
                        char=get_column_letter(15)
                        ws[char+str(1)]="Avg. Cost of Retrofit Constituent (m^2)"
                        char=get_column_letter(16)
                        ws[char+str(1)]="Cost of Retrofit Constituent (m^2) / Improvement based on area"
                        char=get_column_letter(17)
                        ws[char+str(1)]="High Cost of Retrofit Constituent (m^2)"
                        char=get_column_letter(18)
                        ws[char+str(1)]="Cost of Retrofit Constituent (m^2) / Improvement based on area"
                        char=get_column_letter(19)
                        ws[char+str(1)]="Cost of Retrofit Constituent (m^2) * area"
                        char=get_column_letter(20)
                        ws[char+str(1)]="Energy Savings (GBP)"
                        char=get_column_letter(21)
                        ws[char+str(1)]="Energy Savings (kWh)"

                        char=get_column_letter(4)
                        ws[char+str(2)]="Window"
                        char=get_column_letter(5)
                        ws[char+str(2)]=uvalue_2010["window_L"]
                        char=get_column_letter(6)
                        ws[char+str(2)]=uvalue_2010["window_leti"]
                        char=get_column_letter(7)
                        ws[char+str(2)]=uvalue_2010["window_L"]-uvalue_2010["window_leti"]
                        char=get_column_letter(8)
                        ws[char+str(2)]=(math.sqrt(float(floorarea))*heightt)*(0.20)
                        char=get_column_letter(9)
                        ws[char+str(2)]=(((math.sqrt(float(floorarea))*heightt)*(0.20)) * uvalue_2010["window_L"])
                        char=get_column_letter(10)
                        ws[char+str(2)]=((math.sqrt(float(floorarea))*heightt)*(0.20)) * uvalue_2010["window_leti"]
                        char=get_column_letter(11)
                        ws[char+str(2)]=(((math.sqrt(float(floorarea))*heightt)*(0.20)) * uvalue_2010["window_L"]) - (((math.sqrt(float(floorarea))*heightt)*(0.20)) * uvalue_2010["window_leti"])
                        char=get_column_letter(12)
                        ws[char+str(2)]=((((math.sqrt(float(floorarea))*heightt)*(0.20)) * uvalue_2010["window_L"]) - (((math.sqrt(float(floorarea))*heightt)*(0.20)) * uvalue_2010["window_leti"])) / ((((math.sqrt(float(floorarea))*heightt)*(0.20)) * uvalue_2010["window_L"]))
                        #BuildingData.json()['data'][0]['height']) THIS IS THE BUILDINGS HEIGHT FROM WHICH WE CANNOT CALCULATE THE WINDOW SIZE



                        char=get_column_letter(4)
                        ws[char+str(3)]="Door"
                        char=get_column_letter(5)
                        ws[char+str(3)]=uvalue_2010["door_L"]
                        char=get_column_letter(6)
                        ws[char+str(3)]=uvalue_2010["door_leti"]
                        char=get_column_letter(7)
                        ws[char+str(3)]=uvalue_2010["door_L"]-uvalue_2010["door_leti"]
                        char=get_column_letter(8)
                        ws[char+str(3)]=(math.sqrt(float(floorarea))*heightt)*(0.40)
                        char=get_column_letter(9)
                        ws[char+str(3)]=(((math.sqrt(float(floorarea))*heightt)*(0.40)) * uvalue_2010["door_L"])
                        char=get_column_letter(10)
                        ws[char+str(3)]=((math.sqrt(float(floorarea))*heightt)*(0.40)) * uvalue_2010["door_leti"]
                        char=get_column_letter(11)
                        ws[char+str(3)]=((((math.sqrt(float(floorarea))*heightt)*(0.40)) * uvalue_2010["door_L"]) - (((math.sqrt(float(floorarea))*heightt)*(0.40)) * uvalue_2010["door_leti"]))
                        char=get_column_letter(12)
                        ws[char+str(3)]=((((math.sqrt(float(floorarea))*heightt)*(0.40)) * uvalue_2010["door_L"]) - (((math.sqrt(float(floorarea))*heightt)*(0.40)) * uvalue_2010["door_leti"])) / ((((math.sqrt(float(floorarea))*heightt)*(0.40)) * uvalue_2010["door_L"]))

                        char=get_column_letter(4)
                        ws[char+str(4)]="Roof"
                        char=get_column_letter(5)
                        ws[char+str(4)]=uvalue_2010["roof_L"]
                        char=get_column_letter(6)
                        ws[char+str(4)]=uvalue_2010["roof_leti"]
                        char=get_column_letter(7)
                        ws[char+str(4)]=uvalue_2010["roof_L"]-uvalue_2010["roof_leti"]
                        char=get_column_letter(8)
                        ws[char+str(4)]=(float(floorarea))
                        char=get_column_letter(9)
                        ws[char+str(4)]=((float(floorarea)) * uvalue_2010["roof_L"])
                        char=get_column_letter(10)
                        ws[char+str(4)]=(float(floorarea)) * uvalue_2010["roof_leti"]
                        char=get_column_letter(11)
                        ws[char+str(4)]=(((float(floorarea)) * uvalue_2010["roof_L"]) - ((float(floorarea)) * uvalue_2010["roof_leti"]))
                        char=get_column_letter(12)
                        ws[char+str(4)]=(((float(floorarea)) * uvalue_2010["roof_L"]) - ((float(floorarea)) * uvalue_2010["roof_leti"])) / ((float(floorarea)) * uvalue_2010["roof_L"])


                        char=get_column_letter(4)
                        ws[char+str(5)]="Wall"
                        char=get_column_letter(5)
                        ws[char+str(5)]=uvalue_2010["wall_L"]
                        char=get_column_letter(6)
                        ws[char+str(5)]=uvalue_2010["wall_leti"]
                        char=get_column_letter(7)
                        ws[char+str(5)]=uvalue_2010["wall_L"]-uvalue_2010["wall_leti"]
                        char=get_column_letter(8)
                        ws[char+str(5)]=(math.sqrt(float(floorarea))*heightt)*(0.80)
                        char=get_column_letter(9)
                        ws[char+str(5)]=(((math.sqrt(float(floorarea))*heightt)*(0.80)) * uvalue_2010["wall_L"])
                        char=get_column_letter(10)
                        ws[char+str(5)]=((math.sqrt(float(floorarea))*heightt)*(0.80)) * uvalue_2010["wall_leti"]
                        char=get_column_letter(11)
                        ws[char+str(5)]=((((math.sqrt(float(floorarea))*heightt)*(0.80)) * uvalue_2010["wall_L"]) - (((math.sqrt(float(floorarea))*heightt)*(0.80)) * uvalue_2010["wall_leti"]))
                        char=get_column_letter(12)
                        ws[char+str(5)]=((((math.sqrt(float(floorarea))*heightt)*(0.80)) * uvalue_2010["wall_L"]) - (((math.sqrt(float(floorarea))*heightt)*(0.80)) * uvalue_2010["wall_leti"])) / (((math.sqrt(float(floorarea))*heightt)*(0.80)) * uvalue_2010["wall_leti"]  )

                        char=get_column_letter(4)
                        ws[char+str(6)]="Floor"
                        char=get_column_letter(5)
                        ws[char+str(6)]=uvalue_2010["floor_L"]
                        char=get_column_letter(6)
                        ws[char+str(6)]=uvalue_2010["floor_leti"]
                        char=get_column_letter(7)
                        ws[char+str(6)]=uvalue_2010["floor_L"]-uvalue_2010["floor_leti"]
                        char=get_column_letter(8)
                        ws[char+str(6)]=(float(floorarea))
                        char=get_column_letter(9)
                        ws[char+str(6)]=((float(floorarea)) * uvalue_2010["floor_L"])
                        char=get_column_letter(10)
                        ws[char+str(6)]=(float(floorarea)) * uvalue_2010["floor_leti"]
                        char=get_column_letter(11)
                        ws[char+str(6)]=(((float(floorarea)) * uvalue_2010["floor_L"]) - ((float(floorarea)) * uvalue_2010["floor_leti"]))
                        char=get_column_letter(12)
                        ws[char+str(6)]= (((float(floorarea)) * uvalue_2010["floor_L"]) - ((float(floorarea)) * uvalue_2010["floor_leti"])) / ((float(floorarea)) * uvalue_2010["floor_L"])

                        char=get_column_letter(8)
                        ws[char+str(7)]="Total U-Value"

                        ws['L2']='=(K2/I2)'
                        ws['L3']='=(K3/I3)'
                        ws['L4']='=(K4/I4)'
                        ws['L5']='=(K5/I5)'
                        ws['L6']='=(K6/I6)'
                        ws['L7']='=(K7/I7)'
                        ws['I7']='=SUM(I2:I6)'
                        ws['I8']=f"=UserInputs!B{propertyrow}+UserInputs!C{propertyrow}"
                        ws['J7']='=SUM(J2:J6)'
                        ws['K7']='=SUM(K2:K6)'
                        ws['L8']='=(L7)*I8'

                        #const of retrofit * area
                        ws['S2']='=(H2)*O2'
                        ws['S3']='=(H3)*O3'
                        ws['S4']='=(H4)*O4'
                        ws['S5']='=(H5)*O5'
                        ws['S6']='=(H6)*O6'

                        #TOTAL SAVINGS

                        char=get_column_letter(11)
                        ws[char+str(8)]="Total Savings"

                        #Energy Savings
                        ws['T2']='=((((K2)*(Portfolio!D50))/1000)*365*24*(Portfolio!D67))'
                        ws['T3']='=((((K3)*(Portfolio!D50))/1000)*365*24*(Portfolio!D67))'
                        ws['T4']='=((((K4)*(Portfolio!D50))/1000)*365*24*(Portfolio!D67))'
                        ws['T5']='=((((K5)*(Portfolio!D50))/1000)*365*24*(Portfolio!D67))'
                        ws['T6']='=((((K6)*(Portfolio!D50))/1000)*365*24*(Portfolio!D67))'
                        ws['T7']='=SUM(T2:T6)'


                        ws['U2']=f'=((((K2)*(Portfolio!D50))/1000)*365*24)/12'
                        ws['U3']=f'=((((K3)*(Portfolio!D50))/1000)*365*24)/12'
                        ws['U4']=f'=((((K4)*(Portfolio!D50))/1000)*365*24)/12'
                        ws['U5']=f'=((((K5)*(Portfolio!D50))/1000)*365*24)/12'
                        ws['U6']=f'=((((K6)*(Portfolio!D50))/1000)*365*24)/12'
                        ws['U7']='=SUM(U2:U6)'

                        if(add_to_portfolio==True):
                            #TO ADD TO PORTFOLIO

                            #difference u value
                            char=get_column_letter(11)
                            portfolio[char+str(propertyrow)]=uvalue_2010["window_L"]-uvalue_2010["window_leti"]
                            char=get_column_letter(10)
                            portfolio[char+str(propertyrow)]=uvalue_2010["door_L"]-uvalue_2010["door_leti"]
                            char=get_column_letter(12)
                            portfolio[char+str(propertyrow)]=uvalue_2010["roof_L"]-uvalue_2010["roof_leti"]
                            char=get_column_letter(13)
                            portfolio[char+str(propertyrow)]=uvalue_2010["wall_L"]-uvalue_2010["wall_leti"]
                            char=get_column_letter(14)
                            portfolio[char+str(propertyrow)]=uvalue_2010["floor_L"]-uvalue_2010["floor_leti"]

                            #area
                            char=get_column_letter(16)
                            portfolio[char+str(propertyrow)]=(math.sqrt(floorarea)*heightt)*(0.20)
                            char=get_column_letter(15)
                            portfolio[char+str(propertyrow)]=(((math.sqrt(floorarea)*heightt)*(0.20))*(0.20))
                            char=get_column_letter(17)
                            portfolio[char+str(propertyrow)]=(floorarea)
                            char=get_column_letter(18)
                            portfolio[char+str(propertyrow)]=(math.sqrt(floorarea)*heightt)*(0.80)
                            char=get_column_letter(19)
                            portfolio[char+str(propertyrow)]=(floorarea)

                            #improvement
                            char=get_column_letter(21)
                            portfolio[char+str(propertyrow)]=(((math.sqrt(floorarea)*heightt)*(0.20)) * uvalue_2010["window_L"]) - (((math.sqrt(floorarea)*heightt)*(0.20)) * uvalue_2010["window_leti"])
                            char=get_column_letter(20)
                            portfolio[char+str(propertyrow)]=((((((math.sqrt(floorarea)*heightt)*(0.20))*(0.20))) * uvalue_2010["door_L"]) - (((((math.sqrt(floorarea)*heightt)*(0.20))*(0.20))) * uvalue_2010["door_leti"]))
                            char=get_column_letter(22)
                            portfolio[char+str(propertyrow)]=(((floorarea) * uvalue_2010["roof_L"]) - ((floorarea) * uvalue_2010["roof_leti"]))
                            char=get_column_letter(23)
                            portfolio[char+str(propertyrow)]=((((math.sqrt(floorarea)*heightt)*(0.80)) * uvalue_2010["wall_L"]) - (((math.sqrt(floorarea)*heightt)*(0.80)) * uvalue_2010["wall_leti"]))
                            char=get_column_letter(24)
                            portfolio[char+str(propertyrow)]=(((floorarea) * uvalue_2010["floor_L"]) - ((floorarea) * uvalue_2010["floor_leti"]))
                            char=get_column_letter(25)
                            portfolio[char+str(propertyrow)]=(((((math.sqrt(floorarea)*heightt)*(0.20)) * uvalue_2010["window_L"]) - (((math.sqrt(floorarea)*heightt)*(0.20)) * uvalue_2010["window_leti"])) + ((((((math.sqrt(floorarea)*heightt)*(0.20))*(0.20))) * uvalue_2010["door_L"]) - (((((math.sqrt(floorarea)*heightt)*(0.20))*(0.20))) * uvalue_2010["door_leti"])) + (((floorarea) * uvalue_2010["roof_L"]) - ((floorarea) * uvalue_2010["roof_leti"])) + ((((math.sqrt(floorarea)*heightt)*(0.80)) * uvalue_2010["wall_L"]) - (((math.sqrt(floorarea)*heightt)*(0.80)) * uvalue_2010["wall_leti"])) + (((floorarea) * uvalue_2010["floor_L"]) - ((floorarea) * uvalue_2010["floor_leti"])))


                            #savings
                            char=get_column_letter(27)
                            portfolio[char+str(propertyrow)]=((((math.sqrt(floorarea)*heightt)*(0.20)) * uvalue_2010["window_L"]) - (((math.sqrt(floorarea)*heightt)*(0.20)) * uvalue_2010["window_leti"])) / ((((math.sqrt(floorarea)*heightt)*(0.20)) * uvalue_2010["window_L"]))
                            char=get_column_letter(26)
                            portfolio[char+str(propertyrow)]=((((((math.sqrt(floorarea)*heightt)*(0.20))*(0.20))) * uvalue_2010["door_L"]) - (((((math.sqrt(floorarea)*heightt)*(0.20))*(0.20))) * uvalue_2010["door_leti"])) / ((((((math.sqrt(floorarea)*heightt)*(0.20))*(0.20))) * uvalue_2010["door_L"]))
                            char=get_column_letter(28)
                            portfolio[char+str(propertyrow)]=(((floorarea) * uvalue_2010["roof_L"]) - ((floorarea) * uvalue_2010["roof_leti"])) / ((floorarea) * uvalue_2010["roof_L"])
                            char=get_column_letter(29)
                            portfolio[char+str(propertyrow)]=((((math.sqrt(floorarea)*heightt)*(0.80)) * uvalue_2010["wall_L"]) - (((math.sqrt(floorarea)*heightt)*(0.80)) * uvalue_2010["wall_leti"])) / (((math.sqrt(floorarea)*heightt)*(0.80)) * uvalue_2010["wall_L"]  )
                            char=get_column_letter(30)
                            portfolio[char+str(propertyrow)]=(((floorarea) * uvalue_2010["floor_L"]) - ((floorarea) * uvalue_2010["floor_leti"])) / ((floorarea) * uvalue_2010["floor_L"])
                            char=get_column_letter(31)
                            portfolio[char+str(propertyrow)]=(((((math.sqrt(floorarea)*heightt)*(0.20)) * uvalue_2010["window_L"]) - (((math.sqrt(floorarea)*heightt)*(0.20)) * uvalue_2010["window_leti"])) + ((((((math.sqrt(floorarea)*heightt)*(0.20))*(0.20))) * uvalue_2010["door_L"]) - (((((math.sqrt(floorarea)*heightt)*(0.20))*(0.20))) * uvalue_2010["door_leti"])) + (((floorarea) * uvalue_2010["roof_L"]) - ((floorarea) * uvalue_2010["roof_leti"])) + ((((math.sqrt(floorarea)*heightt)*(0.80)) * uvalue_2010["wall_L"]) - (((math.sqrt(floorarea)*heightt)*(0.80)) * uvalue_2010["wall_leti"])) + (((floorarea) * uvalue_2010["floor_L"]) - ((floorarea) * uvalue_2010["floor_leti"]))) / ((((math.sqrt(floorarea)*heightt)*(0.20)) * uvalue_2010["window_L"]) + (((((math.sqrt(floorarea)*heightt)*(0.20))*(0.20))) * uvalue_2010["door_L"]) + ((floorarea) * uvalue_2010["roof_L"]) + (((math.sqrt(floorarea)*heightt)*(0.80)) * uvalue_2010["wall_L"]) + ((floorarea) * uvalue_2010["floor_L"]))

                            #FINAL SAVINGS
                            char=get_column_letter(32)
                            portfolio[char+str(propertyrow)]='='+str(uprn)+'!L8*12'

                            #cost of retrofit per meter square for all constituents
                            char=get_column_letter(33)
                            portfolio[get_column_letter(41)+str(propertyrow)]='=SUM('+str(uprn)+'!S2:S6)'

                            #Energy and Emissions Savings
                            up=str(uprn)

                            char=get_column_letter(51)
                            portfolio[char+str(propertyrow)]=f"={up}!L7*{buildingEmissions}"
                            char=get_column_letter(52)
                            portfolio[char+str(propertyrow)]=f"={up}!T2"
                            char=get_column_letter(53)
                            portfolio[char+str(propertyrow)]=f"={up}!T3"
                            char=get_column_letter(54)
                            portfolio[char+str(propertyrow)]=f"={up}!T4"
                            char=get_column_letter(55)
                            portfolio[char+str(propertyrow)]=f"={up}!T5"
                            char=get_column_letter(56)
                            portfolio[char+str(propertyrow)]=f"={up}!T6"
                            char=get_column_letter(57)
                            portfolio[char+str(propertyrow)]=f"={up}!U7"
                            char=get_column_letter(58)
                            portfolio[char+str(propertyrow)]=f"=INDEX({up}!U2:U6, MATCH(MAX({up}!L2:L6), {up}!L2:L6, 0))"
                            char=get_column_letter(59)
                            portfolio[char+str(propertyrow)]=f"=(INDEX({up}!H2:H6, MATCH(MAX({up}!L2:L6), {up}!L2:L6, 0))/{up}!B36)*{up}!B1"
                            char=get_column_letter(60)
                            portfolio[char+str(propertyrow)]=f"=INDEX({up}!T2:T6, MATCH(MAX({up}!L2:L6), {up}!L2:L6, 0))"
                            char=get_column_letter(61)
                            portfolio[char+str(propertyrow)]=f"={up}!I7"
                        #cost of constituent outputs
                        #cost of constituent outputs
                        char=get_column_letter(13)
                        ws[char+str(2)]=f"=Portfolio!D51"
                        char=get_column_letter(13)
                        ws[char+str(3)]=f"=Portfolio!D52"
                        char=get_column_letter(13)
                        ws[char+str(4)]=f"=Portfolio!D53"
                        char=get_column_letter(13)
                        ws[char+str(5)]=f"=Portfolio!D54"
                        char=get_column_letter(13)
                        ws[char+str(6)]=f"=Portfolio!D55"

                        char=get_column_letter(15)
                        ws[char+str(2)]=f"=Portfolio!D56"
                        char=get_column_letter(15)
                        ws[char+str(3)]=f"=Portfolio!D57"
                        char=get_column_letter(15)
                        ws[char+str(4)]=f"=Portfolio!D58"
                        char=get_column_letter(15)
                        ws[char+str(5)]=f"=Portfolio!D59"
                        char=get_column_letter(15)
                        ws[char+str(6)]=f"=Portfolio!D60"

                        char=get_column_letter(17)
                        ws[char+str(2)]=f"=Portfolio!D61"
                        char=get_column_letter(17)
                        ws[char+str(3)]=f"=Portfolio!D62"
                        char=get_column_letter(17)
                        ws[char+str(4)]=f"=Portfolio!D63"
                        char=get_column_letter(17)
                        ws[char+str(5)]=f"=Portfolio!D64"
                        char=get_column_letter(17)
                        ws[char+str(6)]=f"=Portfolio!D65"


                        ws['N2']='=(M2/K2)'
                        ws['N3']='=(M3/K3)'
                        ws['N4']='=(M4/K4)'
                        ws['N5']='=(M5/K5)'
                        ws['N6']='=(M6/K6)'

                        ws['P2']='=(O2/K2)'
                        ws['P3']='=(O3/K3)'
                        ws['P4']='=(O4/K4)'
                        ws['P5']='=(O5/K5)'
                        ws['P6']='=(O6/K6)'

                        ws['R2']='=(Q2/K2)'
                        ws['R3']='=(Q3/K3)'
                        ws['R4']='=(Q4/K4)'
                        ws['R5']='=(Q5/K5)'
                        ws['R6']='=(Q6/K6)'
                    if(inspectionDate[0:4]=="2011"):


                        char=get_column_letter(4)
                        ws[char+str(1)]="Building Attribute"
                        char=get_column_letter(5)
                        ws[char+str(1)]="U_Value L-Doc (W/m^2K)"
                        char=get_column_letter(6)
                        ws[char+str(1)]="U_Value Leti (W/m^2K)"
                        char=get_column_letter(7)
                        ws[char+str(1)]="U_Value Difference (W/m^2K)"
                        char=get_column_letter(8)
                        ws[char+str(1)]="Area"
                        char=get_column_letter(9)
                        ws[char+str(1)]="U_value L-Doc Based on Area"
                        char=get_column_letter(10)
                        ws[char+str(1)]="U_value Leti Based on Area"
                        char=get_column_letter(11)
                        ws[char+str(1)]="U-value Improvement Based on Area"
                        char=get_column_letter(12)
                        ws[char+str(1)]="Savings"
                        #code for cost of retrofit per contituent
                        char=get_column_letter(13)
                        ws[char+str(1)]="low Cost of Retrofit Constituent (m^2)"
                        char=get_column_letter(14)
                        ws[char+str(1)]="Cost of Retrofit Constituent (m^2) / Improvement based on area"
                        char=get_column_letter(15)
                        ws[char+str(1)]="Avg. Cost of Retrofit Constituent (m^2)"
                        char=get_column_letter(16)
                        ws[char+str(1)]="Cost of Retrofit Constituent (m^2) / Improvement based on area"
                        char=get_column_letter(17)
                        ws[char+str(1)]="High Cost of Retrofit Constituent (m^2)"
                        char=get_column_letter(18)
                        ws[char+str(1)]="Cost of Retrofit Constituent (m^2) / Improvement based on area"
                        char=get_column_letter(19)
                        ws[char+str(1)]="Cost of Retrofit Constituent (m^2) * area"
                        char=get_column_letter(20)
                        ws[char+str(1)]="Energy Savings (GBP)"
                        char=get_column_letter(21)
                        ws[char+str(1)]="Energy Savings (kWh)"



                        char=get_column_letter(4)
                        ws[char+str(2)]="Window"
                        char=get_column_letter(5)
                        ws[char+str(2)]=uvalue_2011["window_L"]
                        char=get_column_letter(6)
                        ws[char+str(2)]=uvalue_2011["window_leti"]
                        char=get_column_letter(7)
                        ws[char+str(2)]=uvalue_2011["window_L"]-uvalue_2011["window_leti"]
                        char=get_column_letter(8)
                        ws[char+str(2)]=(math.sqrt(float(floorarea))*heightt)*(0.20)
                        char=get_column_letter(9)
                        ws[char+str(2)]=(((math.sqrt(float(floorarea))*heightt)*(0.20)) * uvalue_2011["window_L"])
                        char=get_column_letter(10)
                        ws[char+str(2)]=((math.sqrt(float(floorarea))*heightt)*(0.20)) * uvalue_2011["window_leti"]
                        char=get_column_letter(11)
                        ws[char+str(2)]=(((math.sqrt(float(floorarea))*heightt)*(0.20)) * uvalue_2011["window_L"]) - (((math.sqrt(float(floorarea))*heightt)*(0.20)) * uvalue_2011["window_leti"])
                        char=get_column_letter(12)
                        ws[char+str(2)]=((((math.sqrt(float(floorarea))*heightt)*(0.20)) * uvalue_2011["window_L"]) - (((math.sqrt(float(floorarea))*heightt)*(0.20)) * uvalue_2011["window_leti"])) / ((((math.sqrt(float(floorarea))*heightt)*(0.20)) * uvalue_2011["window_L"]))
                        #BuildingData.json()['data'][0]['height']) THIS IS THE BUILDINGS HEIGHT FROM WHICH WE CANNOT CALCULATE THE WINDOW SIZE



                        char=get_column_letter(4)
                        ws[char+str(3)]="Door"
                        char=get_column_letter(5)
                        ws[char+str(3)]=uvalue_2011["door_L"]
                        char=get_column_letter(6)
                        ws[char+str(3)]=uvalue_2011["door_leti"]
                        char=get_column_letter(7)
                        ws[char+str(3)]=uvalue_2011["door_L"]-uvalue_2011["door_leti"]
                        char=get_column_letter(8)
                        ws[char+str(3)]=(math.sqrt(float(floorarea))*heightt)*(0.40)
                        char=get_column_letter(9)
                        ws[char+str(3)]=(((math.sqrt(float(floorarea))*heightt)*(0.40)) * uvalue_2011["door_L"])
                        char=get_column_letter(10)
                        ws[char+str(3)]=((math.sqrt(float(floorarea))*heightt)*(0.40)) * uvalue_2011["door_leti"]
                        char=get_column_letter(11)
                        ws[char+str(3)]=((((math.sqrt(float(floorarea))*heightt)*(0.40)) * uvalue_2011["door_L"]) - (((math.sqrt(float(floorarea))*heightt)*(0.40)) * uvalue_2011["door_leti"]))
                        char=get_column_letter(12)
                        ws[char+str(3)]=((((math.sqrt(float(floorarea))*heightt)*(0.40)) * uvalue_2011["door_L"]) - (((math.sqrt(float(floorarea))*heightt)*(0.40)) * uvalue_2011["door_leti"])) / ((((math.sqrt(float(floorarea))*heightt)*(0.40)) * uvalue_2011["door_L"]))

                        char=get_column_letter(4)
                        ws[char+str(4)]="Roof"
                        char=get_column_letter(5)
                        ws[char+str(4)]=uvalue_2011["roof_L"]
                        char=get_column_letter(6)
                        ws[char+str(4)]=uvalue_2011["roof_leti"]
                        char=get_column_letter(7)
                        ws[char+str(4)]=uvalue_2011["roof_L"]-uvalue_2011["roof_leti"]
                        char=get_column_letter(8)
                        ws[char+str(4)]=(float(floorarea))
                        char=get_column_letter(9)
                        ws[char+str(4)]=((float(floorarea)) * uvalue_2011["roof_L"])
                        char=get_column_letter(10)
                        ws[char+str(4)]=(float(floorarea)) * uvalue_2011["roof_leti"]
                        char=get_column_letter(11)
                        ws[char+str(4)]=(((float(floorarea)) * uvalue_2011["roof_L"]) - ((float(floorarea)) * uvalue_2011["roof_leti"]))
                        char=get_column_letter(12)
                        ws[char+str(4)]=(((float(floorarea)) * uvalue_2011["roof_L"]) - ((float(floorarea)) * uvalue_2011["roof_leti"])) / ((float(floorarea)) * uvalue_2011["roof_L"])


                        char=get_column_letter(4)
                        ws[char+str(5)]="Wall"
                        char=get_column_letter(5)
                        ws[char+str(5)]=uvalue_2011["wall_L"]
                        char=get_column_letter(6)
                        ws[char+str(5)]=uvalue_2011["wall_leti"]
                        char=get_column_letter(7)
                        ws[char+str(5)]=uvalue_2011["wall_L"]-uvalue_2011["wall_leti"]
                        char=get_column_letter(8)
                        ws[char+str(5)]=(math.sqrt(float(floorarea))*heightt)*(0.80)
                        char=get_column_letter(9)
                        ws[char+str(5)]=(((math.sqrt(float(floorarea))*heightt)*(0.80)) * uvalue_2011["wall_L"])
                        char=get_column_letter(10)
                        ws[char+str(5)]=((math.sqrt(float(floorarea))*heightt)*(0.80)) * uvalue_2011["wall_leti"]
                        char=get_column_letter(11)
                        ws[char+str(5)]=((((math.sqrt(float(floorarea))*heightt)*(0.80)) * uvalue_2011["wall_L"]) - (((math.sqrt(float(floorarea))*heightt)*(0.80)) * uvalue_2011["wall_leti"]))
                        char=get_column_letter(12)
                        ws[char+str(5)]=((((math.sqrt(float(floorarea))*heightt)*(0.80)) * uvalue_2011["wall_L"]) - (((math.sqrt(float(floorarea))*heightt)*(0.80)) * uvalue_2011["wall_leti"])) / (((math.sqrt(float(floorarea))*heightt)*(0.80)) * uvalue_2011["wall_leti"]  )

                        char=get_column_letter(4)
                        ws[char+str(6)]="Floor"
                        char=get_column_letter(5)
                        ws[char+str(6)]=uvalue_2011["floor_L"]
                        char=get_column_letter(6)
                        ws[char+str(6)]=uvalue_2011["floor_leti"]
                        char=get_column_letter(7)
                        ws[char+str(6)]=uvalue_2011["floor_L"]-uvalue_2011["floor_leti"]
                        char=get_column_letter(8)
                        ws[char+str(6)]=(float(floorarea))
                        char=get_column_letter(9)
                        ws[char+str(6)]=((float(floorarea)) * uvalue_2011["floor_L"])
                        char=get_column_letter(10)
                        ws[char+str(6)]=(float(floorarea)) * uvalue_2011["floor_leti"]
                        char=get_column_letter(11)
                        ws[char+str(6)]=(((float(floorarea)) * uvalue_2011["floor_L"]) - ((float(floorarea)) * uvalue_2011["floor_leti"]))
                        char=get_column_letter(12)
                        ws[char+str(6)]= (((float(floorarea)) * uvalue_2011["floor_L"]) - ((float(floorarea)) * uvalue_2011["floor_leti"])) / ((float(floorarea)) * uvalue_2011["floor_L"])

                        char=get_column_letter(8)
                        ws[char+str(7)]="Total U-Value"

                        ws['L2']='=(K2/I2)'
                        ws['L3']='=(K3/I3)'
                        ws['L4']='=(K4/I4)'
                        ws['L5']='=(K5/I5)'
                        ws['L6']='=(K6/I6)'
                        ws['L7']='=(K7/I7)'
                        ws['I7']='=SUM(I2:I6)'
                        ws['I8']=f"=UserInputs!B{propertyrow}+UserInputs!C{propertyrow}"
                        ws['J7']='=SUM(J2:J6)'
                        ws['K7']='=SUM(K2:K6)'
                        ws['L8']='=(L7)*I8'

                        #const of retrofit * area
                        ws['S2']='=(H2)*O2'
                        ws['S3']='=(H3)*O3'
                        ws['S4']='=(H4)*O4'
                        ws['S5']='=(H5)*O5'
                        ws['S6']='=(H6)*O6'

                        #TOTAL SAVINGS

                        char=get_column_letter(11)
                        ws[char+str(8)]="Total Savings"

                        #Energy Savings
                        ws['T2']='=((((K2)*(Portfolio!D50))/1000)*365*24*(Portfolio!D67))'
                        ws['T3']='=((((K3)*(Portfolio!D50))/1000)*365*24*(Portfolio!D67))'
                        ws['T4']='=((((K4)*(Portfolio!D50))/1000)*365*24*(Portfolio!D67))'
                        ws['T5']='=((((K5)*(Portfolio!D50))/1000)*365*24*(Portfolio!D67))'
                        ws['T6']='=((((K6)*(Portfolio!D50))/1000)*365*24*(Portfolio!D67))'
                        ws['T7']='=SUM(T2:T6)'


                        ws['U2']=f'=((((K2)*(Portfolio!D50))/1000)*365*24)/12'
                        ws['U3']=f'=((((K3)*(Portfolio!D50))/1000)*365*24)/12'
                        ws['U4']=f'=((((K4)*(Portfolio!D50))/1000)*365*24)/12'
                        ws['U5']=f'=((((K5)*(Portfolio!D50))/1000)*365*24)/12'
                        ws['U6']=f'=((((K6)*(Portfolio!D50))/1000)*365*24)/12'
                        ws['U7']='=SUM(U2:U6)'


                        if(add_to_portfolio==True):
                            #TO ADD TO PORTFOLIO

                            #difference u value
                            char=get_column_letter(11)
                            portfolio[char+str(propertyrow)]=uvalue_2011["window_L"]-uvalue_2011["window_leti"]
                            char=get_column_letter(10)
                            portfolio[char+str(propertyrow)]=uvalue_2011["door_L"]-uvalue_2011["door_leti"]
                            char=get_column_letter(12)
                            portfolio[char+str(propertyrow)]=uvalue_2011["roof_L"]-uvalue_2011["roof_leti"]
                            char=get_column_letter(13)
                            portfolio[char+str(propertyrow)]=uvalue_2011["wall_L"]-uvalue_2011["wall_leti"]
                            char=get_column_letter(14)
                            portfolio[char+str(propertyrow)]=uvalue_2011["floor_L"]-uvalue_2011["floor_leti"]

                            #area
                            char=get_column_letter(16)
                            portfolio[char+str(propertyrow)]=(math.sqrt(floorarea)*heightt)*(0.20)
                            char=get_column_letter(15)
                            portfolio[char+str(propertyrow)]=(((math.sqrt(floorarea)*heightt)*(0.20))*(0.20))
                            char=get_column_letter(17)
                            portfolio[char+str(propertyrow)]=(floorarea)
                            char=get_column_letter(18)
                            portfolio[char+str(propertyrow)]=(math.sqrt(floorarea)*heightt)*(0.80)
                            char=get_column_letter(19)
                            portfolio[char+str(propertyrow)]=(floorarea)

                            #improvement
                            char=get_column_letter(21)
                            portfolio[char+str(propertyrow)]=(((math.sqrt(floorarea)*heightt)*(0.20)) * uvalue_2011["window_L"]) - (((math.sqrt(floorarea)*heightt)*(0.20)) * uvalue_2011["window_leti"])
                            char=get_column_letter(20)
                            portfolio[char+str(propertyrow)]=((((((math.sqrt(floorarea)*heightt)*(0.20))*(0.20))) * uvalue_2011["door_L"]) - (((((math.sqrt(floorarea)*heightt)*(0.20))*(0.20))) * uvalue_2011["door_leti"]))
                            char=get_column_letter(22)
                            portfolio[char+str(propertyrow)]=(((floorarea) * uvalue_2011["roof_L"]) - ((floorarea) * uvalue_2011["roof_leti"]))
                            char=get_column_letter(23)
                            portfolio[char+str(propertyrow)]=((((math.sqrt(floorarea)*heightt)*(0.80)) * uvalue_2011["wall_L"]) - (((math.sqrt(floorarea)*heightt)*(0.80)) * uvalue_2011["wall_leti"]))
                            char=get_column_letter(24)
                            portfolio[char+str(propertyrow)]=(((floorarea) * uvalue_2011["floor_L"]) - ((floorarea) * uvalue_2011["floor_leti"]))
                            char=get_column_letter(25)
                            portfolio[char+str(propertyrow)]=(((((math.sqrt(floorarea)*heightt)*(0.20)) * uvalue_2011["window_L"]) - (((math.sqrt(floorarea)*heightt)*(0.20)) * uvalue_2011["window_leti"])) + ((((((math.sqrt(floorarea)*heightt)*(0.20))*(0.20))) * uvalue_2011["door_L"]) - (((((math.sqrt(floorarea)*heightt)*(0.20))*(0.20))) * uvalue_2011["door_leti"])) + (((floorarea) * uvalue_2011["roof_L"]) - ((floorarea) * uvalue_2011["roof_leti"])) + ((((math.sqrt(floorarea)*heightt)*(0.80)) * uvalue_2011["wall_L"]) - (((math.sqrt(floorarea)*heightt)*(0.80)) * uvalue_2011["wall_leti"])) + (((floorarea) * uvalue_2011["floor_L"]) - ((floorarea) * uvalue_2011["floor_leti"])))


                            #savings
                            char=get_column_letter(27)
                            portfolio[char+str(propertyrow)]=((((math.sqrt(floorarea)*heightt)*(0.20)) * uvalue_2011["window_L"]) - (((math.sqrt(floorarea)*heightt)*(0.20)) * uvalue_2011["window_leti"])) / ((((math.sqrt(floorarea)*heightt)*(0.20)) * uvalue_2011["window_L"]))
                            char=get_column_letter(26)
                            portfolio[char+str(propertyrow)]=((((((math.sqrt(floorarea)*heightt)*(0.20))*(0.20))) * uvalue_2011["door_L"]) - (((((math.sqrt(floorarea)*heightt)*(0.20))*(0.20))) * uvalue_2011["door_leti"])) / ((((((math.sqrt(floorarea)*heightt)*(0.20))*(0.20))) * uvalue_2011["door_L"]))
                            char=get_column_letter(28)
                            portfolio[char+str(propertyrow)]=(((floorarea) * uvalue_2011["roof_L"]) - ((floorarea) * uvalue_2011["roof_leti"])) / ((floorarea) * uvalue_2011["roof_L"])
                            char=get_column_letter(29)
                            portfolio[char+str(propertyrow)]=((((math.sqrt(floorarea)*heightt)*(0.80)) * uvalue_2011["wall_L"]) - (((math.sqrt(floorarea)*heightt)*(0.80)) * uvalue_2011["wall_leti"])) / (((math.sqrt(floorarea)*heightt)*(0.80)) * uvalue_2011["wall_L"]  )
                            char=get_column_letter(30)
                            portfolio[char+str(propertyrow)]=(((floorarea) * uvalue_2011["floor_L"]) - ((floorarea) * uvalue_2011["floor_leti"])) / ((floorarea) * uvalue_2011["floor_L"])
                            char=get_column_letter(31)
                            portfolio[char+str(propertyrow)]=(((((math.sqrt(floorarea)*heightt)*(0.20)) * uvalue_2011["window_L"]) - (((math.sqrt(floorarea)*heightt)*(0.20)) * uvalue_2011["window_leti"])) + ((((((math.sqrt(floorarea)*heightt)*(0.20))*(0.20))) * uvalue_2011["door_L"]) - (((((math.sqrt(floorarea)*heightt)*(0.20))*(0.20))) * uvalue_2011["door_leti"])) + (((floorarea) * uvalue_2011["roof_L"]) - ((floorarea) * uvalue_2011["roof_leti"])) + ((((math.sqrt(floorarea)*heightt)*(0.80)) * uvalue_2011["wall_L"]) - (((math.sqrt(floorarea)*heightt)*(0.80)) * uvalue_2011["wall_leti"])) + (((floorarea) * uvalue_2011["floor_L"]) - ((floorarea) * uvalue_2011["floor_leti"]))) / ((((math.sqrt(floorarea)*heightt)*(0.20)) * uvalue_2011["window_L"]) + (((((math.sqrt(floorarea)*heightt)*(0.20))*(0.20))) * uvalue_2011["door_L"]) + ((floorarea) * uvalue_2011["roof_L"]) + (((math.sqrt(floorarea)*heightt)*(0.80)) * uvalue_2011["wall_L"]) + ((floorarea) * uvalue_2011["floor_L"]))


                            #FINAL SAVINGS
                            char=get_column_letter(32)
                            portfolio[char+str(propertyrow)]='='+str(uprn)+'!L8*12'

                            #cost of retrofit per meter square for all constituents
                            char=get_column_letter(33)
                            portfolio[get_column_letter(41)+str(propertyrow)]='=SUM('+str(uprn)+'!S2:S6)'

                            #Energy and Emissions Savings
                            up=str(uprn)

                            char=get_column_letter(51)
                            portfolio[char+str(propertyrow)]=f"={up}!L7*{buildingEmissions}"
                            char=get_column_letter(52)
                            portfolio[char+str(propertyrow)]=f"={up}!T2"
                            char=get_column_letter(53)
                            portfolio[char+str(propertyrow)]=f"={up}!T3"
                            char=get_column_letter(54)
                            portfolio[char+str(propertyrow)]=f"={up}!T4"
                            char=get_column_letter(55)
                            portfolio[char+str(propertyrow)]=f"={up}!T5"
                            char=get_column_letter(56)
                            portfolio[char+str(propertyrow)]=f"={up}!T6"
                            char=get_column_letter(57)
                            portfolio[char+str(propertyrow)]=f"={up}!U7"
                            char=get_column_letter(58)
                            portfolio[char+str(propertyrow)]=f"=INDEX({up}!U2:U6, MATCH(MAX({up}!L2:L6), {up}!L2:L6, 0))"
                            char=get_column_letter(59)
                            portfolio[char+str(propertyrow)]=f"=(INDEX({up}!H2:H6, MATCH(MAX({up}!L2:L6), {up}!L2:L6, 0))/{up}!B36)*{up}!B1"
                            char=get_column_letter(60)
                            portfolio[char+str(propertyrow)]=f"=INDEX({up}!T2:T6, MATCH(MAX({up}!L2:L6), {up}!L2:L6, 0))"
                            char=get_column_letter(61)
                            portfolio[char+str(propertyrow)]=f"={up}!I7"

                        #cost of constituent outputs
                        #cost of constituent outputs
                        char=get_column_letter(13)
                        ws[char+str(2)]=f"=Portfolio!D51"
                        char=get_column_letter(13)
                        ws[char+str(3)]=f"=Portfolio!D52"
                        char=get_column_letter(13)
                        ws[char+str(4)]=f"=Portfolio!D53"
                        char=get_column_letter(13)
                        ws[char+str(5)]=f"=Portfolio!D54"
                        char=get_column_letter(13)
                        ws[char+str(6)]=f"=Portfolio!D55"

                        char=get_column_letter(15)
                        ws[char+str(2)]=f"=Portfolio!D56"
                        char=get_column_letter(15)
                        ws[char+str(3)]=f"=Portfolio!D57"
                        char=get_column_letter(15)
                        ws[char+str(4)]=f"=Portfolio!D58"
                        char=get_column_letter(15)
                        ws[char+str(5)]=f"=Portfolio!D59"
                        char=get_column_letter(15)
                        ws[char+str(6)]=f"=Portfolio!D60"

                        char=get_column_letter(17)
                        ws[char+str(2)]=f"=Portfolio!D61"
                        char=get_column_letter(17)
                        ws[char+str(3)]=f"=Portfolio!D62"
                        char=get_column_letter(17)
                        ws[char+str(4)]=f"=Portfolio!D63"
                        char=get_column_letter(17)
                        ws[char+str(5)]=f"=Portfolio!D64"
                        char=get_column_letter(17)
                        ws[char+str(6)]=f"=Portfolio!D65"


                        ws['N2']='=(M2/K2)'
                        ws['N3']='=(M3/K3)'
                        ws['N4']='=(M4/K4)'
                        ws['N5']='=(M5/K5)'
                        ws['N6']='=(M6/K6)'

                        ws['P2']='=(O2/K2)'
                        ws['P3']='=(O3/K3)'
                        ws['P4']='=(O4/K4)'
                        ws['P5']='=(O5/K5)'
                        ws['P6']='=(O6/K6)'

                        ws['R2']='=(Q2/K2)'
                        ws['R3']='=(Q3/K3)'
                        ws['R4']='=(Q4/K4)'
                        ws['R5']='=(Q5/K5)'
                        ws['R6']='=(Q6/K6)'
                    if(inspectionDate[0:4]=="2013" or inspectionDate[0:4]=="2012" or inspectionDate[0:4]=="2014" or inspectionDate[0:4]=="2015" or inspectionDate[0:4]=="2016" or inspectionDate[0:4]=="2017"):


                        char=get_column_letter(4)
                        ws[char+str(1)]="Building Attribute"
                        char=get_column_letter(5)
                        ws[char+str(1)]="U_Value L-Doc (W/m^2K)"
                        char=get_column_letter(6)
                        ws[char+str(1)]="U_Value Leti (W/m^2K)"
                        char=get_column_letter(7)
                        ws[char+str(1)]="U_Value Difference (W/m^2K)"
                        char=get_column_letter(8)
                        ws[char+str(1)]="Area"
                        char=get_column_letter(9)
                        ws[char+str(1)]="U_value L-Doc Based on Area"
                        char=get_column_letter(10)
                        ws[char+str(1)]="U_value Leti Based on Area"
                        char=get_column_letter(11)
                        ws[char+str(1)]="U-value Improvement Based on Area"
                        char=get_column_letter(12)
                        ws[char+str(1)]="Savings"
                        #code for cost of retrofit per contituent
                        char=get_column_letter(13)
                        ws[char+str(1)]="low Cost of Retrofit Constituent (m^2)"
                        char=get_column_letter(14)
                        ws[char+str(1)]="Cost of Retrofit Constituent (m^2) / Improvement based on area"
                        char=get_column_letter(15)
                        ws[char+str(1)]="Avg. Cost of Retrofit Constituent (m^2)"
                        char=get_column_letter(16)
                        ws[char+str(1)]="Cost of Retrofit Constituent (m^2) / Improvement based on area"
                        char=get_column_letter(17)
                        ws[char+str(1)]="High Cost of Retrofit Constituent (m^2)"
                        char=get_column_letter(18)
                        ws[char+str(1)]="Cost of Retrofit Constituent (m^2) / Improvement based on area"
                        char=get_column_letter(19)
                        ws[char+str(1)]="Cost of Retrofit Constituent (m^2) * area"
                        char=get_column_letter(20)
                        ws[char+str(1)]="Energy Savings (GBP)"
                        char=get_column_letter(21)
                        ws[char+str(1)]="Energy Savings (kWh)"



                        char=get_column_letter(4)
                        ws[char+str(2)]="Window"
                        char=get_column_letter(5)
                        ws[char+str(2)]=uvalue_2013["window_L"]
                        char=get_column_letter(6)
                        ws[char+str(2)]=uvalue_2013["window_leti"]
                        char=get_column_letter(7)
                        ws[char+str(2)]=uvalue_2013["window_L"]-uvalue_2013["window_leti"]
                        char=get_column_letter(8)
                        ws[char+str(2)]=(math.sqrt(float(floorarea))*heightt)*(0.20)
                        char=get_column_letter(9)
                        ws[char+str(2)]=(((math.sqrt(float(floorarea))*heightt)*(0.20)) * uvalue_2013["window_L"])
                        char=get_column_letter(10)
                        ws[char+str(2)]=((math.sqrt(float(floorarea))*heightt)*(0.20)) * uvalue_2013["window_leti"]
                        char=get_column_letter(11)
                        ws[char+str(2)]=(((math.sqrt(float(floorarea))*heightt)*(0.20)) * uvalue_2013["window_L"]) - (((math.sqrt(float(floorarea))*heightt)*(0.20)) * uvalue_2013["window_leti"])
                        char=get_column_letter(12)
                        ws[char+str(2)]=((((math.sqrt(float(floorarea))*heightt)*(0.20)) * uvalue_2013["window_L"]) - (((math.sqrt(float(floorarea))*heightt)*(0.20)) * uvalue_2013["window_leti"])) / ((((math.sqrt(float(floorarea))*heightt)*(0.20)) * uvalue_2013["window_L"]))
                        #BuildingData.json()['data'][0]['height']) THIS IS THE BUILDINGS HEIGHT FROM WHICH WE CANNOT CALCULATE THE WINDOW SIZE



                        char=get_column_letter(4)
                        ws[char+str(3)]="Door"
                        char=get_column_letter(5)
                        ws[char+str(3)]=uvalue_2013["door_L"]
                        char=get_column_letter(6)
                        ws[char+str(3)]=uvalue_2013["door_leti"]
                        char=get_column_letter(7)
                        ws[char+str(3)]=uvalue_2013["door_L"]-uvalue_2013["door_leti"]
                        char=get_column_letter(8)
                        ws[char+str(3)]=(math.sqrt(float(floorarea))*heightt)*(0.40)
                        char=get_column_letter(9)
                        ws[char+str(3)]=(((math.sqrt(float(floorarea))*heightt)*(0.40)) * uvalue_2013["door_L"])
                        char=get_column_letter(10)
                        ws[char+str(3)]=((math.sqrt(float(floorarea))*heightt)*(0.40)) * uvalue_2013["door_leti"]
                        char=get_column_letter(11)
                        ws[char+str(3)]=((((math.sqrt(float(floorarea))*heightt)*(0.40)) * uvalue_2013["door_L"]) - (((math.sqrt(float(floorarea))*heightt)*(0.40)) * uvalue_2013["door_leti"]))
                        char=get_column_letter(12)
                        ws[char+str(3)]=((((math.sqrt(float(floorarea))*heightt)*(0.40)) * uvalue_2013["door_L"]) - (((math.sqrt(float(floorarea))*heightt)*(0.40)) * uvalue_2013["door_leti"])) / ((((math.sqrt(float(floorarea))*heightt)*(0.40)) * uvalue_2013["door_L"]))

                        char=get_column_letter(4)
                        ws[char+str(4)]="Roof"
                        char=get_column_letter(5)
                        ws[char+str(4)]=uvalue_2013["roof_L"]
                        char=get_column_letter(6)
                        ws[char+str(4)]=uvalue_2013["roof_leti"]
                        char=get_column_letter(7)
                        ws[char+str(4)]=uvalue_2013["roof_L"]-uvalue_2013["roof_leti"]
                        char=get_column_letter(8)
                        ws[char+str(4)]=(float(floorarea))
                        char=get_column_letter(9)
                        ws[char+str(4)]=((float(floorarea)) * uvalue_2013["roof_L"])
                        char=get_column_letter(10)
                        ws[char+str(4)]=(float(floorarea)) * uvalue_2013["roof_leti"]
                        char=get_column_letter(11)
                        ws[char+str(4)]=(((float(floorarea)) * uvalue_2013["roof_L"]) - ((float(floorarea)) * uvalue_2013["roof_leti"]))
                        char=get_column_letter(12)
                        ws[char+str(4)]=(((float(floorarea)) * uvalue_2013["roof_L"]) - ((float(floorarea)) * uvalue_2013["roof_leti"])) / ((float(floorarea)) * uvalue_2013["roof_L"])


                        char=get_column_letter(4)
                        ws[char+str(5)]="Wall"
                        char=get_column_letter(5)
                        ws[char+str(5)]=uvalue_2013["wall_L"]
                        char=get_column_letter(6)
                        ws[char+str(5)]=uvalue_2013["wall_leti"]
                        char=get_column_letter(7)
                        ws[char+str(5)]=uvalue_2013["wall_L"]-uvalue_2013["wall_leti"]
                        char=get_column_letter(8)
                        ws[char+str(5)]=(math.sqrt(float(floorarea))*heightt)*(0.80)
                        char=get_column_letter(9)
                        ws[char+str(5)]=(((math.sqrt(float(floorarea))*heightt)*(0.80)) * uvalue_2013["wall_L"])
                        char=get_column_letter(10)
                        ws[char+str(5)]=((math.sqrt(float(floorarea))*heightt)*(0.80)) * uvalue_2013["wall_leti"]
                        char=get_column_letter(11)
                        ws[char+str(5)]=((((math.sqrt(float(floorarea))*heightt)*(0.80)) * uvalue_2013["wall_L"]) - (((math.sqrt(float(floorarea))*heightt)*(0.80)) * uvalue_2013["wall_leti"]))
                        char=get_column_letter(12)
                        ws[char+str(5)]=((((math.sqrt(float(floorarea))*heightt)*(0.80)) * uvalue_2013["wall_L"]) - (((math.sqrt(float(floorarea))*heightt)*(0.80)) * uvalue_2013["wall_leti"])) / (((math.sqrt(float(floorarea))*heightt)*(0.80)) * uvalue_2013["wall_leti"]  )

                        char=get_column_letter(4)
                        ws[char+str(6)]="Floor"
                        char=get_column_letter(5)
                        ws[char+str(6)]=uvalue_2013["floor_L"]
                        char=get_column_letter(6)
                        ws[char+str(6)]=uvalue_2013["floor_leti"]
                        char=get_column_letter(7)
                        ws[char+str(6)]=uvalue_2013["floor_L"]-uvalue_2013["floor_leti"]
                        char=get_column_letter(8)
                        ws[char+str(6)]=(float(floorarea))
                        char=get_column_letter(9)
                        ws[char+str(6)]=((float(floorarea)) * uvalue_2013["floor_L"])
                        char=get_column_letter(10)
                        ws[char+str(6)]=(float(floorarea)) * uvalue_2013["floor_leti"]
                        char=get_column_letter(11)
                        ws[char+str(6)]=(((float(floorarea)) * uvalue_2013["floor_L"]) - ((float(floorarea)) * uvalue_2013["floor_leti"]))
                        char=get_column_letter(12)
                        ws[char+str(6)]= (((float(floorarea)) * uvalue_2013["floor_L"]) - ((float(floorarea)) * uvalue_2013["floor_leti"])) / ((float(floorarea)) * uvalue_2013["floor_L"])

                        char=get_column_letter(8)
                        ws[char+str(7)]="Total U-Value"

                        ws['L2']='=(K2/I2)'
                        ws['L3']='=(K3/I3)'
                        ws['L4']='=(K4/I4)'
                        ws['L5']='=(K5/I5)'
                        ws['L6']='=(K6/I6)'
                        ws['L7']='=(K7/I7)'
                        ws['I7']='=SUM(I2:I6)'
                        ws['I8']=f"=UserInputs!B{propertyrow}+UserInputs!C{propertyrow}"
                        ws['J7']='=SUM(J2:J6)'
                        ws['K7']='=SUM(K2:K6)'
                        ws['L8']='=(L7)*I8'

                        #const of retrofit * area
                        ws['S2']='=(H2)*O2'
                        ws['S3']='=(H3)*O3'
                        ws['S4']='=(H4)*O4'
                        ws['S5']='=(H5)*O5'
                        ws['S6']='=(H6)*O6'

                        #TOTAL SAVINGS

                        char=get_column_letter(11)
                        ws[char+str(8)]="Total Savings"

                        #Energy Savings
                        ws['T2']='=((((K2)*(Portfolio!D50))/1000)*365*24*(Portfolio!D67))'
                        ws['T3']='=((((K3)*(Portfolio!D50))/1000)*365*24*(Portfolio!D67))'
                        ws['T4']='=((((K4)*(Portfolio!D50))/1000)*365*24*(Portfolio!D67))'
                        ws['T5']='=((((K5)*(Portfolio!D50))/1000)*365*24*(Portfolio!D67))'
                        ws['T6']='=((((K6)*(Portfolio!D50))/1000)*365*24*(Portfolio!D67))'
                        ws['T7']='=SUM(T2:T6)'


                        ws['U2']=f'=((((K2)*(Portfolio!D50))/1000)*365*24)/12'
                        ws['U3']=f'=((((K3)*(Portfolio!D50))/1000)*365*24)/12'
                        ws['U4']=f'=((((K4)*(Portfolio!D50))/1000)*365*24)/12'
                        ws['U5']=f'=((((K5)*(Portfolio!D50))/1000)*365*24)/12'
                        ws['U6']=f'=((((K6)*(Portfolio!D50))/1000)*365*24)/12'
                        ws['U7']='=SUM(U2:U6)'



                        if(add_to_portfolio==True):
                            #TO ADD TO PORTFOLIO

                            #difference u value
                            char=get_column_letter(11)
                            portfolio[char+str(propertyrow)]=uvalue_2013["window_L"]-uvalue_2013["window_leti"]
                            char=get_column_letter(10)
                            portfolio[char+str(propertyrow)]=uvalue_2013["door_L"]-uvalue_2013["door_leti"]
                            char=get_column_letter(12)
                            portfolio[char+str(propertyrow)]=uvalue_2013["roof_L"]-uvalue_2013["roof_leti"]
                            char=get_column_letter(13)
                            portfolio[char+str(propertyrow)]=uvalue_2013["wall_L"]-uvalue_2013["wall_leti"]
                            char=get_column_letter(14)
                            portfolio[char+str(propertyrow)]=uvalue_2013["floor_L"]-uvalue_2013["floor_leti"]

                            #area
                            char=get_column_letter(16)
                            portfolio[char+str(propertyrow)]=(math.sqrt(floorarea)*heightt)*(0.20)
                            char=get_column_letter(15)
                            portfolio[char+str(propertyrow)]=(((math.sqrt(floorarea)*heightt)*(0.20))*(0.20))
                            char=get_column_letter(17)
                            portfolio[char+str(propertyrow)]=(floorarea)
                            char=get_column_letter(18)
                            portfolio[char+str(propertyrow)]=(math.sqrt(floorarea)*heightt)*(0.80)
                            char=get_column_letter(19)
                            portfolio[char+str(propertyrow)]=(floorarea)

                            #improvement
                            char=get_column_letter(21)
                            portfolio[char+str(propertyrow)]=(((math.sqrt(floorarea)*heightt)*(0.20)) * uvalue_2013["window_L"]) - (((math.sqrt(floorarea)*heightt)*(0.20)) * uvalue_2013["window_leti"])
                            char=get_column_letter(20)
                            portfolio[char+str(propertyrow)]=((((((math.sqrt(floorarea)*heightt)*(0.20))*(0.20))) * uvalue_2013["door_L"]) - (((((math.sqrt(floorarea)*heightt)*(0.20))*(0.20))) * uvalue_2013["door_leti"]))
                            char=get_column_letter(22)
                            portfolio[char+str(propertyrow)]=(((floorarea) * uvalue_2013["roof_L"]) - ((floorarea) * uvalue_2013["roof_leti"]))
                            char=get_column_letter(23)
                            portfolio[char+str(propertyrow)]=((((math.sqrt(floorarea)*heightt)*(0.80)) * uvalue_2013["wall_L"]) - (((math.sqrt(floorarea)*heightt)*(0.80)) * uvalue_2013["wall_leti"]))
                            char=get_column_letter(24)
                            portfolio[char+str(propertyrow)]=(((floorarea) * uvalue_2013["floor_L"]) - ((floorarea) * uvalue_2013["floor_leti"]))
                            char=get_column_letter(25)
                            portfolio[char+str(propertyrow)]=(((((math.sqrt(floorarea)*heightt)*(0.20)) * uvalue_2013["window_L"]) - (((math.sqrt(floorarea)*heightt)*(0.20)) * uvalue_2013["window_leti"])) + ((((((math.sqrt(floorarea)*heightt)*(0.20))*(0.20))) * uvalue_2013["door_L"]) - (((((math.sqrt(floorarea)*heightt)*(0.20))*(0.20))) * uvalue_2013["door_leti"])) + (((floorarea) * uvalue_2013["roof_L"]) - ((floorarea) * uvalue_2013["roof_leti"])) + ((((math.sqrt(floorarea)*heightt)*(0.80)) * uvalue_2013["wall_L"]) - (((math.sqrt(floorarea)*heightt)*(0.80)) * uvalue_2013["wall_leti"])) + (((floorarea) * uvalue_2013["floor_L"]) - ((floorarea) * uvalue_2013["floor_leti"])))


                            #savings
                            char=get_column_letter(27)
                            portfolio[char+str(propertyrow)]=((((math.sqrt(floorarea)*heightt)*(0.20)) * uvalue_2013["window_L"]) - (((math.sqrt(floorarea)*heightt)*(0.20)) * uvalue_2013["window_leti"])) / ((((math.sqrt(floorarea)*heightt)*(0.20)) * uvalue_2013["window_L"]))
                            char=get_column_letter(26)
                            portfolio[char+str(propertyrow)]=((((((math.sqrt(floorarea)*heightt)*(0.20))*(0.20))) * uvalue_2013["door_L"]) - (((((math.sqrt(floorarea)*heightt)*(0.20))*(0.20))) * uvalue_2013["door_leti"])) / ((((((math.sqrt(floorarea)*heightt)*(0.20))*(0.20))) * uvalue_2013["door_L"]))
                            char=get_column_letter(28)
                            portfolio[char+str(propertyrow)]=(((floorarea) * uvalue_2013["roof_L"]) - ((floorarea) * uvalue_2013["roof_leti"])) / ((floorarea) * uvalue_2013["roof_L"])
                            char=get_column_letter(29)
                            portfolio[char+str(propertyrow)]=((((math.sqrt(floorarea)*heightt)*(0.80)) * uvalue_2013["wall_L"]) - (((math.sqrt(floorarea)*heightt)*(0.80)) * uvalue_2013["wall_leti"])) / (((math.sqrt(floorarea)*heightt)*(0.80)) * uvalue_2013["wall_L"]  )
                            char=get_column_letter(30)
                            portfolio[char+str(propertyrow)]=(((floorarea) * uvalue_2013["floor_L"]) - ((floorarea) * uvalue_2013["floor_leti"])) / ((floorarea) * uvalue_2013["floor_L"])
                            char=get_column_letter(31)
                            portfolio[char+str(propertyrow)]=(((((math.sqrt(floorarea)*heightt)*(0.20)) * uvalue_2013["window_L"]) - (((math.sqrt(floorarea)*heightt)*(0.20)) * uvalue_2013["window_leti"])) + ((((((math.sqrt(floorarea)*heightt)*(0.20))*(0.20))) * uvalue_2013["door_L"]) - (((((math.sqrt(floorarea)*heightt)*(0.20))*(0.20))) * uvalue_2013["door_leti"])) + (((floorarea) * uvalue_2013["roof_L"]) - ((floorarea) * uvalue_2013["roof_leti"])) + ((((math.sqrt(floorarea)*heightt)*(0.80)) * uvalue_2013["wall_L"]) - (((math.sqrt(floorarea)*heightt)*(0.80)) * uvalue_2013["wall_leti"])) + (((floorarea) * uvalue_2013["floor_L"]) - ((floorarea) * uvalue_2013["floor_leti"]))) / ((((math.sqrt(floorarea)*heightt)*(0.20)) * uvalue_2013["window_L"]) + (((((math.sqrt(floorarea)*heightt)*(0.20))*(0.20))) * uvalue_2013["door_L"]) + ((floorarea) * uvalue_2013["roof_L"]) + (((math.sqrt(floorarea)*heightt)*(0.80)) * uvalue_2013["wall_L"]) + ((floorarea) * uvalue_2013["floor_L"]))

                            #FINAL SAVINGS
                            char=get_column_letter(32)
                            portfolio[char+str(propertyrow)]='='+str(uprn)+'!L8*12'

                            #cost of retrofit per meter square for all constituents
                            char=get_column_letter(33)
                            portfolio[get_column_letter(41)+str(propertyrow)]='=SUM('+str(uprn)+'!S2:S6)'

                            #Energy and Emissions Savings
                            up=str(uprn)

                            char=get_column_letter(51)
                            portfolio[char+str(propertyrow)]=f"={up}!L7*{buildingEmissions}"
                            char=get_column_letter(52)
                            portfolio[char+str(propertyrow)]=f"={up}!T2"
                            char=get_column_letter(53)
                            portfolio[char+str(propertyrow)]=f"={up}!T3"
                            char=get_column_letter(54)
                            portfolio[char+str(propertyrow)]=f"={up}!T4"
                            char=get_column_letter(55)
                            portfolio[char+str(propertyrow)]=f"={up}!T5"
                            char=get_column_letter(56)
                            portfolio[char+str(propertyrow)]=f"={up}!T6"
                            char=get_column_letter(57)
                            portfolio[char+str(propertyrow)]=f"={up}!U7"
                            char=get_column_letter(58)
                            portfolio[char+str(propertyrow)]=f"=INDEX({up}!U2:U6, MATCH(MAX({up}!L2:L6), {up}!L2:L6, 0))"
                            char=get_column_letter(59)
                            portfolio[char+str(propertyrow)]=f"=(INDEX({up}!H2:H6, MATCH(MAX({up}!L2:L6), {up}!L2:L6, 0))/{up}!B36)*{up}!B1"
                            char=get_column_letter(60)
                            portfolio[char+str(propertyrow)]=f"=INDEX({up}!T2:T6, MATCH(MAX({up}!L2:L6), {up}!L2:L6, 0))"
                            char=get_column_letter(61)
                            portfolio[char+str(propertyrow)]=f"={up}!I7"

                        #cost of constituent outputs
                        #cost of constituent outputs
                        char=get_column_letter(13)
                        ws[char+str(2)]=f"=Portfolio!D51"
                        char=get_column_letter(13)
                        ws[char+str(3)]=f"=Portfolio!D52"
                        char=get_column_letter(13)
                        ws[char+str(4)]=f"=Portfolio!D53"
                        char=get_column_letter(13)
                        ws[char+str(5)]=f"=Portfolio!D54"
                        char=get_column_letter(13)
                        ws[char+str(6)]=f"=Portfolio!D55"

                        char=get_column_letter(15)
                        ws[char+str(2)]=f"=Portfolio!D56"
                        char=get_column_letter(15)
                        ws[char+str(3)]=f"=Portfolio!D57"
                        char=get_column_letter(15)
                        ws[char+str(4)]=f"=Portfolio!D58"
                        char=get_column_letter(15)
                        ws[char+str(5)]=f"=Portfolio!D59"
                        char=get_column_letter(15)
                        ws[char+str(6)]=f"=Portfolio!D60"

                        char=get_column_letter(17)
                        ws[char+str(2)]=f"=Portfolio!D61"
                        char=get_column_letter(17)
                        ws[char+str(3)]=f"=Portfolio!D62"
                        char=get_column_letter(17)
                        ws[char+str(4)]=f"=Portfolio!D63"
                        char=get_column_letter(17)
                        ws[char+str(5)]=f"=Portfolio!D64"
                        char=get_column_letter(17)
                        ws[char+str(6)]=f"=Portfolio!D65"


                        ws['N2']='=(M2/K2)'
                        ws['N3']='=(M3/K3)'
                        ws['N4']='=(M4/K4)'
                        ws['N5']='=(M5/K5)'
                        ws['N6']='=(M6/K6)'

                        ws['P2']='=(O2/K2)'
                        ws['P3']='=(O3/K3)'
                        ws['P4']='=(O4/K4)'
                        ws['P5']='=(O5/K5)'
                        ws['P6']='=(O6/K6)'

                        ws['R2']='=(Q2/K2)'
                        ws['R3']='=(Q3/K3)'
                        ws['R4']='=(Q4/K4)'
                        ws['R5']='=(Q5/K5)'
                        ws['R6']='=(Q6/K6)'
                    if(inspectionDate[0:4]=="2018"):


                        char=get_column_letter(4)
                        ws[char+str(1)]="Building Attribute"
                        char=get_column_letter(5)
                        ws[char+str(1)]="U_Value L-Doc (W/m^2K)"
                        char=get_column_letter(6)
                        ws[char+str(1)]="U_Value Leti (W/m^2K)"
                        char=get_column_letter(7)
                        ws[char+str(1)]="U_Value Difference (W/m^2K)"
                        char=get_column_letter(8)
                        ws[char+str(1)]="Area"
                        char=get_column_letter(9)
                        ws[char+str(1)]="U_value L-Doc Based on Area"
                        char=get_column_letter(10)
                        ws[char+str(1)]="U_value Leti Based on Area"
                        char=get_column_letter(11)
                        ws[char+str(1)]="U-value Improvement Based on Area"
                        char=get_column_letter(12)
                        ws[char+str(1)]="Savings"
                        #code for cost of retrofit per contituent
                        char=get_column_letter(13)
                        ws[char+str(1)]="low Cost of Retrofit Constituent (m^2)"
                        char=get_column_letter(14)
                        ws[char+str(1)]="Cost of Retrofit Constituent (m^2) / Improvement based on area"
                        char=get_column_letter(15)
                        ws[char+str(1)]="Avg. Cost of Retrofit Constituent (m^2)"
                        char=get_column_letter(16)
                        ws[char+str(1)]="Cost of Retrofit Constituent (m^2) / Improvement based on area"
                        char=get_column_letter(17)
                        ws[char+str(1)]="High Cost of Retrofit Constituent (m^2)"
                        char=get_column_letter(18)
                        ws[char+str(1)]="Cost of Retrofit Constituent (m^2) / Improvement based on area"
                        char=get_column_letter(19)
                        ws[char+str(1)]="Cost of Retrofit Constituent (m^2) * area"
                        char=get_column_letter(20)
                        ws[char+str(1)]="Energy Savings (GBP)"
                        char=get_column_letter(21)
                        ws[char+str(1)]="Energy Savings (kWh)"



                        char=get_column_letter(4)
                        ws[char+str(2)]="Window"
                        char=get_column_letter(5)
                        ws[char+str(2)]=uvalue_2018["window_L"]
                        char=get_column_letter(6)
                        ws[char+str(2)]=uvalue_2018["window_leti"]
                        char=get_column_letter(7)
                        ws[char+str(2)]=uvalue_2018["window_L"]-uvalue_2018["window_leti"]
                        char=get_column_letter(8)
                        ws[char+str(2)]=(math.sqrt(float(floorarea))*heightt)*(0.20)
                        char=get_column_letter(9)
                        ws[char+str(2)]=(((math.sqrt(float(floorarea))*heightt)*(0.20)) * uvalue_2018["window_L"])
                        char=get_column_letter(10)
                        ws[char+str(2)]=((math.sqrt(float(floorarea))*heightt)*(0.20)) * uvalue_2018["window_leti"]
                        char=get_column_letter(11)
                        ws[char+str(2)]=(((math.sqrt(float(floorarea))*heightt)*(0.20)) * uvalue_2018["window_L"]) - (((math.sqrt(float(floorarea))*heightt)*(0.20)) * uvalue_2018["window_leti"])
                        char=get_column_letter(12)
                        ws[char+str(2)]=((((math.sqrt(float(floorarea))*heightt)*(0.20)) * uvalue_2018["window_L"]) - (((math.sqrt(float(floorarea))*heightt)*(0.20)) * uvalue_2018["window_leti"])) / ((((math.sqrt(float(floorarea))*heightt)*(0.20)) * uvalue_2018["window_L"]))
                        #BuildingData.json()['data'][0]['height']) THIS IS THE BUILDINGS HEIGHT FROM WHICH WE CANNOT CALCULATE THE WINDOW SIZE



                        char=get_column_letter(4)
                        ws[char+str(3)]="Door"
                        char=get_column_letter(5)
                        ws[char+str(3)]=uvalue_2018["door_L"]
                        char=get_column_letter(6)
                        ws[char+str(3)]=uvalue_2018["door_leti"]
                        char=get_column_letter(7)
                        ws[char+str(3)]=uvalue_2018["door_L"]-uvalue_2018["door_leti"]
                        char=get_column_letter(8)
                        ws[char+str(3)]=(math.sqrt(float(floorarea))*heightt)*(0.40)
                        char=get_column_letter(9)
                        ws[char+str(3)]=(((math.sqrt(float(floorarea))*heightt)*(0.40)) * uvalue_2018["door_L"])
                        char=get_column_letter(10)
                        ws[char+str(3)]=((math.sqrt(float(floorarea))*heightt)*(0.40)) * uvalue_2018["door_leti"]
                        char=get_column_letter(11)
                        ws[char+str(3)]=((((math.sqrt(float(floorarea))*heightt)*(0.40)) * uvalue_2018["door_L"]) - (((math.sqrt(float(floorarea))*heightt)*(0.40)) * uvalue_2018["door_leti"]))
                        char=get_column_letter(12)
                        ws[char+str(3)]=((((math.sqrt(float(floorarea))*heightt)*(0.40)) * uvalue_2018["door_L"]) - (((math.sqrt(float(floorarea))*heightt)*(0.40)) * uvalue_2018["door_leti"])) / ((((math.sqrt(float(floorarea))*heightt)*(0.40)) * uvalue_2018["door_L"]))

                        char=get_column_letter(4)
                        ws[char+str(4)]="Roof"
                        char=get_column_letter(5)
                        ws[char+str(4)]=uvalue_2018["roof_L"]
                        char=get_column_letter(6)
                        ws[char+str(4)]=uvalue_2018["roof_leti"]
                        char=get_column_letter(7)
                        ws[char+str(4)]=uvalue_2018["roof_L"]-uvalue_2018["roof_leti"]
                        char=get_column_letter(8)
                        ws[char+str(4)]=(float(floorarea))
                        char=get_column_letter(9)
                        ws[char+str(4)]=((float(floorarea)) * uvalue_2018["roof_L"])
                        char=get_column_letter(10)
                        ws[char+str(4)]=(float(floorarea)) * uvalue_2018["roof_leti"]
                        char=get_column_letter(11)
                        ws[char+str(4)]=(((float(floorarea)) * uvalue_2018["roof_L"]) - ((float(floorarea)) * uvalue_2018["roof_leti"]))
                        char=get_column_letter(12)
                        ws[char+str(4)]=(((float(floorarea)) * uvalue_2018["roof_L"]) - ((float(floorarea)) * uvalue_2018["roof_leti"])) / ((float(floorarea)) * uvalue_2018["roof_L"])


                        char=get_column_letter(4)
                        ws[char+str(5)]="Wall"
                        char=get_column_letter(5)
                        ws[char+str(5)]=uvalue_2018["wall_L"]
                        char=get_column_letter(6)
                        ws[char+str(5)]=uvalue_2018["wall_leti"]
                        char=get_column_letter(7)
                        ws[char+str(5)]=uvalue_2018["wall_L"]-uvalue_2018["wall_leti"]
                        char=get_column_letter(8)
                        ws[char+str(5)]=(math.sqrt(float(floorarea))*heightt)*(0.80)
                        char=get_column_letter(9)
                        ws[char+str(5)]=(((math.sqrt(float(floorarea))*heightt)*(0.80)) * uvalue_2018["wall_L"])
                        char=get_column_letter(10)
                        ws[char+str(5)]=((math.sqrt(float(floorarea))*heightt)*(0.80)) * uvalue_2018["wall_leti"]
                        char=get_column_letter(11)
                        ws[char+str(5)]=((((math.sqrt(float(floorarea))*heightt)*(0.80)) * uvalue_2018["wall_L"]) - (((math.sqrt(float(floorarea))*heightt)*(0.80)) * uvalue_2018["wall_leti"]))
                        char=get_column_letter(12)
                        ws[char+str(5)]=((((math.sqrt(float(floorarea))*heightt)*(0.80)) * uvalue_2018["wall_L"]) - (((math.sqrt(float(floorarea))*heightt)*(0.80)) * uvalue_2018["wall_leti"])) / (((math.sqrt(float(floorarea))*heightt)*(0.80)) * uvalue_2018["wall_leti"]  )

                        char=get_column_letter(4)
                        ws[char+str(6)]="Floor"
                        char=get_column_letter(5)
                        ws[char+str(6)]=uvalue_2018["floor_L"]
                        char=get_column_letter(6)
                        ws[char+str(6)]=uvalue_2018["floor_leti"]
                        char=get_column_letter(7)
                        ws[char+str(6)]=uvalue_2018["floor_L"]-uvalue_2018["floor_leti"]
                        char=get_column_letter(8)
                        ws[char+str(6)]=(float(floorarea))
                        char=get_column_letter(9)
                        ws[char+str(6)]=((float(floorarea)) * uvalue_2018["floor_L"])
                        char=get_column_letter(10)
                        ws[char+str(6)]=(float(floorarea)) * uvalue_2018["floor_leti"]
                        char=get_column_letter(11)
                        ws[char+str(6)]=(((float(floorarea)) * uvalue_2018["floor_L"]) - ((float(floorarea)) * uvalue_2018["floor_leti"]))
                        char=get_column_letter(12)
                        ws[char+str(6)]= (((float(floorarea)) * uvalue_2018["floor_L"]) - ((float(floorarea)) * uvalue_2018["floor_leti"])) / ((float(floorarea)) * uvalue_2018["floor_L"])

                        char=get_column_letter(8)
                        ws[char+str(7)]="Total U-Value"

                        ws['L2']='=(K2/I2)'
                        ws['L3']='=(K3/I3)'
                        ws['L4']='=(K4/I4)'
                        ws['L5']='=(K5/I5)'
                        ws['L6']='=(K6/I6)'
                        ws['L7']='=(K7/I7)'
                        ws['I7']='=SUM(I2:I6)'
                        ws['I8']=f"=UserInputs!B{propertyrow}+UserInputs!C{propertyrow}"
                        ws['J7']='=SUM(J2:J6)'
                        ws['K7']='=SUM(K2:K6)'
                        ws['L8']='=(L7)*I8'


                        #const of retrofit * area
                        ws['S2']='=(H2)*O2'
                        ws['S3']='=(H3)*O3'
                        ws['S4']='=(H4)*O4'
                        ws['S5']='=(H5)*O5'
                        ws['S6']='=(H6)*O6'

                        #TOTAL SAVINGS

                        char=get_column_letter(11)
                        ws[char+str(8)]="Total Savings"

                        #Energy Savings
                        ws['T2']='=((((K2)*(Portfolio!D50))/1000)*365*24*(Portfolio!D67))'
                        ws['T3']='=((((K3)*(Portfolio!D50))/1000)*365*24*(Portfolio!D67))'
                        ws['T4']='=((((K4)*(Portfolio!D50))/1000)*365*24*(Portfolio!D67))'
                        ws['T5']='=((((K5)*(Portfolio!D50))/1000)*365*24*(Portfolio!D67))'
                        ws['T6']='=((((K6)*(Portfolio!D50))/1000)*365*24*(Portfolio!D67))'
                        ws['T7']='=SUM(T2:T6)'


                        ws['U2']=f'=((((K2)*(Portfolio!D50))/1000)*365*24)/12'
                        ws['U3']=f'=((((K3)*(Portfolio!D50))/1000)*365*24)/12'
                        ws['U4']=f'=((((K4)*(Portfolio!D50))/1000)*365*24)/12'
                        ws['U5']=f'=((((K5)*(Portfolio!D50))/1000)*365*24)/12'
                        ws['U6']=f'=((((K6)*(Portfolio!D50))/1000)*365*24)/12'
                        ws['U7']='=SUM(U2:U6)'


                        if(add_to_portfolio==True):

                            #TO ADD TO PORTFOLIO

                            #difference u value
                            char=get_column_letter(11)
                            portfolio[char+str(propertyrow)]=uvalue_2018["window_L"]-uvalue_2018["window_leti"]
                            char=get_column_letter(10)
                            portfolio[char+str(propertyrow)]=uvalue_2018["door_L"]-uvalue_2018["door_leti"]
                            char=get_column_letter(12)
                            portfolio[char+str(propertyrow)]=uvalue_2018["roof_L"]-uvalue_2018["roof_leti"]
                            char=get_column_letter(13)
                            portfolio[char+str(propertyrow)]=uvalue_2018["wall_L"]-uvalue_2018["wall_leti"]
                            char=get_column_letter(14)
                            portfolio[char+str(propertyrow)]=uvalue_2018["floor_L"]-uvalue_2018["floor_leti"]

                            #area
                            char=get_column_letter(16)
                            portfolio[char+str(propertyrow)]=(math.sqrt(floorarea)*heightt)*(0.20)
                            char=get_column_letter(15)
                            portfolio[char+str(propertyrow)]=(((math.sqrt(floorarea)*heightt)*(0.20))*(0.20))
                            char=get_column_letter(17)
                            portfolio[char+str(propertyrow)]=(floorarea)
                            char=get_column_letter(18)
                            portfolio[char+str(propertyrow)]=(math.sqrt(floorarea)*heightt)*(0.80)
                            char=get_column_letter(19)
                            portfolio[char+str(propertyrow)]=(floorarea)

                            #improvement
                            char=get_column_letter(21)
                            portfolio[char+str(propertyrow)]=(((math.sqrt(floorarea)*heightt)*(0.20)) * uvalue_2018["window_L"]) - (((math.sqrt(floorarea)*heightt)*(0.20)) * uvalue_2018["window_leti"])
                            char=get_column_letter(20)
                            portfolio[char+str(propertyrow)]=((((((math.sqrt(floorarea)*heightt)*(0.20))*(0.20))) * uvalue_2018["door_L"]) - (((((math.sqrt(floorarea)*heightt)*(0.20))*(0.20))) * uvalue_2018["door_leti"]))
                            char=get_column_letter(22)
                            portfolio[char+str(propertyrow)]=(((floorarea) * uvalue_2018["roof_L"]) - ((floorarea) * uvalue_2018["roof_leti"]))
                            char=get_column_letter(23)
                            portfolio[char+str(propertyrow)]=((((math.sqrt(floorarea)*heightt)*(0.80)) * uvalue_2018["wall_L"]) - (((math.sqrt(floorarea)*heightt)*(0.80)) * uvalue_2018["wall_leti"]))
                            char=get_column_letter(24)
                            portfolio[char+str(propertyrow)]=(((floorarea) * uvalue_2018["floor_L"]) - ((floorarea) * uvalue_2018["floor_leti"]))
                            char=get_column_letter(25)
                            portfolio[char+str(propertyrow)]=(((((math.sqrt(floorarea)*heightt)*(0.20)) * uvalue_2018["window_L"]) - (((math.sqrt(floorarea)*heightt)*(0.20)) * uvalue_2018["window_leti"])) + ((((((math.sqrt(floorarea)*heightt)*(0.20))*(0.20))) * uvalue_2018["door_L"]) - (((((math.sqrt(floorarea)*heightt)*(0.20))*(0.20))) * uvalue_2018["door_leti"])) + (((floorarea) * uvalue_2018["roof_L"]) - ((floorarea) * uvalue_2018["roof_leti"])) + ((((math.sqrt(floorarea)*heightt)*(0.80)) * uvalue_2018["wall_L"]) - (((math.sqrt(floorarea)*heightt)*(0.80)) * uvalue_2018["wall_leti"])) + (((floorarea) * uvalue_2018["floor_L"]) - ((floorarea) * uvalue_2018["floor_leti"])))


                            #savings
                            char=get_column_letter(27)
                            portfolio[char+str(propertyrow)]=((((math.sqrt(floorarea)*heightt)*(0.20)) * uvalue_2018["window_L"]) - (((math.sqrt(floorarea)*heightt)*(0.20)) * uvalue_2018["window_leti"])) / ((((math.sqrt(floorarea)*heightt)*(0.20)) * uvalue_2018["window_L"]))
                            char=get_column_letter(26)
                            portfolio[char+str(propertyrow)]=((((((math.sqrt(floorarea)*heightt)*(0.20))*(0.20))) * uvalue_2018["door_L"]) - (((((math.sqrt(floorarea)*heightt)*(0.20))*(0.20))) * uvalue_2018["door_leti"])) / ((((((math.sqrt(floorarea)*heightt)*(0.20))*(0.20))) * uvalue_2018["door_L"]))
                            char=get_column_letter(28)
                            portfolio[char+str(propertyrow)]=(((floorarea) * uvalue_2018["roof_L"]) - ((floorarea) * uvalue_2018["roof_leti"])) / ((floorarea) * uvalue_2018["roof_L"])
                            char=get_column_letter(29)
                            portfolio[char+str(propertyrow)]=((((math.sqrt(floorarea)*heightt)*(0.80)) * uvalue_2018["wall_L"]) - (((math.sqrt(floorarea)*heightt)*(0.80)) * uvalue_2018["wall_leti"])) / (((math.sqrt(floorarea)*heightt)*(0.80)) * uvalue_2018["wall_L"]  )
                            char=get_column_letter(30)
                            portfolio[char+str(propertyrow)]=(((floorarea) * uvalue_2018["floor_L"]) - ((floorarea) * uvalue_2018["floor_leti"])) / ((floorarea) * uvalue_2018["floor_L"])
                            char=get_column_letter(31)
                            portfolio[char+str(propertyrow)]=(((((math.sqrt(floorarea)*heightt)*(0.20)) * uvalue_2018["window_L"]) - (((math.sqrt(floorarea)*heightt)*(0.20)) * uvalue_2018["window_leti"])) + ((((((math.sqrt(floorarea)*heightt)*(0.20))*(0.20))) * uvalue_2018["door_L"]) - (((((math.sqrt(floorarea)*heightt)*(0.20))*(0.20))) * uvalue_2018["door_leti"])) + (((floorarea) * uvalue_2018["roof_L"]) - ((floorarea) * uvalue_2018["roof_leti"])) + ((((math.sqrt(floorarea)*heightt)*(0.80)) * uvalue_2018["wall_L"]) - (((math.sqrt(floorarea)*heightt)*(0.80)) * uvalue_2018["wall_leti"])) + (((floorarea) * uvalue_2018["floor_L"]) - ((floorarea) * uvalue_2018["floor_leti"]))) / ((((math.sqrt(floorarea)*heightt)*(0.20)) * uvalue_2018["window_L"]) + (((((math.sqrt(floorarea)*heightt)*(0.20))*(0.20))) * uvalue_2018["door_L"]) + ((floorarea) * uvalue_2018["roof_L"]) + (((math.sqrt(floorarea)*heightt)*(0.80)) * uvalue_2018["wall_L"]) + ((floorarea) * uvalue_2018["floor_L"]))

                            #FINAL SAVINGS
                            char=get_column_letter(32)
                            portfolio[char+str(propertyrow)]='='+str(uprn)+'!L8*12'

                            #cost of retrofit per meter square for all constituents
                            char=get_column_letter(33)
                            portfolio[get_column_letter(41)+str(propertyrow)]='=SUM('+str(uprn)+'!S2:S6)'

                            up=str(uprn)

                            char=get_column_letter(51)
                            portfolio[char+str(propertyrow)]=f"={up}!L7*{buildingEmissions}"
                            char=get_column_letter(52)
                            portfolio[char+str(propertyrow)]=f"={up}!T2"
                            char=get_column_letter(53)
                            portfolio[char+str(propertyrow)]=f"={up}!T3"
                            char=get_column_letter(54)
                            portfolio[char+str(propertyrow)]=f"={up}!T4"
                            char=get_column_letter(55)
                            portfolio[char+str(propertyrow)]=f"={up}!T5"
                            char=get_column_letter(56)
                            portfolio[char+str(propertyrow)]=f"={up}!T6"
                            char=get_column_letter(57)
                            portfolio[char+str(propertyrow)]=f"={up}!U7"
                            char=get_column_letter(58)
                            portfolio[char+str(propertyrow)]=f"=INDEX({up}!U2:U6, MATCH(MAX({up}!L2:L6), {up}!L2:L6, 0))"
                            char=get_column_letter(59)
                            portfolio[char+str(propertyrow)]=f"=(INDEX({up}!H2:H6, MATCH(MAX({up}!L2:L6), {up}!L2:L6, 0))/{up}!B36)*{up}!B1"
                            char=get_column_letter(60)
                            portfolio[char+str(propertyrow)]=f"=INDEX({up}!T2:T6, MATCH(MAX({up}!L2:L6), {up}!L2:L6, 0))"
                            char=get_column_letter(61)
                            portfolio[char+str(propertyrow)]=f"={up}!I7"

                        #cost of constituent outputs
                        #cost of constituent outputs
                        char=get_column_letter(13)
                        ws[char+str(2)]=f"=Portfolio!D51"
                        char=get_column_letter(13)
                        ws[char+str(3)]=f"=Portfolio!D52"
                        char=get_column_letter(13)
                        ws[char+str(4)]=f"=Portfolio!D53"
                        char=get_column_letter(13)
                        ws[char+str(5)]=f"=Portfolio!D54"
                        char=get_column_letter(13)
                        ws[char+str(6)]=f"=Portfolio!D55"

                        char=get_column_letter(15)
                        ws[char+str(2)]=f"=Portfolio!D56"
                        char=get_column_letter(15)
                        ws[char+str(3)]=f"=Portfolio!D57"
                        char=get_column_letter(15)
                        ws[char+str(4)]=f"=Portfolio!D58"
                        char=get_column_letter(15)
                        ws[char+str(5)]=f"=Portfolio!D59"
                        char=get_column_letter(15)
                        ws[char+str(6)]=f"=Portfolio!D60"

                        char=get_column_letter(17)
                        ws[char+str(2)]=f"=Portfolio!D61"
                        char=get_column_letter(17)
                        ws[char+str(3)]=f"=Portfolio!D62"
                        char=get_column_letter(17)
                        ws[char+str(4)]=f"=Portfolio!D63"
                        char=get_column_letter(17)
                        ws[char+str(5)]=f"=Portfolio!D64"
                        char=get_column_letter(17)
                        ws[char+str(6)]=f"=Portfolio!D65"


                        ws['N2']='=(M2/K2)'
                        ws['N3']='=(M3/K3)'
                        ws['N4']='=(M4/K4)'
                        ws['N5']='=(M5/K5)'
                        ws['N6']='=(M6/K6)'

                        ws['P2']='=(O2/K2)'
                        ws['P3']='=(O3/K3)'
                        ws['P4']='=(O4/K4)'
                        ws['P5']='=(O5/K5)'
                        ws['P6']='=(O6/K6)'

                        ws['R2']='=(Q2/K2)'
                        ws['R3']='=(Q3/K3)'
                        ws['R4']='=(Q4/K4)'
                        ws['R5']='=(Q5/K5)'
                        ws['R6']='=(Q6/K6)'
                    if(inspectionDate[0:4]=="2019"):


                        char=get_column_letter(4)
                        ws[char+str(1)]="Building Attribute"
                        char=get_column_letter(5)
                        ws[char+str(1)]="U_Value L-Doc (W/m^2K)"
                        char=get_column_letter(6)
                        ws[char+str(1)]="U_Value Leti (W/m^2K)"
                        char=get_column_letter(7)
                        ws[char+str(1)]="U_Value Difference (W/m^2K)"
                        char=get_column_letter(8)
                        ws[char+str(1)]="Area"
                        char=get_column_letter(9)
                        ws[char+str(1)]="U_value L-Doc Based on Area"
                        char=get_column_letter(10)
                        ws[char+str(1)]="U_value Leti Based on Area"
                        char=get_column_letter(11)
                        ws[char+str(1)]="U-value Improvement Based on Area"
                        char=get_column_letter(12)
                        ws[char+str(1)]="Savings"
                        #code for cost of retrofit per contituent
                        char=get_column_letter(13)
                        ws[char+str(1)]="low Cost of Retrofit Constituent (m^2)"
                        char=get_column_letter(14)
                        ws[char+str(1)]="Cost of Retrofit Constituent (m^2) / Improvement based on area"
                        char=get_column_letter(15)
                        ws[char+str(1)]="Avg. Cost of Retrofit Constituent (m^2)"
                        char=get_column_letter(16)
                        ws[char+str(1)]="Cost of Retrofit Constituent (m^2) / Improvement based on area"
                        char=get_column_letter(17)
                        ws[char+str(1)]="High Cost of Retrofit Constituent (m^2)"
                        char=get_column_letter(18)
                        ws[char+str(1)]="Cost of Retrofit Constituent (m^2) / Improvement based on area"
                        char=get_column_letter(19)
                        ws[char+str(1)]="Cost of Retrofit Constituent (m^2) * area"
                        char=get_column_letter(19)
                        ws[char+str(1)]="Cost of Retrofit Constituent (m^2) * area"
                        char=get_column_letter(20)
                        ws[char+str(1)]="Energy Savings (GBP)"
                        char=get_column_letter(21)
                        ws[char+str(1)]="Energy Savings (kWh)"

                        char=get_column_letter(4)
                        ws[char+str(2)]="Window"
                        char=get_column_letter(5)
                        ws[char+str(2)]=uvalue_2019["window_L"]
                        char=get_column_letter(6)
                        ws[char+str(2)]=uvalue_2019["window_leti"]
                        char=get_column_letter(7)
                        ws[char+str(2)]=uvalue_2019["window_L"]-uvalue_2019["window_leti"]
                        char=get_column_letter(8)
                        ws[char+str(2)]=(math.sqrt(float(floorarea))*heightt)*(0.20)
                        char=get_column_letter(9)
                        ws[char+str(2)]=(((math.sqrt(float(floorarea))*heightt)*(0.20)) * uvalue_2019["window_L"])
                        char=get_column_letter(10)
                        ws[char+str(2)]=((math.sqrt(float(floorarea))*heightt)*(0.20)) * uvalue_2019["window_leti"]
                        char=get_column_letter(11)
                        ws[char+str(2)]=(((math.sqrt(float(floorarea))*heightt)*(0.20)) * uvalue_2019["window_L"]) - (((math.sqrt(float(floorarea))*heightt)*(0.20)) * uvalue_2019["window_leti"])
                        char=get_column_letter(12)
                        ws[char+str(2)]=((((math.sqrt(float(floorarea))*heightt)*(0.20)) * uvalue_2019["window_L"]) - (((math.sqrt(float(floorarea))*heightt)*(0.20)) * uvalue_2019["window_leti"])) / ((((math.sqrt(float(floorarea))*heightt)*(0.20)) * uvalue_2019["window_L"]))
                        #BuildingData.json()['data'][0]['height']) THIS IS THE BUILDINGS HEIGHT FROM WHICH WE CANNOT CALCULATE THE WINDOW SIZE



                        char=get_column_letter(4)
                        ws[char+str(3)]="Door"
                        char=get_column_letter(5)
                        ws[char+str(3)]=uvalue_2019["door_L"]
                        char=get_column_letter(6)
                        ws[char+str(3)]=uvalue_2019["door_leti"]
                        char=get_column_letter(7)
                        ws[char+str(3)]=uvalue_2019["door_L"]-uvalue_2019["door_leti"]
                        char=get_column_letter(8)
                        ws[char+str(3)]=(math.sqrt(float(floorarea))*heightt)*(0.40)
                        char=get_column_letter(9)
                        ws[char+str(3)]=(((math.sqrt(float(floorarea))*heightt)*(0.40)) * uvalue_2019["door_L"])
                        char=get_column_letter(10)
                        ws[char+str(3)]=((math.sqrt(float(floorarea))*heightt)*(0.40)) * uvalue_2019["door_leti"]
                        char=get_column_letter(11)
                        ws[char+str(3)]=((((math.sqrt(float(floorarea))*heightt)*(0.40)) * uvalue_2019["door_L"]) - (((math.sqrt(float(floorarea))*heightt)*(0.40)) * uvalue_2019["door_leti"]))
                        char=get_column_letter(12)
                        ws[char+str(3)]=((((math.sqrt(float(floorarea))*heightt)*(0.40)) * uvalue_2019["door_L"]) - (((math.sqrt(float(floorarea))*heightt)*(0.40)) * uvalue_2019["door_leti"])) / ((((math.sqrt(float(floorarea))*heightt)*(0.40)) * uvalue_2019["door_L"]))


                        char=get_column_letter(4)
                        ws[char+str(4)]="Roof"
                        char=get_column_letter(5)
                        ws[char+str(4)]=uvalue_2019["roof_L"]
                        char=get_column_letter(6)
                        ws[char+str(4)]=uvalue_2019["roof_leti"]
                        char=get_column_letter(7)
                        ws[char+str(4)]=uvalue_2019["roof_L"]-uvalue_2019["roof_leti"]
                        char=get_column_letter(8)
                        ws[char+str(4)]=(float(floorarea))
                        char=get_column_letter(9)
                        ws[char+str(4)]=((float(floorarea)) * uvalue_2019["roof_L"])
                        char=get_column_letter(10)
                        ws[char+str(4)]=(float(floorarea)) * uvalue_2019["roof_leti"]
                        char=get_column_letter(11)
                        ws[char+str(4)]=(((float(floorarea)) * uvalue_2019["roof_L"]) - ((float(floorarea)) * uvalue_2019["roof_leti"]))
                        char=get_column_letter(12)
                        ws[char+str(4)]=(((float(floorarea)) * uvalue_2019["roof_L"]) - ((float(floorarea)) * uvalue_2019["roof_leti"])) / ((float(floorarea)) * uvalue_2019["roof_L"])


                        char=get_column_letter(4)
                        ws[char+str(5)]="Wall"
                        char=get_column_letter(5)
                        ws[char+str(5)]=uvalue_2019["wall_L"]
                        char=get_column_letter(6)
                        ws[char+str(5)]=uvalue_2019["wall_leti"]
                        char=get_column_letter(7)
                        ws[char+str(5)]=uvalue_2019["wall_L"]-uvalue_2019["wall_leti"]
                        char=get_column_letter(8)
                        ws[char+str(5)]=(math.sqrt(float(floorarea))*heightt)*(0.80)
                        char=get_column_letter(9)
                        ws[char+str(5)]=(((math.sqrt(float(floorarea))*heightt)*(0.80)) * uvalue_2019["wall_L"])
                        char=get_column_letter(10)
                        ws[char+str(5)]=((math.sqrt(float(floorarea))*heightt)*(0.80)) * uvalue_2019["wall_leti"]
                        char=get_column_letter(11)
                        ws[char+str(5)]=((((math.sqrt(float(floorarea))*heightt)*(0.80)) * uvalue_2019["wall_L"]) - (((math.sqrt(float(floorarea))*heightt)*(0.80)) * uvalue_2019["wall_leti"]))
                        char=get_column_letter(12)
                        ws[char+str(5)]=((((math.sqrt(float(floorarea))*heightt)*(0.80)) * uvalue_2019["wall_L"]) - (((math.sqrt(float(floorarea))*heightt)*(0.80)) * uvalue_2019["wall_leti"])) / (((math.sqrt(float(floorarea))*heightt)*(0.80)) * uvalue_2019["wall_leti"]  )

                        char=get_column_letter(4)
                        ws[char+str(6)]="Floor"
                        char=get_column_letter(5)
                        ws[char+str(6)]=uvalue_2019["floor_L"]
                        char=get_column_letter(6)
                        ws[char+str(6)]=uvalue_2019["floor_leti"]
                        char=get_column_letter(7)
                        ws[char+str(6)]=uvalue_2019["floor_L"]-uvalue_2019["floor_leti"]
                        char=get_column_letter(8)
                        ws[char+str(6)]=(float(floorarea))
                        char=get_column_letter(9)
                        ws[char+str(6)]=((float(floorarea)) * uvalue_2019["floor_L"])
                        char=get_column_letter(10)
                        ws[char+str(6)]=(float(floorarea)) * uvalue_2019["floor_leti"]
                        char=get_column_letter(11)
                        ws[char+str(6)]=(((float(floorarea)) * uvalue_2019["floor_L"]) - ((float(floorarea)) * uvalue_2019["floor_leti"]))
                        char=get_column_letter(12)
                        ws[char+str(6)]= (((float(floorarea)) * uvalue_2019["floor_L"]) - ((float(floorarea)) * uvalue_2019["floor_leti"])) / ((float(floorarea)) * uvalue_2019["floor_L"])

                        char=get_column_letter(8)
                        ws[char+str(7)]="Total U-Value"

                        ws['L2']='=(K2/I2)'
                        ws['L3']='=(K3/I3)'
                        ws['L4']='=(K4/I4)'
                        ws['L5']='=(K5/I5)'
                        ws['L6']='=(K6/I6)'
                        ws['L7']='=(K7/I7)'
                        ws['I7']='=SUM(I2:I6)'
                        ws['I8']=f"=UserInputs!B{propertyrow}+UserInputs!C{propertyrow}"
                        ws['J7']='=SUM(J2:J6)'
                        ws['K7']='=SUM(K2:K6)'
                        ws['L8']='=(L7)*I8'


                        #const of retrofit * area
                        ws['S2']='=(H2)*O2'
                        ws['S3']='=(H3)*O3'
                        ws['S4']='=(H4)*O4'
                        ws['S5']='=(H5)*O5'
                        ws['S6']='=(H6)*O6'

                        #TOTAL SAVINGS

                        char=get_column_letter(11)
                        ws[char+str(8)]="Total Savings"

                        #Energy Savings
                        ws['T2']='=((((K2)*(Portfolio!D50))/1000)*365*24*(Portfolio!D67))'
                        ws['T3']='=((((K3)*(Portfolio!D50))/1000)*365*24*(Portfolio!D67))'
                        ws['T4']='=((((K4)*(Portfolio!D50))/1000)*365*24*(Portfolio!D67))'
                        ws['T5']='=((((K5)*(Portfolio!D50))/1000)*365*24*(Portfolio!D67))'
                        ws['T6']='=((((K6)*(Portfolio!D50))/1000)*365*24*(Portfolio!D67))'
                        ws['T7']='=SUM(T2:T6)'


                        ws['U2']=f'=((((K2)*(Portfolio!D50))/1000)*365*24)/12'
                        ws['U3']=f'=((((K3)*(Portfolio!D50))/1000)*365*24)/12'
                        ws['U4']=f'=((((K4)*(Portfolio!D50))/1000)*365*24)/12'
                        ws['U5']=f'=((((K5)*(Portfolio!D50))/1000)*365*24)/12'
                        ws['U6']=f'=((((K6)*(Portfolio!D50))/1000)*365*24)/12'
                        ws['U7']='=SUM(U2:U6)'




                        if(add_to_portfolio==True):
                            #TO ADD TO PORTFOLIO

                            #difference u value
                            char=get_column_letter(11)
                            portfolio[char+str(propertyrow)]=uvalue_2019["window_L"]-uvalue_2019["window_leti"]
                            char=get_column_letter(10)
                            portfolio[char+str(propertyrow)]=uvalue_2019["door_L"]-uvalue_2019["door_leti"]
                            char=get_column_letter(12)
                            portfolio[char+str(propertyrow)]=uvalue_2019["roof_L"]-uvalue_2019["roof_leti"]
                            char=get_column_letter(13)
                            portfolio[char+str(propertyrow)]=uvalue_2019["wall_L"]-uvalue_2019["wall_leti"]
                            char=get_column_letter(14)
                            portfolio[char+str(propertyrow)]=uvalue_2019["floor_L"]-uvalue_2019["floor_leti"]

                            #area
                            char=get_column_letter(16)
                            portfolio[char+str(propertyrow)]=(math.sqrt(floorarea)*heightt)*(0.20)
                            char=get_column_letter(15)
                            portfolio[char+str(propertyrow)]=(((math.sqrt(floorarea)*heightt)*(0.20))*(0.20))
                            char=get_column_letter(17)
                            portfolio[char+str(propertyrow)]=(floorarea)
                            char=get_column_letter(18)
                            portfolio[char+str(propertyrow)]=(math.sqrt(floorarea)*heightt)*(0.80)
                            char=get_column_letter(19)
                            portfolio[char+str(propertyrow)]=(floorarea)

                            #improvement
                            char=get_column_letter(21)
                            portfolio[char+str(propertyrow)]=(((math.sqrt(floorarea)*heightt)*(0.20)) * uvalue_2019["window_L"]) - (((math.sqrt(floorarea)*heightt)*(0.20)) * uvalue_2019["window_leti"])
                            char=get_column_letter(20)
                            portfolio[char+str(propertyrow)]=((((((math.sqrt(floorarea)*heightt)*(0.20))*(0.20))) * uvalue_2019["door_L"]) - (((((math.sqrt(floorarea)*heightt)*(0.20))*(0.20))) * uvalue_2019["door_leti"]))
                            char=get_column_letter(22)
                            portfolio[char+str(propertyrow)]=(((floorarea) * uvalue_2019["roof_L"]) - ((floorarea) * uvalue_2019["roof_leti"]))
                            char=get_column_letter(23)
                            portfolio[char+str(propertyrow)]=((((math.sqrt(floorarea)*heightt)*(0.80)) * uvalue_2019["wall_L"]) - (((math.sqrt(floorarea)*heightt)*(0.80)) * uvalue_2019["wall_leti"]))
                            char=get_column_letter(24)
                            portfolio[char+str(propertyrow)]=(((floorarea) * uvalue_2019["floor_L"]) - ((floorarea) * uvalue_2019["floor_leti"]))
                            char=get_column_letter(25)
                            portfolio[char+str(propertyrow)]=(((((math.sqrt(floorarea)*heightt)*(0.20)) * uvalue_2019["window_L"]) - (((math.sqrt(floorarea)*heightt)*(0.20)) * uvalue_2019["window_leti"])) + ((((((math.sqrt(floorarea)*heightt)*(0.20))*(0.20))) * uvalue_2019["door_L"]) - (((((math.sqrt(floorarea)*heightt)*(0.20))*(0.20))) * uvalue_2019["door_leti"])) + (((floorarea) * uvalue_2019["roof_L"]) - ((floorarea) * uvalue_2019["roof_leti"])) + ((((math.sqrt(floorarea)*heightt)*(0.80)) * uvalue_2019["wall_L"]) - (((math.sqrt(floorarea)*heightt)*(0.80)) * uvalue_2019["wall_leti"])) + (((floorarea) * uvalue_2019["floor_L"]) - ((floorarea) * uvalue_2019["floor_leti"])))


                            #savings
                            char=get_column_letter(27)
                            portfolio[char+str(propertyrow)]=((((math.sqrt(floorarea)*heightt)*(0.20)) * uvalue_2019["window_L"]) - (((math.sqrt(floorarea)*heightt)*(0.20)) * uvalue_2019["window_leti"])) / ((((math.sqrt(floorarea)*heightt)*(0.20)) * uvalue_2019["window_L"]))
                            char=get_column_letter(26)
                            portfolio[char+str(propertyrow)]=((((((math.sqrt(floorarea)*heightt)*(0.20))*(0.20))) * uvalue_2019["door_L"]) - (((((math.sqrt(floorarea)*heightt)*(0.20))*(0.20))) * uvalue_2019["door_leti"])) / ((((((math.sqrt(floorarea)*heightt)*(0.20))*(0.20))) * uvalue_2019["door_L"]))
                            char=get_column_letter(28)
                            portfolio[char+str(propertyrow)]=(((floorarea) * uvalue_2019["roof_L"]) - ((floorarea) * uvalue_2019["roof_leti"])) / ((floorarea) * uvalue_2019["roof_L"])
                            char=get_column_letter(29)
                            portfolio[char+str(propertyrow)]=((((math.sqrt(floorarea)*heightt)*(0.80)) * uvalue_2019["wall_L"]) - (((math.sqrt(floorarea)*heightt)*(0.80)) * uvalue_2019["wall_leti"])) / (((math.sqrt(floorarea)*heightt)*(0.80)) * uvalue_2019["wall_L"]  )
                            char=get_column_letter(30)
                            portfolio[char+str(propertyrow)]=(((floorarea) * uvalue_2019["floor_L"]) - ((floorarea) * uvalue_2019["floor_leti"])) / ((floorarea) * uvalue_2019["floor_L"])
                            char=get_column_letter(31)
                            portfolio[char+str(propertyrow)]=(((((math.sqrt(floorarea)*heightt)*(0.20)) * uvalue_2019["window_L"]) - (((math.sqrt(floorarea)*heightt)*(0.20)) * uvalue_2019["window_leti"])) + ((((((math.sqrt(floorarea)*heightt)*(0.20))*(0.20))) * uvalue_2019["door_L"]) - (((((math.sqrt(floorarea)*heightt)*(0.20))*(0.20))) * uvalue_2019["door_leti"])) + (((floorarea) * uvalue_2019["roof_L"]) - ((floorarea) * uvalue_2019["roof_leti"])) + ((((math.sqrt(floorarea)*heightt)*(0.80)) * uvalue_2019["wall_L"]) - (((math.sqrt(floorarea)*heightt)*(0.80)) * uvalue_2019["wall_leti"])) + (((floorarea) * uvalue_2019["floor_L"]) - ((floorarea) * uvalue_2019["floor_leti"]))) / ((((math.sqrt(floorarea)*heightt)*(0.20)) * uvalue_2019["window_L"]) + (((((math.sqrt(floorarea)*heightt)*(0.20))*(0.20))) * uvalue_2019["door_L"]) + ((floorarea) * uvalue_2019["roof_L"]) + (((math.sqrt(floorarea)*heightt)*(0.80)) * uvalue_2019["wall_L"]) + ((floorarea) * uvalue_2019["floor_L"]))


                            #FINAL SAVINGS
                            char=get_column_letter(32)
                            portfolio[char+str(propertyrow)]='='+str(uprn)+'!L8*12'

                            #cost of retrofit per meter square for all constituents
                            char=get_column_letter(33)
                            portfolio[get_column_letter(41)+str(propertyrow)]='=SUM('+str(uprn)+'!S2:S6)'

                            up=str(uprn)

                            char=get_column_letter(51)
                            portfolio[char+str(propertyrow)]=f"={up}!L7*{buildingEmissions}"
                            char=get_column_letter(52)
                            portfolio[char+str(propertyrow)]=f"={up}!T2"
                            char=get_column_letter(53)
                            portfolio[char+str(propertyrow)]=f"={up}!T3"
                            char=get_column_letter(54)
                            portfolio[char+str(propertyrow)]=f"={up}!T4"
                            char=get_column_letter(55)
                            portfolio[char+str(propertyrow)]=f"={up}!T5"
                            char=get_column_letter(56)
                            portfolio[char+str(propertyrow)]=f"={up}!T6"
                            char=get_column_letter(57)
                            portfolio[char+str(propertyrow)]=f"={up}!U7"
                            char=get_column_letter(58)
                            portfolio[char+str(propertyrow)]=f"=INDEX({up}!U2:U6, MATCH(MAX({up}!L2:L6), {up}!L2:L6, 0))"
                            char=get_column_letter(59)
                            portfolio[char+str(propertyrow)]=f"=INDEX({up}!L2:L6, MATCH(MAX({up}!L2:L6), {up}!L2:L6, 0))*{up}!B1"
                            char=get_column_letter(60)
                            portfolio[char+str(propertyrow)]=f"=INDEX({up}!T2:T6, MATCH(MAX({up}!L2:L6), {up}!L2:L6, 0))"
                            char=get_column_letter(61)
                            portfolio[char+str(propertyrow)]=f"={up}!I7"

                        #cost of constituent outputs
                        char=get_column_letter(13)
                        ws[char+str(2)]=f"=Portfolio!D51"
                        char=get_column_letter(13)
                        ws[char+str(3)]=f"=Portfolio!D52"
                        char=get_column_letter(13)
                        ws[char+str(4)]=f"=Portfolio!D53"
                        char=get_column_letter(13)
                        ws[char+str(5)]=f"=Portfolio!D54"
                        char=get_column_letter(13)
                        ws[char+str(6)]=f"=Portfolio!D55"

                        char=get_column_letter(15)
                        ws[char+str(2)]=f"=Portfolio!D56"
                        char=get_column_letter(15)
                        ws[char+str(3)]=f"=Portfolio!D57"
                        char=get_column_letter(15)
                        ws[char+str(4)]=f"=Portfolio!D58"
                        char=get_column_letter(15)
                        ws[char+str(5)]=f"=Portfolio!D59"
                        char=get_column_letter(15)
                        ws[char+str(6)]=f"=Portfolio!D60"

                        char=get_column_letter(17)
                        ws[char+str(2)]=f"=Portfolio!D61"
                        char=get_column_letter(17)
                        ws[char+str(3)]=f"=Portfolio!D62"
                        char=get_column_letter(17)
                        ws[char+str(4)]=f"=Portfolio!D63"
                        char=get_column_letter(17)
                        ws[char+str(5)]=f"=Portfolio!D64"
                        char=get_column_letter(17)
                        ws[char+str(6)]=f"=Portfolio!D65"


                        ws['N2']='=(M2/K2)'
                        ws['N3']='=(M3/K3)'
                        ws['N4']='=(M4/K4)'
                        ws['N5']='=(M5/K5)'
                        ws['N6']='=(M6/K6)'

                        ws['P2']='=(O2/K2)'
                        ws['P3']='=(O3/K3)'
                        ws['P4']='=(O4/K4)'
                        ws['P5']='=(O5/K5)'
                        ws['P6']='=(O6/K6)'

                        ws['R2']='=(Q2/K2)'
                        ws['R3']='=(Q3/K3)'
                        ws['R4']='=(Q4/K4)'
                        ws['R5']='=(Q5/K5)'
                        ws['R6']='=(Q6/K6)'


                    if(add_to_portfolio==True):
                        propertyy=str(propertyrow)
                        char=get_column_letter(33)
                        portfolio[char+str(propertyrow)]='=(0.84-D42)*AF'+str(propertyrow)
                        char=get_column_letter(34)
                        portfolio[char+str(propertyrow)]='=AG'+str(propertyrow)+'/12'

                        char=get_column_letter(35)
                        portfolio[char+str(propertyrow)]=f"=AO{propertyy}*D41"
                        char=get_column_letter(36)

                        portfolio[char+str(propertyrow)]=f"=D42*AF{propertyy}"

                        char=get_column_letter(37)
                        portfolio[char+str(propertyrow)]='=AI'+str(propertyrow)+'*AJ'+str(propertyrow)

                        char=get_column_letter(38)
                        portfolio[char+str(propertyrow)]=f"=AF{propertyy}-(AF{propertyy}*D47)-(((1-D47)*AF{propertyy})*D42)"


                        char=get_column_letter(39)
                        portfolio[char+str(propertyrow)]=f"=AF{propertyy}*D47"

                        char=get_column_letter(40)
                        portfolio[char+str(propertyrow)]=f"=AL{propertyy}*UserInputs!D{propertyy}"


                        #ROI
                        up=str(uprn)
                        char=get_column_letter(42)
                        portfolio[char+str(1)]="High ROI Savings"
                        char=get_column_letter(43)
                        portfolio[char+str(1)]="High ROI savings passed to landlord"
                        char=get_column_letter(44)
                        portfolio[char+str(1)]="High ROI monthly tenant fee"
                        char=get_column_letter(45)
                        portfolio[char+str(1)]="Referral fee on high ROI components"
                        char=get_column_letter(46)
                        portfolio[char+str(1)]="UpGreen Annual income on high ROI components"
                        char=get_column_letter(47)
                        portfolio[char+str(1)]="High ROI savings passed to tenants"
                        char=get_column_letter(48)
                        portfolio[char+str(1)]="High ROI savings passed back for remainder of lease"
                        char=get_column_letter(49)
                        portfolio[char+str(1)]="Cost of retrofit for prioritized components"
                        char=get_column_letter(50)
                        portfolio[char+str(1)]="Constituent to Focus on"
                        #AP
                        char=get_column_letter(42)
                        portfolio[char+str(propertyrow)]=f"=MAX(Z{propertyy}:AD{propertyy})*AF{propertyy}"
                        #AQ
                        char=get_column_letter(43)
                        portfolio[char+str(propertyrow)]=f"=AP{propertyy}*(1-D47)"
                        #AR
                        char=get_column_letter(44)
                        portfolio[char+str(propertyrow)]=f"=AQ{propertyy}/12"
                        #AS
                        char=get_column_letter(45)
                        portfolio[char+str(propertyrow)]=f"=AW{propertyy}*D41"
                        #AT
                        char=get_column_letter(46)
                        portfolio[char+str(propertyrow)]=f"=AS{propertyy}+D42"
                        #AU
                        char=get_column_letter(47)
                        portfolio[char+str(propertyrow)]=f"=AP{propertyy}*D47"
                        #AV
                        char=get_column_letter(48)
                        portfolio[char+str(propertyrow)]=f"=AQ{propertyy}*UserInputs!D{propertyy}"
                        #AW
                        char=get_column_letter(49)
                        portfolio[char+str(propertyrow)]=f"=INDEX({up}!S2:S6, MATCH(MAX({up}!L2:L6), {up}!L2:L6, 0))"
                        #AX
                        char=get_column_letter(50)
                        portfolio[char+str(propertyrow)]=f"=INDEX({up}!D2:D6, MATCH(MAX({up}!L2:L6), {up}!L2:L6, 0))"


                    #print(BuildingData.json())

                    #print(EPCData.json()['data'][0]['inspectionDate'])

            char=get_column_letter(4)
            portfolio[char+str(68)]=f"=SUM(BI2:BI12)/{propertyrow}"


            #COST OF RETROFITTING
            char=get_column_letter(4)
            portfolio[char+str(43)]='=SUM(AO2:AO'+str(propertyrow)+')'

            secondtable_headingrow=properties+1
            secondtable_datarow=properties+2

            # FOR THE FINAL TOTAL FROM PUJAS LOOM VIDEO
            totalproperties=str(propertyrow)

            char=get_column_letter(32)
            portfolio[char+str(secondtable_headingrow)]="TOTAL FINAL SAVINGS (Area-Based)"
            char=get_column_letter(33)
            portfolio[char+str(secondtable_headingrow)]="Annual Saving Passed to Landlord"
            char=get_column_letter(34)
            portfolio[char+str(secondtable_headingrow)]="Monthly tenant Fee"
            char=get_column_letter(35)
            portfolio[char+str(secondtable_headingrow)]="Annual Payback to Landlord"
            char=get_column_letter(36)
            portfolio[char+str(secondtable_headingrow)]="Annual Savings Passed to Tenants"
            char=get_column_letter(37)
            portfolio[char+str(secondtable_headingrow)]="Value from fees passed back to the landlord for the remainder of the lease"
            char=get_column_letter(38)
            portfolio[char+str(secondtable_headingrow)]="Total cost of Retrofitting"
            char=get_column_letter(39)
            portfolio[char+str(secondtable_headingrow)]="Unit economics work? (total value of fees passed from the tenant over the course of the lease > cost of the projects)"
            char=get_column_letter(40)
            portfolio[char+str(secondtable_headingrow)]="Energy Savings(kWh)"
            char=get_column_letter(41)
            portfolio[char+str(secondtable_headingrow)]="Building Emissions Savings (kg co2/m2)"
            char=get_column_letter(42)
            portfolio[char+str(secondtable_headingrow)]="Total Energy Savings (GBP - Energy Based)"
            char=get_column_letter(32)
            portfolio[char+str(secondtable_datarow)]=f"=SUM(AF2:AF{totalproperties})"
            char=get_column_letter(33)
            portfolio[char+str(secondtable_datarow)]=f"=(0.84-D42)*AF{secondtable_datarow}"
            char=get_column_letter(34)
            portfolio[char+str(secondtable_datarow)]=f"=AG{secondtable_datarow}/12"
            char=get_column_letter(35)
            portfolio[char+str(secondtable_datarow)]=f"=AF{secondtable_datarow}-(AF{secondtable_datarow}*D47)-(((1-D47)*AF{secondtable_datarow})*D42)"
            char=get_column_letter(36)
            portfolio[char+str(secondtable_datarow)]=f"=AF{secondtable_datarow}*D47"
            char=get_column_letter(37)
            portfolio[char+str(secondtable_datarow)]=f"=SUM(AN2:AN{totalproperties})"
            char=get_column_letter(38)
            portfolio[char+str(secondtable_datarow)]=f'=SUM(AO2:AO{totalproperties})'
            char=get_column_letter(39)
            portfolio[char+str(secondtable_datarow)]=f'=IF(AK{secondtable_datarow} >= D43, "Yes", "No")'
            char=get_column_letter(40)
            portfolio[char+str(secondtable_datarow)]=f'=SUM(BE2:BE{totalproperties})'
            char=get_column_letter(41)
            portfolio[char+str(secondtable_datarow)]=f'=SUM(AY2:AY{totalproperties})'
            char=get_column_letter(42)
            portfolio[char+str(secondtable_datarow)]=f'=SUM(AZ2:AZ{totalproperties})+SUM(BA2:BA{totalproperties})+SUM(BB2:BB{totalproperties})+SUM(BC2:BC{totalproperties})+SUM(BD2:BD{totalproperties})'


            #Upgreen Calculations
            char=get_column_letter(32)
            portfolio[char+str(secondtable_headingrow+4)]="Referral Fee"
            char=get_column_letter(33)
            portfolio[char+str(secondtable_headingrow+4)]="Cut of Tenant Fee"
            char=get_column_letter(34)
            portfolio[char+str(secondtable_headingrow+4)]="UpGreen Annual Income"
            char=get_column_letter(32)
            portfolio[char+str(secondtable_datarow+4)]=f"=D43*D41"
            char=get_column_letter(33)
            portfolio[char+str(secondtable_datarow+4)]=f"=D42*AF{secondtable_datarow}"
            char=get_column_letter(34)
            portfolio[char+str(secondtable_datarow+4)]=f"=AI{secondtable_datarow} + AJ{secondtable_datarow}"


            #ROI TABLE
            char=get_column_letter(32)
            portfolio[char+str(secondtable_headingrow+2)]="High ROI FINAL SAVINGS"
            char=get_column_letter(33)
            portfolio[char+str(secondtable_headingrow+2)]="Annual Saving Passed to Landlord"
            char=get_column_letter(34)
            portfolio[char+str(secondtable_headingrow+2)]="Monthly tenant Fee"
            char=get_column_letter(35)
            portfolio[char+str(secondtable_headingrow+2)]="Annual Payback to Landlord"
            char=get_column_letter(36)
            portfolio[char+str(secondtable_headingrow+2)]="Annual Savings Passed to Tenants"
            char=get_column_letter(37)
            portfolio[char+str(secondtable_headingrow+2)]="Total fees passed back over lease"
            char=get_column_letter(38)
            portfolio[char+str(secondtable_headingrow+2)]="Total cost of retrofit"
            char=get_column_letter(39)
            portfolio[char+str(secondtable_headingrow+2)]="Unit economics work? (total value of fees passed from the tenant over the course of the lease > cost of the projects)"
            char=get_column_letter(40)
            portfolio[char+str(secondtable_headingrow+2)]="High ROI Energy Savings (kWh)"
            char=get_column_letter(41)
            portfolio[char+str(secondtable_headingrow+2)]="High ROI Emission Savings (kg co2/m2)"
            char=get_column_letter(42)
            portfolio[char+str(secondtable_headingrow+2)]="High ROI Energy Savings (GBP)"
            char=get_column_letter(32)
            portfolio[char+str(secondtable_headingrow+3)]=f"=SUM(AP2:AP{totalproperties})"
            char=get_column_letter(33)
            portfolio[char+str(secondtable_headingrow+3)]=f"=SUM(AQ2:AQ{totalproperties})"
            char=get_column_letter(34)
            portfolio[char+str(secondtable_headingrow+3)]=f"=SUM(AR2:AR{totalproperties})"
            char=get_column_letter(35)
            portfolio[char+str(secondtable_headingrow+3)]=f"=AF{secondtable_datarow+2}-(AF{secondtable_datarow+2}*D47)-(((1-D47)*AF{secondtable_datarow+2})*D42)"
            char=get_column_letter(36)
            portfolio[char+str(secondtable_headingrow+3)]=f"=SUM(AU2:AU{totalproperties})"
            char=get_column_letter(37)
            portfolio[char+str(secondtable_headingrow+3)]=f"=SUM(AV2:AV{totalproperties})"
            char=get_column_letter(38)
            portfolio[char+str(secondtable_headingrow+3)]=f"=SUM(AW2:AW{totalproperties})"
            char=get_column_letter(39)
            portfolio[char+str(secondtable_headingrow+3)]=f'=IF(AK{secondtable_datarow+2} >= AL{secondtable_datarow+2}, "Yes", "No")'
            char=get_column_letter(40)
            portfolio[char+str(secondtable_headingrow+3)]=f"=SUM(BF2:BF{totalproperties})"
            char=get_column_letter(41)
            portfolio[char+str(secondtable_headingrow+3)]=f"=SUM(BG2:BG{totalproperties})"
            char=get_column_letter(42)
            portfolio[char+str(secondtable_headingrow+3)]=f"=SUM(BH2:BH{totalproperties})"


            totalproperties=propertyrow+2
            # FINAL TABLE SHEET
            secondtable_headingrow=1
            secondtable_datarow=2
            #totalproperties=str(propertyrow)
            char=get_column_letter(2)
            finaltable[char+str(secondtable_headingrow)]="TOTAL FINAL SAVINGS (Area-Based)"
            char=get_column_letter(3)
            finaltable[char+str(secondtable_headingrow)]="Annual Saving Passed to Landlord"
            char=get_column_letter(4)
            finaltable[char+str(secondtable_headingrow)]="Monthly tenant Fee"
            char=get_column_letter(5)
            finaltable[char+str(secondtable_headingrow)]="Annual Payback to Landlord"
            char=get_column_letter(6)
            finaltable[char+str(secondtable_headingrow)]="Annual Savings Passed to Tenants"
            char=get_column_letter(7)
            finaltable[char+str(secondtable_headingrow)]="Value from fees passed back to the landlord for the remainder of the lease"
            char=get_column_letter(8)
            finaltable[char+str(secondtable_headingrow)]="Total cost of Retrofitting"
            char=get_column_letter(9)
            finaltable[char+str(secondtable_headingrow)]="Unit economics work? (total value of fees passed from the tenant over the course of the lease > cost of the projects)"
            char=get_column_letter(10)
            finaltable[char+str(secondtable_headingrow)]="Energy Savings(kWh)"
            char=get_column_letter(11)
            finaltable[char+str(secondtable_headingrow)]="Building Emissions Savings (kg co2/m2)"
            char=get_column_letter(12)
            finaltable[char+str(secondtable_headingrow)]="Total Energy Savings (GBP Energy-Based)"
            char=get_column_letter(2)
            finaltable[char+str(secondtable_datarow)]=f"=Portfolio!AF{totalproperties}"
            char=get_column_letter(3)
            finaltable[char+str(secondtable_datarow)]=f"=Portfolio!AG{totalproperties}"
            char=get_column_letter(4)
            finaltable[char+str(secondtable_datarow)]=f"=Portfolio!AH{totalproperties}"
            char=get_column_letter(5)
            finaltable[char+str(secondtable_datarow)]=f"=Portfolio!AI{totalproperties}"
            char=get_column_letter(6)
            finaltable[char+str(secondtable_datarow)]=f"=Portfolio!AJ{totalproperties}"
            char=get_column_letter(7)
            finaltable[char+str(secondtable_datarow)]=f"=Portfolio!AK{totalproperties}"
            char=get_column_letter(8)
            finaltable[char+str(secondtable_datarow)]=f"=Portfolio!AL{totalproperties}"
            char=get_column_letter(9)
            finaltable[char+str(secondtable_datarow)]=f"=Portfolio!AM{totalproperties}"
            char=get_column_letter(10)
            finaltable[char+str(secondtable_datarow)]=f"=Portfolio!AN{totalproperties}"
            char=get_column_letter(11)
            finaltable[char+str(secondtable_datarow)]=f"=Portfolio!AO{totalproperties}"
            char=get_column_letter(12)
            finaltable[char+str(secondtable_datarow)]=f"=Portfolio!AP{totalproperties}"

            totalproperties=totalproperties+2
            #ROI TABLE
            char=get_column_letter(2)
            finaltable[char+str(secondtable_headingrow+2)]="High ROI FINAL SAVINGS"
            char=get_column_letter(3)
            finaltable[char+str(secondtable_headingrow+2)]="Annual Saving Passed to Landlord"
            char=get_column_letter(4)
            finaltable[char+str(secondtable_headingrow+2)]="Monthly tenant Fee"
            char=get_column_letter(5)
            finaltable[char+str(secondtable_headingrow+2)]="Annual Payback to Landlord"
            char=get_column_letter(6)
            finaltable[char+str(secondtable_headingrow+2)]="Annual Savings Passed to Tenants"
            char=get_column_letter(7)
            finaltable[char+str(secondtable_headingrow+2)]="Total fees passed back over lease"
            char=get_column_letter(8)
            finaltable[char+str(secondtable_headingrow+2)]="Total cost of retrofit"
            char=get_column_letter(9)
            finaltable[char+str(secondtable_headingrow+2)]="Unit economics work? (total value of fees passed from the tenant over the course of the lease > cost of the projects)"
            char=get_column_letter(10)
            finaltable[char+str(secondtable_headingrow+2)]="High ROI Energy Savings(kWh)"
            char=get_column_letter(11)
            finaltable[char+str(secondtable_headingrow+2)]="High ROI Emissions Savings(kg co2/m2)"
            char=get_column_letter(12)
            finaltable[char+str(secondtable_headingrow+2)]="High ROI Energy Savings (GBP)"

            char=get_column_letter(2)
            finaltable[char+str(secondtable_headingrow+3)]=f"=Portfolio!AF{totalproperties}"
            char=get_column_letter(3)
            finaltable[char+str(secondtable_headingrow+3)]=f"=Portfolio!AG{totalproperties}"
            char=get_column_letter(4)
            finaltable[char+str(secondtable_headingrow+3)]=f"=Portfolio!AH{totalproperties}"
            char=get_column_letter(5)
            finaltable[char+str(secondtable_headingrow+3)]=f"=Portfolio!AI{totalproperties}"
            char=get_column_letter(6)
            finaltable[char+str(secondtable_headingrow+3)]=f"=Portfolio!AJ{totalproperties}"
            char=get_column_letter(7)
            finaltable[char+str(secondtable_headingrow+3)]=f"=Portfolio!AK{totalproperties}"
            char=get_column_letter(8)
            finaltable[char+str(secondtable_headingrow+3)]=f"=Portfolio!AL{totalproperties}"
            char=get_column_letter(9)
            finaltable[char+str(secondtable_headingrow+3)]=f"=Portfolio!AM{totalproperties}"
            char=get_column_letter(10)
            finaltable[char+str(secondtable_headingrow+3)]=f"=Portfolio!AN{totalproperties}"
            char=get_column_letter(11)
            finaltable[char+str(secondtable_headingrow+3)]=f"=Portfolio!AO{totalproperties}"
            char=get_column_letter(12)
            finaltable[char+str(secondtable_headingrow+3)]=f"=Portfolio!AP{totalproperties}"



            filename='uvalue_'+query_params["postcode"]+'.xlsx'


            #wb.save(filename)
            response = HttpResponse(save_virtual_workbook(wb), content_type='application/vnd.ms-excel')
            att='attachment; filename='+filename
            response['Content-Disposition'] = att
            return response
        else:
            return JsonResponse({'invalidpostcode':"Please Enter a Valid Postcode"}, status=200)
